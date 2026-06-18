"""
Fase 9 · QA fixes - bugs encontrados en simulación final.

1. Fix Recipes: perfil_narrativo apuntaba a "Streaming USA" que no existe.
   Mapear a los 4 perfiles reales: Tech / AdTech, CPG / Retail,
   Media / Streaming, Consultoría.

2. Fix 30 logros (IDs 19-48): sin Tags Industria, Tags Habilidad, ni
   Incluir CV EN. Eran los logros heredados de Phase 1-3 que nunca
   recibieron metadata completa.

3. (Fix separado en export_master_json.py): parser de 01 Perfil leía solo
   2 columnas en lugar de 3.
"""
from __future__ import annotations
import shutil
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "data" / "master" / "cv_master_raul_mercado.xlsx"
BACKUP = ROOT / "data" / "master" / "backups" / "v9_pre_qa_fixes.xlsx"


# ═══ 1. RECIPES: mapeo correcto a perfiles narrativos reales ════════════
RECIPE_PERFIL_FIX = {
    "R01": "Media / Streaming",   # Director Research US Streaming/Media
    "R02": "Media / Streaming",   # Research Director LATAM Media Group
    "R03": "CPG / Retail",         # Sr Insights Manager CPG Global
    "R04": "Tech / AdTech",        # Head Research Tech/AdTech Startup
    "R05": "Consultoría",          # Sr Strategic Consultant
    "R06": "CPG / Retail",         # Account Manager Retail Analytics
}


# ═══ 2. ENRIQUECIMIENTO de logros 19-48 (tags + incluir EN) ═════════════
# logro_id → (tags_industria, tags_habilidad)
# Todos los Incluir CV EN se setean a "Sí" porque todos tienen texto EN.
LOGRO_TAGS = {
    # Independiente Consultor 2014 (ID Empresa 9)
    19: ("Consulting, B2B", "Consulting, Strategy, Account Management"),
    20: ("Consulting, B2B", "Strategy, Customer Insights"),
    21: ("Consulting, B2B", "Project Management, Leadership, Cross-functional"),

    # TIBA IT Services / KIO (ID Empresa 10, 11)
    22: ("Tech, B2B, Telecom", "Consultative Sales, Competitive Intelligence"),
    23: ("Tech, B2B, Telecom", "Consultative Sales, Strategy"),
    24: ("Tech, B2B, Telecom", "Consultative Sales, Client Management, Storytelling"),
    25: ("Tech, B2B, Telecom", "Strategy, Competitive Intelligence, Account Growth"),
    26: ("Tech, B2B, Telecom", "Strategy, Competitive Intelligence, Analysis"),
    27: ("Tech, B2B, Telecom", "Leadership, Strategy, Competitive Intelligence"),
    28: ("Tech, B2B, Telecom", "Strategy, Storytelling, Leadership"),
    29: ("Tech, B2B, Telecom", "Strategy, Content Strategy, Leadership"),
    30: ("Tech, B2B, Telecom", "Storytelling, Content Strategy"),
    31: ("Tech, B2B, Telecom", "Strategy, Content Strategy, Storytelling, Leadership"),

    # Nodo Research (ID Empresa 8) — UX Research pioneer
    32: ("Research-Industry, Tech, Digital", "UX Research, Methodology, Research, Leadership"),
    33: ("Research-Industry, Tech, Consulting", "Research, Storytelling, Leadership"),
    34: ("Research-Industry, Global, Consulting", "Research, Multi-Country, Leadership, Methodology"),
    35: ("Research-Industry, Tech, Digital", "UX Research, Research, Methodology"),

    # Xabón'art (ID Empresa 14) — microempresa propia
    36: ("Startup, Retail", "Creative Production, Strategy, Visual Design"),
    37: ("Startup, Retail", "Creative Production, Visual Design"),
    38: ("Startup, Retail", "Creative Production, Project Management"),

    # Focus Vision / Channel M3 (ID Empresa 15)
    39: ("Tech, Digital, Research-Industry", "Process Design"),
    40: ("Tech, Digital, Research-Industry", "Process Design, Project Management"),

    # PTA/GBA productoras (ID Empresa 16)
    41: ("Media, Arts, Entertainment", "Creative Production"),
    42: ("Media, Arts, Entertainment", "Creative Production, Visual Design"),

    # Onírica (ID Empresa 17) — emprendimiento creativo
    43: ("Media, Digital, Startup", "Creative Production, Project Management"),
    44: ("Media, Digital, Startup", "Creative Production, Project Management"),
    45: ("Media, Digital, Startup", "Creative Production, Project Management"),

    # Central de Registradoras (ID Empresa 18) — early hardware
    46: ("Tech, Retail", "Process Design, Project Management"),
    47: ("Tech, Retail", "Process Design"),
    48: ("Tech, Retail", "Process Design"),
}


def main():
    print(f"📂 Master: {MASTER}")
    if not BACKUP.exists():
        shutil.copy2(MASTER, BACKUP)
        print(f"🛟 Backup: {BACKUP.name}")

    wb = load_workbook(str(MASTER))

    # ── 1. Fix Recipes ──
    print(f"\n[1] Fixing Recipes perfil_narrativo references…")
    ws = wb["12 Recipes"]
    headers = [c.value for c in ws[2]]
    idx_rid = headers.index("Recipe ID") + 1
    idx_perfil = headers.index("Perfil Narrativo (idx)") + 1
    fixes = 0
    for r in range(3, ws.max_row + 1):
        rid = ws.cell(row=r, column=idx_rid).value
        if rid in RECIPE_PERFIL_FIX:
            old = ws.cell(row=r, column=idx_perfil).value
            new = RECIPE_PERFIL_FIX[rid]
            ws.cell(row=r, column=idx_perfil, value=new)
            print(f"  {rid}: '{old}' → '{new}'")
            fixes += 1
    print(f"  ✅ {fixes} recipes corregidas")

    # ── 2. Enrich logros 19-48 with tags + Incluir CV EN ──
    print(f"\n[2] Enriching logros 19-48 with tags + Incluir CV EN…")
    ws = wb["03 Logros"]
    headers = [c.value for c in ws[2]]
    idx_id = headers.index("ID Logro") + 1
    idx_ind = headers.index("Tags Industria") + 1
    idx_skl = headers.index("Tags Habilidad") + 1
    idx_inc_en = headers.index("Incluir CV EN") + 1

    n_tagged = 0
    n_flagged = 0
    for r in range(3, ws.max_row + 1):
        lid = ws.cell(row=r, column=idx_id).value
        if lid is None:
            continue
        try:
            lid_int = int(lid)
        except (ValueError, TypeError):
            continue
        if lid_int in LOGRO_TAGS:
            ind, skl = LOGRO_TAGS[lid_int]
            if not ws.cell(row=r, column=idx_ind).value:
                ws.cell(row=r, column=idx_ind, value=ind)
                n_tagged += 1
            if not ws.cell(row=r, column=idx_skl).value:
                ws.cell(row=r, column=idx_skl, value=skl)
        # Set Incluir CV EN = "Sí" si el logro tiene texto EN y el flag está vacío
        idx_en = headers.index("Achievement EN") + 1
        en_text = ws.cell(row=r, column=idx_en).value
        if en_text and not ws.cell(row=r, column=idx_inc_en).value:
            ws.cell(row=r, column=idx_inc_en, value="Sí")
            n_flagged += 1

    print(f"  ✅ {n_tagged} logros con tags agregados, {n_flagged} con Incluir CV EN='Sí'")

    wb.save(str(MASTER))
    print(f"\n✅ Master guardado")


if __name__ == "__main__":
    main()
