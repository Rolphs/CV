"""
Fase 10 · Agregar cursos faltantes a sheet 05 Educación.

Cursos confirmados por el usuario:
  1. La mirada del cinematógrafo (1998) · Maestro Joaquín Rubio · UIA
  2. Vanguardias Cinematográficas (2001) · Maestro Joaquín Rubio · UIA
  3. Preparación como terapeuta — Constelaciones familiares + eneagrama
     (2003-2007) · Bert Hellinger, Ingala Robl, Lawrence (apellido pendiente)

Bonus: la cátedra Anáhuac ya existe en sheet 10 (verificado: 2013,
'Introducción al Análisis Cualitativo', Maestría).
"""
from __future__ import annotations
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "data" / "master" / "cv_master_raul_mercado.xlsx"
BACKUP = ROOT / "data" / "master" / "backups" / "v10_pre_courses.xlsx"


# Cada tupla: (tipo, institucion, titulo_es, degree_en, area, pais,
#              ano_inicio, ano_fin, estado, relevancia)
NEW_COURSES = [
    ("Diplomado",
     "Universidad Iberoamericana, CDMX",
     "Diplomado en Cine: La mirada del cinematógrafo",
     "Cinema Diploma: The Cinematographer's Gaze",
     "Cine / Comunicación Visual",
     "México",
     1998, 1998, "Completo",
     "Media — formación visual y narrativa con el Maestro Joaquín Rubio. "
     "Fundamenta lectura de imagen, dirección de fotografía y análisis cinematográfico."),

    ("Diplomado",
     "Universidad Iberoamericana, CDMX",
     "Diplomado en Cine: Vanguardias Cinematográficas",
     "Cinema Diploma: Cinematographic Avant-Gardes",
     "Cine / Teoría",
     "México",
     2001, 2001, "Completo",
     "Media — continuación de la formación cinematográfica con Joaquín Rubio. "
     "Análisis de movimientos vanguardistas y teoría del lenguaje audiovisual."),

    ("Formación Continua",
     "Bert Hellinger · Ingala Robl · Lawrence",
     "Formación en Terapias Sistémicas y Eneagrama",
     "Systemic Therapy & Enneagram Training",
     "Terapia / Constelaciones Familiares",
     "México / Internacional",
     2003, 2007, "Completo",
     "Media — formación directa con Bert Hellinger (fundador de las "
     "Constelaciones Familiares) e Ingala Robl. Aporta lectura sistémica "
     "de dinámicas grupales y organizacionales aplicable a research cualitativo "
     "y liderazgo de equipos."),
]


def main():
    print(f"📂 Master: {MASTER}")
    if not BACKUP.exists():
        shutil.copy2(MASTER, BACKUP)
        print(f"🛟 Backup: {BACKUP.name}")

    wb = load_workbook(str(MASTER))
    ws = wb["05 Educación"]

    headers = [c.value for c in ws[2]]
    print(f"\nHeaders detectados: {headers}")
    print(f"Filas actuales: {ws.max_row - 2} (rows 3 a {ws.max_row})")

    # Append rows
    print(f"\n[+] Agregando {len(NEW_COURSES)} cursos:")
    for course in NEW_COURSES:
        ws.append(course)
        print(f"    • {course[6]}-{course[7]} · {course[1]} · {course[2]}")

    # Apply same styling as existing rows (banded + wrap)
    band_fill = PatternFill("solid", fgColor="F5F5F5")
    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    new_start = ws.max_row - len(NEW_COURSES) + 1
    for r in range(new_start, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = wrap
            if r % 2 == 0:
                cell.fill = band_fill

    wb.save(str(MASTER))
    print(f"\n✅ Master guardado. Educación: {ws.max_row - 2} filas totales.")


if __name__ == "__main__":
    main()
