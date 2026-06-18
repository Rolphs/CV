"""
Fase 3 v2 · Cleaner candidate extraction.

Improvements over v1:
  - Reconstruct multi-line bullets from PDF wrap (merge lines until period/end).
  - Skip skill/tool lists (lines with "Word:" or "Word, Word, Word" patterns).
  - Skip CV-header-like lines (tabs, role · company \t date).
  - Skip the current 2026 CV from candidate generation (canonical, already represents truth).
  - Lower thresholds: 0.25 = VARIANT, 0.45 = DUPLICATE.
  - Quality score per candidate: metric (+3), action verb (+2), length (+1).
  - Cap top 15 per empresa+lang for review.
"""

from __future__ import annotations
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
CATALOG = ROOT / "data" / "reports" / "file_catalog.json"
MASTER = ROOT / "data" / "master" / "cv_master_raul_mercado.xlsx"
OUT_DIR = ROOT / "data" / "reports"

EMPRESAS = [
    (1,  "Canela Media",                    [r"Canela\s+Media"]),
    (2,  "Yoor Company",                    [r"\bYoor(?:co)?\b"]),
    (3,  "Altazor Intelligence",            [r"\bAltazor\b"]),
    (4,  "Televisa",                        [r"\bTelevisa\b"]),
    (5,  "Kantar (Millward Brown)",         [r"\bKantar\b", r"Millward\s+Brown"]),
    (6,  "Independiente (BRICKO / EIDOS)",  [r"\bBRICKO\b", r"\bEIDOS\b", r"\bBricko\b"]),
    (7,  "KIO Networks",                    [r"KIO\s+Networks"]),
    (8,  "Nodo Research",                   [r"Nodo\s+Research"]),
    (10, "TIBA IT Services",                [r"\bTIBA\b"]),
    (13, "NODO Investigación + Estrategia", [r"NODO\s+Investigaci"]),
    (14, "Xabón'art",                       [r"Xab[óo]n['’`]?\s*art"]),
    (15, "Focus Vision / Channel M3",       [r"Focus\s*Vision", r"Channel\s*M3"]),
    (17, "Onírica",                         [r"\bOn[íi]rica\b"]),
]

# Files to SKIP (canonical 2026 = truth; non-empty book = pending OCR)
SKIP_FILES = {
    "cv_raul_mercado_2026.pdf",
    "cv_raul_mercado_2026.docx",
    "_PENDING_OCR_intro_book_art_director_49p.pdf",
    "Curriculum (1).pdf",  # intruso (already moved but text still in catalog json)
}


# ─── text utilities ───────────────────────────────────────────────────────────

ACTION_VERBS_ES = {
    "lideré", "lideró", "liderar", "diseñé", "diseñó", "implementé", "implementó",
    "incrementé", "incrementó", "reduje", "redujo", "aumenté", "aumentó",
    "desarrollé", "desarrolló", "automaticé", "automatizó", "gestioné", "gestionó",
    "construí", "construyó", "logré", "logró", "presenté", "presentó",
    "coordiné", "coordinó", "ejecuté", "ejecutó", "establecí", "estableció",
    "dupliqué", "duplicó", "generé", "generó", "transformé", "transformó",
    "creé", "creó", "rediseñé", "rediseñó", "optimicé", "optimizó",
    "facilité", "facilitó", "supervisé", "supervisó", "moderé", "moderó",
}
ACTION_VERBS_EN = {
    "led", "designed", "implemented", "increased", "reduced", "developed",
    "automated", "managed", "built", "achieved", "presented", "coordinated",
    "executed", "established", "doubled", "generated", "transformed", "created",
    "redesigned", "optimized", "facilitated", "supervised", "moderated",
    "delivered", "drove", "spearheaded", "launched", "founded", "scaled",
    "streamlined", "negotiated", "secured", "expanded", "pioneered",
}
METRIC_RE = re.compile(r"\d+\s*(?:%|x|m\b|k\b|mm|\+|años?|years?)|\$\s*[\d.,]+|[+\-]\s*\d+", re.I)

SKILL_LIST_RE = re.compile(
    r"^\s*[A-Z][\w/&\s\-]{2,30}:\s+",  # "Tools: ...", "Frameworks: ..."
)
PIPE_TAB_RE = re.compile(r"\t+|\s\|\s")  # CV header rows usually have tabs / pipes


def normalize_text(text: str) -> str:
    text = unicodedata.normalize("NFKD", text)
    return "".join(c for c in text if not unicodedata.combining(c)).lower()


def tokens(text: str) -> set[str]:
    norm = normalize_text(text)
    toks = re.findall(r"[a-z0-9]+", norm)
    STOP = {
        "de", "la", "el", "en", "y", "a", "del", "los", "las", "para", "con",
        "por", "un", "una", "the", "of", "and", "to", "in", "for", "on", "at",
        "from", "with", "by", "as", "an", "or", "que", "se", "su", "sus",
        "es", "lo", "le", "al", "fue", "ha", "han", "muy", "mas", "como",
    }
    return {t for t in toks if len(t) >= 3 and t not in STOP}


def jaccard(a: set, b: set) -> float:
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


def quality_score(text: str, lang: str) -> int:
    score = 0
    if METRIC_RE.search(text):
        score += 3
    verbs = ACTION_VERBS_EN if lang == "EN" else ACTION_VERBS_ES
    if any(v in normalize_text(text) for v in verbs):
        score += 2
    if 40 <= len(text) <= 250:
        score += 1
    return score


# ─── bullet extraction with line-merge ────────────────────────────────────────

BULLET_PREFIX_RE = re.compile(r"^\s*[•●▪■◾▶►‣\-\*–—·]+\s+")
NUMBER_PREFIX_RE = re.compile(r"^\s*\d+[\.\)]\s+")


def merge_wrapped_lines(text: str) -> list[str]:
    """Reconstruct sentences split by PDF line-wrap.

    Heuristic: if a line doesn't end in . ! ? : and next line starts with lowercase,
    they belong together.
    """
    raw = [ln.strip() for ln in text.splitlines() if ln.strip()]
    merged: list[str] = []
    buf = ""
    for line in raw:
        if not buf:
            buf = line
            continue
        prev_ends_sentence = buf[-1] in ".!?:;"
        starts_lower = line[0].islower() if line else False
        starts_bullet = bool(BULLET_PREFIX_RE.match(line) or NUMBER_PREFIX_RE.match(line))
        if not prev_ends_sentence and starts_lower and not starts_bullet:
            buf += " " + line
        else:
            merged.append(buf)
            buf = line
    if buf:
        merged.append(buf)
    return merged


def extract_bullets(segment: str) -> list[str]:
    merged = merge_wrapped_lines(segment)
    bullets = []
    for line in merged:
        clean = BULLET_PREFIX_RE.sub("", line)
        clean = NUMBER_PREFIX_RE.sub("", clean)
        clean = clean.strip()

        # filters
        if len(clean) < 35 or len(clean) > 400:
            continue
        if clean.isupper():
            continue
        if PIPE_TAB_RE.search(clean):
            continue
        if SKILL_LIST_RE.match(clean):
            continue
        if sum(1 for c in clean if c.islower()) < 8:
            continue
        if re.match(r"^[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*\s*[·•|]", clean):
            # "Director of X · Company · Date" looking header
            continue
        # must contain at least one verb (filter out noun phrases / lists)
        norm = normalize_text(clean)
        if not any(v in norm for v in (ACTION_VERBS_ES | ACTION_VERBS_EN)):
            # Allow if it has a metric (still a strong signal even without verb)
            if not METRIC_RE.search(clean):
                continue

        bullets.append(clean)
    return bullets


# ─── language detection ───────────────────────────────────────────────────────

EN_HINT = re.compile(r"\b(the|and|of|with|through|achieved|delivered|increased|reduced|led|driven|management|strategy|client|research|across|via)\b", re.I)
ES_HINT = re.compile(r"\b(de|la|el|en|con|para|por|mediante|investigaci[óo]n|estrategia|clientes|gesti[óo]n|aument[óoé]|reduj[óoe]|lider[éóa])\b", re.I)


def detect_lang_bullet(text: str) -> str:
    """Per-bullet language. Counts matching hint words."""
    en = len(EN_HINT.findall(text))
    es = len(ES_HINT.findall(text))
    if en == es == 0:
        return "ES"  # default
    return "EN" if en > es else "ES"


def detect_lang(text: str) -> str:
    en_markers = ["Professional Summary", "PROFILE SUMMARY", "EXECUTIVE SUMMARY",
                  "Achievements", "WORK EXPERIENCE", "EDUCATION"]
    es_markers = ["PERFIL", "RESUMEN", "EXPERIENCIA", "EDUCACIÓN",
                  "FORMACIÓN", "Logros", "OBJETIVOS"]
    en = sum(1 for m in en_markers if m in text)
    es = sum(1 for m in es_markers if m in text)
    return "EN" if en > es else "ES"


# ─── per-empresa segmentation ─────────────────────────────────────────────────

def segment_by_empresa(text: str) -> dict[int, str]:
    hits = []
    for emp_id, _name, aliases in EMPRESAS:
        for alias in aliases:
            for m in re.finditer(alias, text):
                hits.append((m.start(), emp_id))
    if not hits:
        return {}
    hits.sort()
    segments: dict[int, str] = defaultdict(str)
    for i, (pos, emp_id) in enumerate(hits):
        end = hits[i + 1][0] if i + 1 < len(hits) else min(pos + 1200, len(text))
        segments[emp_id] += "\n" + text[pos:end]
    return dict(segments)


# ─── master logros loader ─────────────────────────────────────────────────────

def load_master_logros() -> dict[int, list[dict]]:
    wb = load_workbook(str(MASTER), data_only=True)
    ws = wb["03 Logros"]
    by_emp: dict[int, list[dict]] = defaultdict(list)
    for row in ws.iter_rows(min_row=3, max_col=14, values_only=True):
        if row[0] is None:
            continue
        try:
            emp_id = int(row[1]) if row[1] is not None else None
        except (TypeError, ValueError):
            emp_id = None
        if emp_id is None:
            continue
        by_emp[emp_id].append({
            "id": int(row[0]),
            "tipo": row[3] or "",
            "es": row[4] or "",
            "en": row[5] or "",
        })
    wb.close()
    return dict(by_emp)


# ─── main pipeline ────────────────────────────────────────────────────────────

def main():
    catalog = json.loads(CATALOG.read_text(encoding="utf-8"))
    cv_records = [
        r for r in catalog
        if r.get("extract", {}).get("full_text")
        and r["ext"] in (".pdf", ".docx")
        and r["status"] == "ok"
        and r["name"] not in SKIP_FILES
    ]
    print(f"Processing {len(cv_records)} CVs (skipped {len(SKIP_FILES)} canonical/intruso)")

    master_logros = load_master_logros()

    candidates: dict[tuple[int, str], list[dict]] = defaultdict(list)

    for rec in cv_records:
        text = rec["extract"]["full_text"]
        for emp_id, segment in segment_by_empresa(text).items():
            for bullet in extract_bullets(segment):
                lang = detect_lang_bullet(bullet)
                candidates[(emp_id, lang)].append({
                    "text": bullet,
                    "source": rec["name"],
                    "quality": quality_score(bullet, lang),
                })

    # cluster-dedupe within (emp, lang)
    def signature(s: str) -> frozenset:
        # use top-5 longest tokens as cluster key
        ts = sorted(tokens(s), key=len, reverse=True)[:5]
        return frozenset(ts)

    deduped: dict[tuple[int, str], list[dict]] = {}
    for key, items in candidates.items():
        clusters: dict[frozenset, dict] = {}
        for it in sorted(items, key=lambda x: -x["quality"]):
            sig = signature(it["text"])
            # check if cluster overlaps with existing (>=3 shared tokens)
            merged = False
            for ckey in clusters:
                if len(sig & ckey) >= 3:
                    clusters[ckey]["sources"].append(it["source"])
                    merged = True
                    break
            if not merged:
                clusters[sig] = {
                    "text": it["text"],
                    "sources": [it["source"]],
                    "quality": it["quality"],
                }
        deduped[key] = sorted(clusters.values(), key=lambda x: -x["quality"])

    # classify ES vs master, keep top 15 per empresa+lang
    TOP_N = 15
    classified: dict[int, dict] = defaultdict(lambda: {
        "ES_NEW": [], "ES_VARIANT": [], "ES_DUPLICATE": [], "EN_FOUND": []
    })
    for (emp_id, lang), items in deduped.items():
        existing = master_logros.get(emp_id, [])
        existing_tokens = [(e, tokens(e["es"])) for e in existing]
        for it in items[:TOP_N]:
            ct = tokens(it["text"])
            best = (None, 0.0)
            for ex_row, ex_t in existing_tokens:
                s = jaccard(ct, ex_t)
                if s > best[1]:
                    best = (ex_row, s)
            it["best_match_id"] = best[0]["id"] if best[0] else None
            it["best_match_es"] = best[0]["es"][:140] if best[0] else ""
            it["best_score"] = round(best[1], 2)
            if lang == "EN":
                classified[emp_id]["EN_FOUND"].append(it)
            else:
                if best[1] >= 0.45:
                    classified[emp_id]["ES_DUPLICATE"].append(it)
                elif best[1] >= 0.25:
                    classified[emp_id]["ES_VARIANT"].append(it)
                else:
                    classified[emp_id]["ES_NEW"].append(it)

    # ─── outputs ──────────────────────────────────────────────────────────────
    out_json = OUT_DIR / "logros_phase3_candidates.json"
    out_json.write_text(json.dumps(
        {str(k): v for k, v in classified.items()},
        indent=2, ensure_ascii=False,
    ), encoding="utf-8")

    emp_names = {eid: name for eid, name, _ in EMPRESAS}
    out_md = OUT_DIR / "logros_phase3_candidates.md"
    with open(out_md, "w", encoding="utf-8") as f:
        f.write("# 🎯 Fase 3 · Candidatos de logros desde CVs (v2)\n\n")
        new = sum(len(c["ES_NEW"]) for c in classified.values())
        var = sum(len(c["ES_VARIANT"]) for c in classified.values())
        dup = sum(len(c["ES_DUPLICATE"]) for c in classified.values())
        en = sum(len(c["EN_FOUND"]) for c in classified.values())
        f.write(f"**Totales:** {new} NUEVOS ES · {var} VARIANTES ES · {dup} duplicados · {en} bullets EN\n\n")
        f.write("**Quality score:** métrica numérica (+3), verbo de acción (+2), longitud ideal (+1)\n\n")
        f.write("**Cómo usar este reporte:**\n")
        f.write("- ✨ **NUEVOS ES**: logros no cubiertos en el maestro — candidatos a agregar.\n")
        f.write("- 🔁 **VARIANTES ES**: misma idea con frase distinta — útil para enriquecer/reemplazar el original.\n")
        f.write("- 🌐 **EN**: bullets en inglés — fuente para llenar `Achievement EN` de logros existentes.\n\n")

        for emp_id in sorted(classified.keys()):
            name = emp_names.get(emp_id, f"#{emp_id}")
            c = classified[emp_id]
            existing = master_logros.get(emp_id, [])
            f.write(f"\n## 🏢 #{emp_id} · {name}\n")
            f.write(f"_{len(existing)} logros/resp. en maestro_\n\n")

            if c["ES_NEW"]:
                f.write(f"### ✨ Nuevos ES ({len(c['ES_NEW'])})\n\n")
                for item in c["ES_NEW"]:
                    sources = ", ".join(sorted(set(item["sources"]))[:3])
                    f.write(f"- `Q{item['quality']}` **{item['text']}**\n")
                    f.write(f"  - _fuentes: {sources}_\n")
                f.write("\n")
            if c["ES_VARIANT"]:
                f.write(f"### 🔁 Variantes ES ({len(c['ES_VARIANT'])})\n\n")
                for item in c["ES_VARIANT"]:
                    sources = ", ".join(sorted(set(item["sources"]))[:2])
                    f.write(f"- `Q{item['quality']}` **{item['text']}**\n")
                    f.write(f"  - matches #{item['best_match_id']} (score {item['best_score']}): _{item['best_match_es']}_\n")
                    f.write(f"  - _fuentes: {sources}_\n")
                f.write("\n")
            if c["EN_FOUND"]:
                f.write(f"### 🌐 EN bullets ({len(c['EN_FOUND'])})\n\n")
                for item in c["EN_FOUND"]:
                    sources = ", ".join(sorted(set(item["sources"]))[:2])
                    f.write(f"- `Q{item['quality']}` **{item['text']}**\n")
                    if item["best_match_id"]:
                        f.write(f"  - likely → logro #{item['best_match_id']}: _{item['best_match_es']}_\n")
                    f.write(f"  - _fuentes: {sources}_\n")
                f.write("\n")
            if c["ES_DUPLICATE"]:
                f.write(f"<details><summary>📋 Duplicados ES ({len(c['ES_DUPLICATE'])})</summary>\n\n")
                for item in c["ES_DUPLICATE"]:
                    f.write(f"- {item['text']} _(matches #{item['best_match_id']}, score {item['best_score']})_\n")
                f.write("\n</details>\n\n")

    print(f"✅ {out_md.name}")
    print(f"✅ {out_json.name}")
    print(f"\nTotales: NUEVOS ES={new} · VARIANTES ES={var} · DUP ES={dup} · EN={en}")


if __name__ == "__main__":
    main()
