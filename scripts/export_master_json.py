"""
Export master xlsx → JSON nativo para consumo de LLM.

Genera data/master/cv_master.json con estructura limpia, IDs cruzadas
resueltas, y un schema que cualquier LLM puede ingerir directo.

Uso:
    python scripts/export_master_json.py

Re-ejecutable: corre cada vez que actualices el xlsx para mantener
sincronizado el JSON.
"""
from __future__ import annotations
import json
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "data" / "master" / "cv_master_raul_mercado.xlsx"
OUT = ROOT / "data" / "master" / "cv_master.json"


def _cell(v):
    """Normalize a cell value for JSON."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    s = str(v).strip()
    return s if s else None


def _bool(v):
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ("sí", "si", "yes", "true", "1"):
        return True
    if s in ("no", "false", "0"):
        return False
    return None


def _list(v):
    if not v:
        return []
    return [t.strip() for t in str(v).replace(";", ",").split(",") if t.strip()]


def read_sheet_as_dicts(ws):
    """Read sheet starting row 3 (row 2 = headers, row 1 = title)."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return [], []
    headers = [str(c).strip() if c is not None else f"col{i}"
               for i, c in enumerate(rows[1])]
    data = []
    for r in rows[2:]:
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in r):
            continue
        d = {h: _cell(r[i]) if i < len(r) else None for i, h in enumerate(headers)}
        data.append(d)
    return headers, data


def read_meta_sheet(ws):
    """Read key/value sheets (00, 01) supporting layouts:
        - 2-col: [SECTION] o [key, value]
        - 3-col: [SECTION] o [None, key, value]
    Returns list of {section, items: [{key, value} | {text}]}.
    """
    out = []
    for row in ws.iter_rows(values_only=True):
        cells = [_cell(c) for c in row[:4]]
        # Pad to 4
        while len(cells) < 4:
            cells.append(None)
        a, b, c, d = cells

        # Skip empty rows
        if not any(cells):
            continue

        # Section header: col A populated + uppercase + nothing in B/C
        if a and not b and not c and a == a.upper() and len(a) < 80:
            out.append({"section": a, "items": []})
            continue

        # Ensure we have a section bucket
        if not out:
            out.append({"section": "INTRO", "items": []})

        # Layout detection:
        # - [None, key, value] → 3-col key/value
        # - [key, value, None] → 2-col key/value
        # - [text, None, None] → free text block
        if a is None and b and c:
            out[-1]["items"].append({"key": b, "value": c})
        elif a and b and not c:
            out[-1]["items"].append({"key": a, "value": b})
        elif a and not b and not c:
            out[-1]["items"].append({"text": a})
        elif a and b and c:
            # 3-col with all populated: treat as key=A, value=B+C
            out[-1]["items"].append({"key": a, "value": f"{b} {c}"})
        elif b and not c and not a:
            out[-1]["items"].append({"text": b})
    return out


def build_json():
    wb = load_workbook(str(MASTER), data_only=True)
    snapshot = {
        "_meta": {
            "generated_at": datetime.now().isoformat(),
            "source": MASTER.name,
            "schema_version": "1.0",
            "owner": "Raúl Mercado Bustamante",
            "purpose": (
                "Master DB for LLM-driven CV generation. "
                "See sheet '00 Instrucciones' for operating manual. "
                "Never invent data, always cite IDs."
            ),
        }
    }

    # 00 Instrucciones — flat structured
    snapshot["instrucciones"] = read_meta_sheet(wb["00 Instrucciones"])

    # 01 Perfil — flat key/value
    perfil_raw = read_meta_sheet(wb["01 Perfil"])
    perfil_flat = {}
    for block in perfil_raw:
        section_data = {}
        for item in block.get("items", []):
            if "key" in item:
                section_data[item["key"]] = item["value"]
        if section_data:
            perfil_flat[block["section"]] = section_data
    snapshot["perfil"] = perfil_flat

    # 02 Experiencia
    _, exp_rows = read_sheet_as_dicts(wb["02 Experiencia"])
    experiencia = []
    for r in exp_rows:
        try:
            eid = int(r["ID"]) if r.get("ID") else None
        except (ValueError, TypeError):
            eid = r.get("ID")
        experiencia.append({
            "id": eid,
            "empresa": r.get("Empresa / Company") or r.get("Empresa"),
            "industria": r.get("Industria / Industry") or r.get("Industria"),
            "tipo": r.get("Tipo"),
            "puesto_es": r.get("Puesto ES"),
            "puesto_en": r.get("Puesto EN"),
            "descripcion_empresa_es": r.get("Descripción Empresa ES"),
            "descripcion_empresa_en": r.get("Descripción Empresa EN"),
            "fecha_inicio": r.get("Fecha Inicio"),
            "fecha_fin": r.get("Fecha Fin"),
            "actual": _bool(r.get("¿Actual?")),
            "pais": r.get("País"),
            "personas_a_cargo": r.get("Personas a cargo"),
            "reporta_a": r.get("Reporta a"),
            "departamento": r.get("Dept / Área"),
        })
    snapshot["experiencia"] = experiencia

    # Build company lookup
    company_by_id = {e["id"]: e["empresa"] for e in experiencia if e["id"] is not None}

    # 03 Logros
    _, log_rows = read_sheet_as_dicts(wb["03 Logros"])
    logros = []
    for r in log_rows:
        try:
            lid = int(r["ID Logro"]) if r.get("ID Logro") else None
        except (ValueError, TypeError):
            lid = r.get("ID Logro")
        try:
            eid = int(r["ID Empresa"]) if r.get("ID Empresa") else None
        except (ValueError, TypeError):
            eid = r.get("ID Empresa")
        logros.append({
            "id": lid,
            "id_empresa": eid,
            "empresa": r.get("Empresa") or company_by_id.get(eid),
            "tipo": r.get("Tipo"),
            "es": r.get("Logro / Resp. ES"),
            "en": r.get("Achievement EN"),
            "impacto": r.get("Impacto"),
            "tiene_metrica": _bool(r.get("Tiene Métrica")),
            "metrica_kpi": r.get("Métrica / KPI"),
            "categoria": r.get("Categoría"),
            "tags_industria": _list(r.get("Tags Industria")),
            "tags_skill": _list(r.get("Tags Habilidad")),
            "seniority_fit": _list(r.get("Seniority Fit")),
            "story_evidence": r.get("Story / Evidence"),
            "pair_with": [int(x) for x in _list(r.get("Pair With (IDs)"))
                          if x.isdigit()],
            "incluir_cv_es": _bool(r.get("Incluir CV ES")),
            "incluir_cv_en": _bool(r.get("Incluir CV EN")),
        })
    snapshot["logros"] = logros

    # 04 Skills
    _, skill_rows = read_sheet_as_dicts(wb["04 Skills"])
    snapshot["skills"] = [
        {
            "categoria": r.get("Categoría"),
            "es": r.get("Habilidad ES"),
            "en": r.get("Skill EN"),
            "nivel": r.get("Nivel"),
            "anos_exp": r.get("Años exp."),
            "descripcion_es": r.get("Descripción ES"),
            "descripcion_en": r.get("Description EN"),
            "destacar_en": _list(r.get("Destacar en")),
        } for r in skill_rows
    ]

    # 05 Educación
    _, edu_rows = read_sheet_as_dicts(wb["05 Educación"])
    snapshot["educacion"] = [
        {
            "institucion": r.get("Institución"),
            "titulo_es": r.get("Título / Grado ES"),
            "titulo_en": r.get("Degree EN"),
            "area": r.get("Área"),
            "anos": f"{r.get('Año inicio') or ''}-{r.get('Año fin') or ''}".strip("-"),
            "pais": r.get("País"),
            "relevancia": r.get("Relevancia para CV") or r.get("Relevancia"),
            "incluir_cv_es": _bool(r.get("Incluir CV ES")),
            "incluir_cv_en": _bool(r.get("Incluir CV EN")),
        } for r in edu_rows
    ]

    # 06 Certificaciones
    _, cert_rows = read_sheet_as_dicts(wb["06 Certificaciones"])
    snapshot["certificaciones"] = [
        {
            "es": r.get("Certificado ES"),
            "en": r.get("Certificate EN"),
            "institucion": r.get("Institución"),
            "ano": r.get("Año"),
            "vigencia": r.get("Vigencia"),
            "url": r.get("URL / Credencial"),
        } for r in cert_rows
    ]

    # 07 Voluntariado
    _, vol_rows = read_sheet_as_dicts(wb["07 Voluntariado"])
    snapshot["voluntariado"] = [
        {
            "organizacion": r.get("Organización"),
            "actividad_es": r.get("Actividad ES"),
            "actividad_en": r.get("Activity EN"),
            "descripcion_es": r.get("Descripción ES"),
            "descripcion_en": r.get("Description EN"),
            "anos": f"{r.get('Año inicio') or ''}-{r.get('Año fin') or ''}".strip("-"),
            "incluir_cv": _bool(r.get("Incluir CV ES")) or _bool(r.get("Incluir CV EN")),
        } for r in vol_rows
    ]

    # 08 Empresas Target (NO usar para CV — solo tracking del usuario)
    _, tgt_rows = read_sheet_as_dicts(wb["08 Empresas Target"])
    snapshot["empresas_target"] = [
        {k: v for k, v in r.items() if v is not None} for r in tgt_rows
    ]

    # 09 Perfiles Narrativos
    _, narr_rows = read_sheet_as_dicts(wb["09 Perfiles Narrativos"])
    snapshot["perfiles_narrativos"] = [
        {
            "tipo_empresa": r.get("Tipo de empresa"),
            "audiencia": r.get("Audiencia"),
            "perfil_es": r.get("Perfil ES (≈120 palabras)"),
            "perfil_en": r.get("Profile EN (≈120 words)"),
            "skills_a_enfatizar": _list(r.get("Skills a enfatizar")),
        } for r in narr_rows
    ]

    # 10 Conferencias y Speaker
    _, conf_rows = read_sheet_as_dicts(wb["10 Conferencias y Speaker"])
    snapshot["conferencias"] = [
        {
            "ano": r.get("Año"),
            "tipo": r.get("Tipo"),
            "titulo_es": r.get("Título ES"),
            "titulo_en": r.get("Title EN"),
            "evento_institucion": r.get("Evento / Institución") or r.get("Institución"),
            "pais": r.get("País"),
            "url": r.get("URL / Credencial") or r.get("URL"),
            "incluir_cv_es": _bool(r.get("Incluir CV ES")),
            "incluir_cv_en": _bool(r.get("Incluir CV EN")),
        } for r in conf_rows
    ]

    # 11 Reconocimientos
    _, rec_rows = read_sheet_as_dicts(wb["11 Reconocimientos y Membresías"])
    snapshot["reconocimientos"] = [
        {
            "tipo": r.get("Tipo"),
            "nombre_es": r.get("Nombre ES"),
            "nombre_en": r.get("Name EN"),
            "institucion": r.get("Institución"),
            "ano": r.get("Año"),
            "estado": r.get("Estado"),
        } for r in rec_rows
    ]

    # 12 Recipes
    _, rec_rows = read_sheet_as_dicts(wb["12 Recipes"])
    snapshot["recipes"] = [
        {
            "id": r.get("Recipe ID"),
            "name": r.get("Nombre / Target Role"),
            "when_to_use": r.get("Cuándo usar"),
            "perfil_narrativo": r.get("Perfil Narrativo (idx)"),
            "logros": [int(x) for x in _list(r.get("Logros a usar (IDs)")) if x.isdigit()],
            "skills_categorias": _list(r.get("Skills (categorías)")),
            "notas_llm": r.get("Notas críticas para el LLM"),
        } for r in rec_rows
    ]

    # 13 Anti-Patterns
    _, ap_rows = read_sheet_as_dicts(wb["13 Anti-Patterns"])
    snapshot["anti_patterns"] = [
        {
            "id": r.get("ID"),
            "categoria": r.get("Categoría"),
            "evita": r.get("Evita esto"),
            "usa_en_lugar": r.get("En lugar usa"),
            "razon": r.get("Por qué"),
        } for r in ap_rows
    ]

    return snapshot


def main():
    print(f"📂 Loading {MASTER.name}")
    snapshot = build_json()
    OUT.write_text(json.dumps(snapshot, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"📁 → {OUT.name}")

    # Stats
    print(f"\n📊 Snapshot stats:")
    print(f"   {len(snapshot['experiencia'])} puestos")
    print(f"   {len(snapshot['logros'])} logros")
    print(f"   {len(snapshot['skills'])} skills")
    print(f"   {len(snapshot['educacion'])} educación")
    print(f"   {len(snapshot['certificaciones'])} certificaciones")
    print(f"   {len(snapshot['voluntariado'])} voluntariado")
    print(f"   {len(snapshot['perfiles_narrativos'])} perfiles narrativos")
    print(f"   {len(snapshot['conferencias'])} conferencias / speaker")
    print(f"   {len(snapshot['reconocimientos'])} reconocimientos")
    print(f"   {len(snapshot['recipes'])} recipes")
    print(f"   {len(snapshot['anti_patterns'])} anti-patterns")

    size_kb = OUT.stat().st_size / 1024
    print(f"\n   📦 Tamaño: {size_kb:.1f} KB")


if __name__ == "__main__":
    main()
