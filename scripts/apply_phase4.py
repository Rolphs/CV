"""
Fase 4 · Complete missing fields in master.

Tasks:
  1. Add Walmart Scintilla as new puesto #19 in '02 Experiencia'.
  2. Add 2-3 starter logros for Walmart Scintilla in '03 Logros'.
  3. Fill 30 missing 'Achievement EN' translations for logros 19-48.
  4. Fill executive bio ES (01 Perfil R18C3) + EN (R22C3).
"""

from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "data" / "master" / "cv_master_raul_mercado.xlsx"


# ─── Walmart Scintilla puesto ─────────────────────────────────────────────────

WALMART_PUESTO = {
    "id": 19,
    "empresa": "Walmart (Scintilla / Data Ventures)",
    "industria": "Retail / Data & Analytics",
    "pais": "México",
    "puesto_es": "Account Manager · Walmart Scintilla",
    "puesto_en": "Account Manager · Walmart Scintilla",
    "nivel": "Gerente",
    "tipo": "Full-time",
    "fecha_inicio": "08/2025",
    "fecha_fin": "Actual",
    "actual": "Sí",
    "dept": "Walmart Data Ventures · Scintilla",
    "reporta_a": "",
    "personas_a_cargo": "",
    "desc_es": "Walmart Data Ventures es la unidad de monetización de datos de Walmart. Scintilla es su plataforma de retail analytics que ofrece a marcas CPG insights transaccionales en tiempo real para decisiones de surtido, precio y promoción.",
    "desc_en": "Walmart Data Ventures is Walmart's data monetization arm. Scintilla is its retail analytics platform delivering real-time transactional insights to CPG brands for assortment, pricing, and promotion decisions.",
}

WALMART_LOGROS = [
    {
        "tipo": "Responsabilidad",
        "es": "Gestión de cuentas CPG estratégicas en la plataforma de retail analytics de Walmart, traduciendo datos transaccionales en decisiones de surtido, precio y promoción.",
        "en": "Manage strategic CPG accounts on Walmart's retail analytics platform, translating transactional data into assortment, pricing and promotion decisions.",
        "impacto": "Alto",
        "metrica": "No",
        "metrica_val": "",
        "categoria": "Account Management",
        "tags_ind": "Retail,CPG,Data",
        "tags_skill": "Account Management,Retail Analytics,Insights",
    },
    {
        "tipo": "Logro",
        "es": "Elección deliberada de cercanía a datos conductuales a escala retail por encima de seniority de título, alineando trayectoria con el siguiente salto a métodos conductuales aplicados.",
        "en": "Deliberate choice of proximity to retail-scale behavioral data over title seniority, aligning career path with applied behavioral methods.",
        "impacto": "Medio",
        "metrica": "No",
        "metrica_val": "",
        "categoria": "Estrategia de Carrera",
        "tags_ind": "Retail,Data",
        "tags_skill": "Behavioral Data,Estrategia",
    },
]


# ─── EN translations for logros 19-48 ─────────────────────────────────────────

EN_TRANSLATIONS = {
    19: "Led projects as independent consultant and strategic partner.",
    20: "Designed solutions tailored by segment.",
    21: "Led research and marketing projects; formed ad-hoc teams adapted to each client and engagement.",
    22: "Won contracts against international enterprise competitors.",
    23: "Effective customization of technology solutions.",
    24: "Translated client needs into high-value proposals for technology support contracts.",
    25: "Annual doubling of client base.",
    26: "Consolidation of a competitive portfolio based on sector analysis.",
    27: "Led endo-marketing strategy and competitive analysis; generated reports for the board and managed multidisciplinary teams.",
    28: "Consolidation of institutional positioning.",
    29: "B2B marketing strategy.",
    30: "Greater internal visibility and sense of belonging.",
    31: "Portfolio design, board-level presentation development, conceptualization of B2B marketing strategy.",
    32: "Pioneer of usability research in Mexico.",
    33: "Presentation to C-level executives.",
    34: "Leadership of international multi-country studies.",
    35: "Study moderation, qualitative analysis, results presentation, methodologies in usability and UX.",
    36: "Sustainability as own business.",
    37: "Integration of design, production and commercialization.",
    38: "Development, production and sale of artisanal health and beauty products.",
    39: "Key technical support for stable streaming of remote research sessions.",
    40: "Transmission and technical support for online research sessions.",
    41: "Participation in notable productions for renowned national clients.",
    42: "Visualization, assistant direction and art direction for advertising campaigns.",
    43: "End-to-end content production for clients.",
    44: "Development of multimedia products.",
    45: "Audiovisual production and web development.",
    46: "Early technical experience.",
    47: "Fundamentals in hardware and technical support.",
    48: "Service and repair of computers, cash registers and peripherals.",
}


# ─── Executive bio ────────────────────────────────────────────────────────────

BIO_ES = (
    "Director de investigación e insights con 20+ años de trayectoria en medios, "
    "FMCG, tecnología y consultoría. Especialista en integrar metodologías cualitativas, "
    "cuantitativas y de IA generativa para acelerar la toma de decisiones ejecutivas. "
    "Ha liderado equipos en Televisa, Canela Media, Kantar/Millward Brown y proyectos "
    "independientes con startups y marcas globales. Reconocido por traducir datos "
    "complejos en estrategia accionable y por construir o transformar áreas de "
    "investigación con automatización y arquitecturas escalables."
)

BIO_EN = (
    "Research & insights director with 20+ years across media, FMCG, technology and "
    "consulting. Specialist in blending qualitative, quantitative and generative-AI "
    "methodologies to accelerate executive decision-making. Has led research functions "
    "at Televisa, Canela Media, Kantar/Millward Brown and independent engagements with "
    "startups and global brands. Recognized for translating complex data into actionable "
    "strategy and for building or transforming research areas with automation and "
    "scalable architectures."
)


def main():
    print("📂 Loading master...")
    wb = load_workbook(str(MASTER))

    # ─── Task 1: Add Walmart Scintilla puesto ────────────────────────────────
    ws_exp = wb["02 Experiencia"]
    # Find max existing ID
    existing_ids = set()
    for row in ws_exp.iter_rows(min_row=3, max_col=1, values_only=True):
        try:
            existing_ids.add(int(row[0]))
        except (TypeError, ValueError):
            pass
    if 19 not in existing_ids:
        wp = WALMART_PUESTO
        new_row = ws_exp.max_row + 1
        values = [
            wp["id"], wp["empresa"], wp["industria"], wp["pais"],
            wp["puesto_es"], wp["puesto_en"], wp["nivel"], wp["tipo"],
            wp["fecha_inicio"], wp["fecha_fin"], wp["actual"],
            wp["dept"], wp["reporta_a"], wp["personas_a_cargo"],
            wp["desc_es"], wp["desc_en"],
        ]
        for cidx, val in enumerate(values, start=1):
            ws_exp.cell(row=new_row, column=cidx, value=val)
        print(f"✅ Added puesto #19: {wp['empresa']}")
    else:
        print(f"ℹ️  Puesto #19 already exists, skipping")

    # ─── Task 2: Add Walmart logros ──────────────────────────────────────────
    ws_log = wb["03 Logros"]
    max_logro_id = 0
    walmart_logro_ids = set()
    for row in ws_log.iter_rows(min_row=3, max_col=14, values_only=True):
        if row[0] is None:
            continue
        rid = int(row[0])
        max_logro_id = max(max_logro_id, rid)
        try:
            if int(row[1]) == 19:
                walmart_logro_ids.add(rid)
        except (TypeError, ValueError):
            pass

    if not walmart_logro_ids:
        next_id = max_logro_id + 1
        for logro in WALMART_LOGROS:
            new_row = ws_log.max_row + 1
            values = [
                next_id, 19, WALMART_PUESTO["empresa"],
                logro["tipo"], logro["es"], logro["en"],
                logro["impacto"], logro["metrica"], logro["metrica_val"],
                logro["categoria"], logro["tags_ind"], logro["tags_skill"],
                "Sí", "Sí",
            ]
            for cidx, val in enumerate(values, start=1):
                ws_log.cell(row=new_row, column=cidx, value=val)
            print(f"  + Walmart logro #{next_id}: {logro['es'][:60]}…")
            next_id += 1
    else:
        print(f"ℹ️  Walmart already has {len(walmart_logro_ids)} logros")

    # ─── Task 3: Fill EN translations for logros 19-48 ───────────────────────
    en_filled = 0
    for ridx, row in enumerate(ws_log.iter_rows(min_row=3, max_col=14, values_only=False), start=3):
        if row[0].value is None:
            continue
        try:
            rid = int(row[0].value)
        except (TypeError, ValueError):
            continue
        if rid in EN_TRANSLATIONS:
            current_en = (row[5].value or "").strip()
            if not current_en:
                ws_log.cell(row=ridx, column=6, value=EN_TRANSLATIONS[rid])
                en_filled += 1
    print(f"✅ Filled {en_filled} EN translations for logros 19-48")

    # ─── Task 4: Fill executive bio ──────────────────────────────────────────
    ws_perfil = wb["01 Perfil"]
    current_es = ws_perfil.cell(row=18, column=3).value
    current_en = ws_perfil.cell(row=22, column=3).value
    if not (current_es and current_es.strip()):
        ws_perfil.cell(row=18, column=3, value=BIO_ES)
        print(f"✅ Filled Bio ES (R18C3): {BIO_ES[:80]}…")
    if not (current_en and current_en.strip()):
        ws_perfil.cell(row=22, column=3, value=BIO_EN)
        print(f"✅ Filled Bio EN (R22C3): {BIO_EN[:80]}…")

    # Save
    wb.save(str(MASTER))
    print(f"\n💾 Saved master")


if __name__ == "__main__":
    main()
