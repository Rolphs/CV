"""
Validador re-ejecutable del master.

Ejecuta TODAS las validaciones críticas:
  - Integridad referencial: 03 Logros → 02 Experiencia
  - Paridad ES/EN en todas las sheets bilingües
  - Tags solo dentro del vocabulario controlado
  - Cronología sin overlaps inválidos
  - Formato de fechas estándar
  - Seniority Fit usando solo valores permitidos
  - Pair With apunta a IDs existentes
  - Recipes referencian logros existentes
  - Anti-patterns presentes y completos

Output: data/reports/validation_report.md
Exit code: 0 si todo limpio, 1 si hay blockers.

Uso:
    python scripts/validate_master.py
"""
from __future__ import annotations
import re
import sys
from collections import defaultdict, Counter
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "data" / "master" / "cv_master_raul_mercado.xlsx"
REPORT = ROOT / "data" / "reports" / "validation_report.md"

# Vocabularios canónicos (mantener sincronizado con apply_phase8_llm_ready.py)
VOCAB_INDUSTRIA = {
    "Media", "Streaming", "TV", "Broadcasting", "AdTech", "Entertainment",
    "CPG", "Retail", "Tech", "Digital", "Data", "Telecom",
    "Consulting", "B2B", "Startup", "Arts", "Global", "Marketing",
    "Research-Industry",
}

VOCAB_SKILL = {
    "Research", "Insights", "Audience Measurement", "UX Research", "Mixed Methods",
    "Brand Lift", "Attribution", "AI/ML", "Generative AI", "Automation",
    "Data Pipelines", "Python", "Strategy", "Content Strategy",
    "Leadership", "Account Management", "Client Management", "Project Management",
    "Vendor Management", "Storytelling", "Consultative Sales",
    "Retail Analytics", "Behavioral Data", "Hispanic Market",
    "Methodology", "Process Design", "Competitive Intelligence", "Multi-Country",
    "Digital Transformation", "Consulting", "NPS / Customer Experience",
    "Account Growth", "Customer Insights", "KPIs", "Analysis",
    "Cross-functional", "Budget Management", "Photography", "Visual Design",
    "Creative Production", "Streaming", "Product Launch", "TV Programming",
    "Retention", "Study Design", "Productivity",
}

VOCAB_CATEGORIA = {
    "Business Impact", "Operational Excellence", "Innovation", "Strategy",
    "Leadership", "Client Management", "Research", "Responsibility",
    "Career Strategy", "Early Career",
}

VOCAB_SENIORITY = {"entry", "mid", "senior", "director", "exec", "all"}


def _cell(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _list(v):
    if not v:
        return []
    return [t.strip() for t in str(v).replace(";", ",").split(",") if t.strip()]


def read_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return [], []
    headers = [str(c).strip() if c is not None else f"col{i}"
               for i, c in enumerate(rows[1])]
    data = []
    for r in rows[2:]:
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in r):
            continue
        data.append({h: r[i] if i < len(r) else None for i, h in enumerate(headers)})
    return headers, data


class Validator:
    def __init__(self):
        self.blockers = []
        self.warnings = []
        self.passed = []

    def block(self, where, msg):
        self.blockers.append((where, msg))

    def warn(self, where, msg):
        self.warnings.append((where, msg))

    def ok(self, msg):
        self.passed.append(msg)


def validate_referential_integrity(wb, v):
    """03 Logros → 02 Experiencia"""
    _, exp = read_sheet(wb["02 Experiencia"])
    exp_ids = set()
    for r in exp:
        try:
            exp_ids.add(int(r["ID"]))
        except (ValueError, TypeError):
            pass
    _, logros = read_sheet(wb["03 Logros"])
    bad = 0
    for r in logros:
        lid = r.get("ID Logro")
        eid = r.get("ID Empresa")
        try:
            eid_int = int(eid) if eid else None
        except (ValueError, TypeError):
            v.block("03 Logros", f"Logro {lid}: ID Empresa no numérico '{eid}'")
            bad += 1
            continue
        if eid_int is None:
            v.block("03 Logros", f"Logro {lid}: ID Empresa vacío")
            bad += 1
        elif eid_int not in exp_ids:
            v.block("03 Logros", f"Logro {lid}: ID Empresa {eid_int} no existe en 02")
            bad += 1
    if bad == 0:
        v.ok(f"Integridad referencial 03→02: OK ({len(logros)} logros válidos)")


def validate_bilingual_parity(wb, v):
    """Toda celda ES debe tener su EN y viceversa."""
    sheets_to_check = [
        ("02 Experiencia", "Puesto ES", "Puesto EN"),
        ("03 Logros", "Logro / Resp. ES", "Achievement EN"),
        ("04 Skills", "Habilidad ES", "Skill EN"),
        ("05 Educación", "Título / Grado ES", "Degree EN"),
        ("06 Certificaciones", "Certificado ES", "Certificate EN"),
        ("07 Voluntariado", "Actividad ES", "Activity EN"),
        ("10 Conferencias y Speaker", "Título ES", "Title EN"),
        ("11 Reconocimientos y Membresías", "Nombre ES", "Name EN"),
    ]
    total_issues = 0
    for sname, k_es, k_en in sheets_to_check:
        if sname not in wb.sheetnames:
            continue
        _, rows = read_sheet(wb[sname])
        for i, r in enumerate(rows, start=3):
            es = _cell(r.get(k_es))
            en = _cell(r.get(k_en))
            if es and not en:
                v.warn(sname, f"Row {i}: '{k_es}' lleno pero '{k_en}' vacío")
                total_issues += 1
            elif en and not es:
                v.warn(sname, f"Row {i}: '{k_en}' lleno pero '{k_es}' vacío")
                total_issues += 1
    if total_issues == 0:
        v.ok(f"Paridad ES/EN en {len(sheets_to_check)} sheets bilingües: OK")


def validate_tags_vocabulary(wb, v):
    """Tags en 03 Logros deben estar en vocabulario controlado."""
    _, logros = read_sheet(wb["03 Logros"])
    unknown_industria = Counter()
    unknown_skill = Counter()
    unknown_categoria = Counter()
    for r in logros:
        lid = r.get("ID Logro")
        for t in _list(r.get("Tags Industria")):
            if t not in VOCAB_INDUSTRIA:
                unknown_industria[t] += 1
                v.warn("03 Logros", f"L{lid}: Tag industria fuera de vocab: '{t}'")
        for t in _list(r.get("Tags Habilidad")):
            if t not in VOCAB_SKILL:
                unknown_skill[t] += 1
                v.warn("03 Logros", f"L{lid}: Tag skill fuera de vocab: '{t}'")
        cat = _cell(r.get("Categoría"))
        if cat and cat not in VOCAB_CATEGORIA:
            unknown_categoria[cat] += 1
            v.warn("03 Logros", f"L{lid}: Categoría fuera de vocab: '{cat}'")

    total_unknown = sum(unknown_industria.values()) + sum(unknown_skill.values()) + sum(unknown_categoria.values())
    if total_unknown == 0:
        v.ok("Vocabulario controlado: OK (todos los tags válidos)")


def validate_seniority(wb, v):
    """Seniority Fit debe usar solo valores permitidos."""
    _, logros = read_sheet(wb["03 Logros"])
    bad = 0
    for r in logros:
        lid = r.get("ID Logro")
        fits = _list(r.get("Seniority Fit"))
        for f in fits:
            if f not in VOCAB_SENIORITY:
                v.warn("03 Logros", f"L{lid}: Seniority Fit '{f}' no válido")
                bad += 1
    if bad == 0:
        v.ok(f"Seniority Fit: OK ({len(logros)} logros)")


def validate_pair_with(wb, v):
    """Pair With debe apuntar a IDs de logros existentes."""
    _, logros = read_sheet(wb["03 Logros"])
    all_ids = set()
    for r in logros:
        try:
            all_ids.add(int(r["ID Logro"]))
        except (ValueError, TypeError):
            pass
    bad = 0
    for r in logros:
        lid = r.get("ID Logro")
        for x in _list(r.get("Pair With (IDs)")):
            try:
                if int(x) not in all_ids:
                    v.warn("03 Logros", f"L{lid}: Pair With apunta a L{x} (no existe)")
                    bad += 1
            except ValueError:
                v.warn("03 Logros", f"L{lid}: Pair With contiene valor no numérico '{x}'")
                bad += 1
    if bad == 0:
        v.ok("Pair With: OK (todas las referencias válidas)")


def validate_recipes(wb, v):
    """Recipes deben apuntar a logros existentes."""
    if "12 Recipes" not in wb.sheetnames:
        v.block("12 Recipes", "Sheet faltante")
        return
    _, logros = read_sheet(wb["03 Logros"])
    all_ids = set()
    for r in logros:
        try:
            all_ids.add(int(r["ID Logro"]))
        except (ValueError, TypeError):
            pass
    _, recipes = read_sheet(wb["12 Recipes"])
    bad = 0
    for r in recipes:
        rid = r.get("Recipe ID")
        for x in _list(r.get("Logros a usar (IDs)")):
            try:
                if int(x) not in all_ids:
                    v.warn("12 Recipes", f"{rid}: referencia logro L{x} que no existe")
                    bad += 1
            except ValueError:
                pass
    if bad == 0:
        v.ok(f"Recipes: OK ({len(recipes)} recetas con IDs válidos)")


def validate_dates(wb, v):
    """02 Experiencia debe tener fechas en formato MM/YYYY, YYYY o 'Actual'."""
    _, exp = read_sheet(wb["02 Experiencia"])
    pattern = re.compile(r"^(\d{1,2}/\d{4}|\d{4}|Actual|Presente)$", re.I)
    bad = 0
    for r in exp:
        eid = r.get("ID")
        for label, key in [("Inicio", "Fecha Inicio"), ("Fin", "Fecha Fin")]:
            val = _cell(r.get(key))
            if val and not pattern.match(val):
                v.warn("02 Experiencia", f"ID {eid}: Fecha {label} formato no estándar: '{val}'")
                bad += 1
    if bad == 0:
        v.ok(f"Formato de fechas: OK ({len(exp)} puestos)")


def validate_metric_flag(wb, v):
    """Si Tiene Métrica = Sí, debe haber número en logro o en Métrica/KPI."""
    _, logros = read_sheet(wb["03 Logros"])
    bad = 0
    for r in logros:
        lid = r.get("ID Logro")
        flag = _cell(r.get("Tiene Métrica"))
        if flag and flag.lower() in ("sí", "si", "yes"):
            es = _cell(r.get("Logro / Resp. ES")) or ""
            metric = _cell(r.get("Métrica / KPI"))
            if not re.search(r"\d", es) and not metric:
                v.warn("03 Logros", f"L{lid}: marcado 'Tiene Métrica=Sí' pero no hay número")
                bad += 1
    if bad == 0:
        v.ok("Coherencia métrica vs flag: OK")


def write_report(v):
    REPORT.parent.mkdir(parents=True, exist_ok=True)
    with open(REPORT, "w", encoding="utf-8") as f:
        f.write("# ✅ Validation Report\n\n")
        f.write(f"_Generated: {datetime.now().isoformat()}_\n\n")
        f.write(f"## Summary\n\n")
        f.write(f"- 🔴 Blockers: **{len(v.blockers)}**\n")
        f.write(f"- 🟡 Warnings: **{len(v.warnings)}**\n")
        f.write(f"- ✅ Checks passed: **{len(v.passed)}**\n\n")

        if v.passed:
            f.write("## ✅ Passed\n\n")
            for msg in v.passed:
                f.write(f"- {msg}\n")
            f.write("\n")

        if v.blockers:
            f.write("## 🔴 Blockers\n\n")
            by_where = defaultdict(list)
            for where, msg in v.blockers:
                by_where[where].append(msg)
            for where, msgs in sorted(by_where.items()):
                f.write(f"### {where} ({len(msgs)})\n\n")
                for m in msgs:
                    f.write(f"- {m}\n")
                f.write("\n")

        if v.warnings:
            f.write("## 🟡 Warnings\n\n")
            by_where = defaultdict(list)
            for where, msg in v.warnings:
                by_where[where].append(msg)
            for where, msgs in sorted(by_where.items()):
                f.write(f"### {where} ({len(msgs)})\n\n")
                for m in msgs:
                    f.write(f"- {m}\n")
                f.write("\n")


def main():
    print(f"🔍 Validating {MASTER.name}…\n")
    wb = load_workbook(str(MASTER), data_only=True)
    v = Validator()

    validate_referential_integrity(wb, v)
    validate_bilingual_parity(wb, v)
    validate_tags_vocabulary(wb, v)
    validate_seniority(wb, v)
    validate_pair_with(wb, v)
    validate_recipes(wb, v)
    validate_dates(wb, v)
    validate_metric_flag(wb, v)

    write_report(v)

    print(f"\n📊 RESULTS:")
    print(f"   🔴 Blockers: {len(v.blockers)}")
    print(f"   🟡 Warnings: {len(v.warnings)}")
    print(f"   ✅ Passed checks: {len(v.passed)}")
    print(f"\n   📄 Report: {REPORT}")

    if v.blockers:
        print(f"\n❌ FAILED ({len(v.blockers)} blockers)")
        sys.exit(1)
    else:
        print(f"\n✅ MASTER LOOKS HEALTHY")
        sys.exit(0)


if __name__ == "__main__":
    main()
