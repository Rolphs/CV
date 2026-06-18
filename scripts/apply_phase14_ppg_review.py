"""
Fase 14 · Master updates per Rolph 2026-05-17 (post-PPG-review).

Changes:
  A. Empresa #6 Bricko: fecha_inicio 09/2020 -> 11/2020 (Televisa cerró Oct 2020)
  B. NEW Empresa #29 "Independent Consulting" (Feb 2025 - Aug 2025)
     - Real entry to back the CV's "Independent Consultant — Research, AI & UX
       Strategy · Feb 2025 – Aug 2025" entry that was floating without master row.
  C. Update L18 Yoorco: remove "30% NPS" (duplicates Canela L62-ish), replace
     with new narrative about nearshore proxy team, 95%+ satisfaction, 10+
     concurrent projects, 4-month discovery-to-live.
  D. Deprecate L63 Yoorco (also "30% NPS" duplicate) -> Incluir CV ES/EN = No.
  E. NEW L78 for Independent Consulting #29: FMCG/Pharma/HORECA clients
     with brand books, AI-augmented competitive analyses, iterative UX
     optimization roadmaps.
  F. Add 2 skills under "Analytics / IA":
     - Retail Analytics
     - Retail Intelligence

Anti-pattern A14 honored: zero invented metrics. All numbers and sector
names confirmed by Rolph in conversation (HORECA confirmed, proxy-team
geography confirmed, 95%/10/2/4 dictated by Rolph).
"""
import shutil
import sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "data" / "master" / "cv_master_raul_mercado.xlsx"
BACKUP = ROOT / "data" / "master" / "backups" / "v14_pre_ppg_review.xlsx"


# ─────────────────────────── Payloads ───────────────────────────


BRICKO_DATE_FIX = {
    "Fecha Inicio": "11/2020",
}


INDEP_CONSULTING_2025 = {
    "ID": 29,
    "Empresa / Company": "Independent Consulting",
    "Industria / Industry": "Consultoría · Research / AI / UX",
    "País": "México",
    "Puesto ES": "Consultor Independiente",
    "Puesto EN": "Independent Consultant — Research, AI & UX Strategy",
    "Nivel": "Director / Principal",
    "Tipo": "Freelance",
    "Fecha Inicio": "02/2025",
    "Fecha Fin": "08/2025",
    "Dept / Área": "Consultoría",
    "Reporta a": "Directo a clientes",
    "Personas a cargo": "Equipos ad-hoc por proyecto",
    "Descripción Empresa ES": (
        "Práctica independiente de consultoría aplicada en investigación, "
        "IA generativa y diseño UX para clientes en FMCG, Pharma y HORECA."
    ),
    "Descripción Empresa EN": (
        "Independent consulting practice in applied research, generative AI "
        "and UX design for clients across FMCG, Pharma and HORECA."
    ),
}


L18_REWRITE = {
    "EN": (
        "Founded and led Yoorco as an AI-based digital startup (UX + No-Code "
        "tools), building a nearshore developer team from Venezuela, Argentina "
        "and Colombia delivering e-commerce, light apps and ad-hoc features to "
        "clients in Europe and the US — bringing better-paid projects back to "
        "the region. Yoorco achieved 95%+ aggregated client satisfaction, "
        "managing 10+ concurrent projects (2 launches/month) with a 4-month "
        "average discovery-to-live-product time."
    ),
    "ES": (
        "Fundé y lideré Yoorco como startup digital basada en IA (UX + "
        "herramientas No-Code), construyendo un equipo nearshore de "
        "desarrolladores de Venezuela, Argentina y Colombia para entregar "
        "e-commerce, light apps y features ad-hoc a clientes de Europa y "
        "EE.UU. — regresando proyectos mejor pagados a la región. Yoorco "
        "alcanzó >95% de satisfacción agregada, gestionando 10+ proyectos "
        "en paralelo (2 lanzamientos/mes) y un tiempo promedio de 4 meses "
        "de discovery a producto funcional."
    ),
}


LOGRO_L78 = {
    "ID Logro": 78,
    "ID Empresa": 29,
    "Empresa": "Independent Consulting",
    "Tipo": "Logro",
    "Logro / Resp. ES": (
        "Atendí clientes de FMCG, Pharma y HORECA con entregables que "
        "incluyeron brand books, análisis competitivos potenciados con IA y "
        "rediseños iterativos de UX guiados por investigación de usuario "
        "(mapeo de patrones de navegación + propuesta de mejoras incrementales "
        "para sitios digitales de clientes)."
    ),
    "Achievement EN": (
        "Served clients across FMCG, Pharma and HORECA with deliverables "
        "spanning brand books, AI-augmented competitive analyses, and "
        "research-led iterative UX optimization roadmaps (user-navigation "
        "mapping + incremental improvement blueprints for client digital "
        "properties)."
    ),
    "Impacto": (
        "Demuestra rango sectorial (FMCG/Pharma/HORECA), capacidad de combinar "
        "IA + research + UX, y entregables tangibles. Cubre cronología "
        "Feb-Ago 2025."
    ),
    "Tiene Métrica": "No",
    "Métrica / KPI": "",
    "Categoría": "Consultoría",
    "Tags Industria": "['FMCG', 'Pharma', 'HORECA']",
    "Tags Habilidad": "['UX', 'AI', 'Research', 'Competitive Intelligence', 'Brand Strategy']",
    "Seniority Fit": "['director', 'exec']",
    "Story / Evidence": (
        "Práctica independiente Feb-Ago 2025 entre Canela y Walmart. "
        "Sectores confirmados por Rolph 2026-05-17."
    ),
    "Pair With (IDs)": "",
    "Incluir CV ES": None,
    "Incluir CV EN": None,
}


NEW_SKILLS = [
    {
        "Categoría": "Analytics / IA",
        "Habilidad ES": "Retail Analytics",
        "Skill EN": "Retail Analytics",
        "Nivel": "Avanzado",
        "Años exp.": "3+",
        "Descripción ES": "Análisis transaccional retail (Walmart Scintilla / Data Ventures) aplicado a pricing, promoción y assortment.",
        "Description EN": "Transactional retail analytics (Walmart Scintilla / Data Ventures) applied to pricing, promotion and assortment.",
        "Destacar en": "R03",
    },
    {
        "Categoría": "Analytics / IA",
        "Habilidad ES": "Retail Intelligence",
        "Skill EN": "Retail Intelligence",
        "Nivel": "Avanzado",
        "Años exp.": "3+",
        "Descripción ES": "Inteligencia comercial sobre datos retail para fabricantes FMCG/CPG, traduciendo comportamiento de compra en recomendaciones ejecutivas.",
        "Description EN": "Commercial intelligence on retail data for FMCG/CPG manufacturers, translating purchase behavior into executive recommendations.",
        "Destacar en": "R03",
    },
]


# ─────────────────────────── Engine ───────────────────────────


def _header_map(ws, header_row: int = 2) -> dict[str, int]:
    return {
        ws.cell(row=header_row, column=c).value: c
        for c in range(1, 30)
        if ws.cell(row=header_row, column=c).value
    }


def _existing_ids(ws, id_col: int, header_row: int) -> set:
    out = set()
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=id_col).value
        if v not in (None, ""):
            out.add(v)
    return out


def _first_empty_row(ws, header_row: int, key_col: int) -> int:
    for r in range(header_row + 1, ws.max_row + 2):
        if ws.cell(row=r, column=key_col).value in (None, ""):
            return r
    return ws.max_row + 1


def _apply_row(ws, row: int, payload: dict, headers: dict[str, int]):
    for field_name, value in payload.items():
        col = headers.get(field_name)
        if col is None:
            print(f"   ⚠ Column '{field_name}' not in sheet — skipped")
            continue
        ws.cell(row=row, column=col).value = value


# ─────────────────────────── Operations ───────────────────────────


def fix_bricko_date(wb):
    ws = wb["02 Experiencia"]
    h = _header_map(ws)
    for r in range(3, ws.max_row + 1):
        if str(ws.cell(row=r, column=h["Empresa / Company"]).value).lower() == "bricko":
            cur = ws.cell(row=r, column=h["Fecha Inicio"]).value
            if cur == "11/2020":
                print(f"   ⏭  Bricko fecha_inicio already 11/2020 — skipped")
                return False
            print(f"   📍 Bricko row {r} · fecha_inicio: {cur!r} -> '11/2020'")
            _apply_row(ws, r, BRICKO_DATE_FIX, h)
            return True
    print("   ❌ Bricko row not found")
    return False


def insert_indep_2025(wb):
    ws = wb["02 Experiencia"]
    h = _header_map(ws)
    existing = _existing_ids(ws, h["ID"], header_row=2)
    new_id = INDEP_CONSULTING_2025["ID"]
    if new_id in existing:
        print(f"   ⏭  Independent Consulting #{new_id} already exists — skipped")
        return False
    row = _first_empty_row(ws, header_row=2, key_col=h["ID"])
    _apply_row(ws, row, INDEP_CONSULTING_2025, h)
    print(f"   ✅ Added Independent Consulting (#{new_id}) at row {row}")
    return True


def rewrite_l18(wb):
    ws = wb["03 Logros"]
    h = _header_map(ws)
    for r in range(3, ws.max_row + 1):
        if ws.cell(row=r, column=h["ID Logro"]).value == 18:
            cur_en = ws.cell(row=r, column=h["Achievement EN"]).value or ""
            if "nearshore" in cur_en.lower():
                print(f"   ⏭  L18 already rewritten — skipped")
                return False
            print(f"   📍 L18 BEFORE: {cur_en[:100]}...")
            ws.cell(row=r, column=h["Achievement EN"]).value = L18_REWRITE["EN"]
            ws.cell(row=r, column=h["Logro / Resp. ES"]).value = L18_REWRITE["ES"]
            print(f"   ✅ L18 AFTER: {L18_REWRITE['EN'][:100]}...")
            return True
    print("   ❌ L18 not found")
    return False


def deprecate_l63(wb):
    ws = wb["03 Logros"]
    h = _header_map(ws)
    cv_es_col = h.get("Incluir CV ES")
    cv_en_col = h.get("Incluir CV EN")
    for r in range(3, ws.max_row + 1):
        if ws.cell(row=r, column=h["ID Logro"]).value == 63:
            already = ws.cell(row=r, column=cv_es_col).value if cv_es_col else None
            if already == "No":
                print(f"   ⏭  L63 already deprecated — skipped")
                return False
            if cv_es_col:
                ws.cell(row=r, column=cv_es_col).value = "No"
            if cv_en_col:
                ws.cell(row=r, column=cv_en_col).value = "No"
            print(f"   ✅ L63 marked Incluir CV = 'No' (duplicate 30% NPS metric)")
            return True
    print("   ❌ L63 not found")
    return False


def add_logro_l78(wb):
    ws = wb["03 Logros"]
    h = _header_map(ws)
    existing = _existing_ids(ws, h["ID Logro"], header_row=2)
    if LOGRO_L78["ID Logro"] in existing:
        print(f"   ⏭  L78 already exists — skipped")
        return False
    row = _first_empty_row(ws, header_row=2, key_col=h["ID Logro"])
    _apply_row(ws, row, LOGRO_L78, h)
    print(f"   ✅ Added L78 at row {row} for Independent Consulting 2025")
    return True


def add_new_skills(wb):
    ws = wb["04 Skills"]
    h = _header_map(ws)
    skill_col = h["Skill EN"]
    existing_skills = {
        str(ws.cell(row=r, column=skill_col).value).lower()
        for r in range(3, ws.max_row + 1)
        if ws.cell(row=r, column=skill_col).value
    }
    added = 0
    for payload in NEW_SKILLS:
        key = payload["Skill EN"].lower()
        if key in existing_skills:
            print(f"   ⏭  Skill '{payload['Skill EN']}' already exists — skipped")
            continue
        row = _first_empty_row(ws, header_row=2, key_col=h["Categoría"])
        _apply_row(ws, row, payload, h)
        print(f"   ✅ Added skill '{payload['Skill EN']}' at row {row}")
        added += 1
    return added > 0


# ─────────────────────────── Main ───────────────────────────


def main():
    print(f"📂 Loading {MASTER.name}")
    BACKUP.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(MASTER, BACKUP)
    print(f"💾 Backup: {BACKUP.relative_to(ROOT)}")

    wb = load_workbook(MASTER)
    n = 0

    print("\n🏢 A · Fix Bricko fecha_inicio 09/2020 -> 11/2020")
    n += int(fix_bricko_date(wb))

    print("\n🏢 B · Insert Independent Consulting 2025 (#29)")
    n += int(insert_indep_2025(wb))

    print("\n🍳 C · Rewrite L18 Yoorco (remove 30% NPS, add proxy-team narrative)")
    n += int(rewrite_l18(wb))

    print("\n🍳 D · Deprecate L63 Yoorco (duplicate 30% NPS)")
    n += int(deprecate_l63(wb))

    print("\n🍳 E · Add L78 (Independent Consulting · FMCG/Pharma/HORECA)")
    n += int(add_logro_l78(wb))

    print("\n🧩 F · Add 2 skills (Retail Analytics, Retail Intelligence)")
    n += int(add_new_skills(wb))

    if n:
        wb.save(MASTER)
        print(f"\n✅ Saved {n} change(s) to {MASTER.name}")
        print("\nNext:  python scripts/export_master_json.py")
    else:
        print("\nℹ Nothing changed — already applied.")


if __name__ == "__main__":
    main()
