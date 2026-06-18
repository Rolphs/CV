"""
Fase 3 apply · Update master with curated candidates from extract_logros_phase3.

Logic:
  1. Load candidates JSON.
  2. Load master logros.
  3. For each empresa, take top Q>=4 EN candidates (sorted by quality desc).
  4. For each candidate:
       a) Find best match in existing logros via token overlap (cross-lang Jaccard).
       b) If existing EN column empty AND match >= 0.10 (loose) AND quality >=5,
          UPDATE the existing row's Achievement EN.
       c) Otherwise, ADD as new logro row with EN filled and ES placeholder.
  5. Cap NEW additions at 3 per empresa to avoid bloat.
  6. Save master in-place (backup v2_pre_phase3 already taken).

Special: Altazor (#3) has 0 logros — add curated 4 picks directly.
"""

from __future__ import annotations
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "data" / "master" / "cv_master_raul_mercado.xlsx"
CAND = ROOT / "data" / "reports" / "logros_phase3_candidates.json"

MAX_NEW_PER_EMPRESA = 3
MIN_QUALITY_FOR_NEW = 4
MIN_QUALITY_FOR_FILL_EN = 5
MIN_MATCH_FOR_FILL_EN = 0.10  # cross-lang naturally low


def normalize_text(t: str) -> str:
    t = unicodedata.normalize("NFKD", t)
    return "".join(c for c in t if not unicodedata.combining(c)).lower()


def tokens(t: str) -> set:
    norm = normalize_text(t)
    toks = re.findall(r"[a-z0-9]+", norm)
    STOP = {"de","la","el","en","y","a","del","los","las","para","con","por",
            "un","una","the","of","and","to","in","for","on","at","from","with",
            "by","as","an","or","que","se","su","sus","es","lo","le","al","ha",
            "fue","han","muy","mas","como"}
    return {t for t in toks if len(t) >= 3 and t not in STOP}


def jaccard(a: set, b: set) -> float:
    return len(a & b) / len(a | b) if (a and b) else 0.0


# Curation for Altazor (#3) — 0 logros in master, need solid starter set
ALTAZOR_CURATED = [
    {
        "es": "Diseño y ejecución de procesos de investigación con automatización por IA, reduciendo tiempos de entrega hasta 90%.",
        "en": "Built AI-powered insight pipelines reducing reporting turnaround by 90%.",
        "tipo": "Logro",
        "categoria": "Innovación / IA",
        "impacto": "Alto",
        "metrica": "90% reducción",
        "tags_ind": "Consultoría,Research",
        "tags_skill": "IA,Automatización,Research",
    },
    {
        "es": "Rediseño de procesos analíticos y operativos en investigación, con enfoque en KPIs predictivos.",
        "en": "Designed predictive KPI ecosystems enabling real-time media performance optimization.",
        "tipo": "Logro",
        "categoria": "Estrategia / Arquitectura",
        "impacto": "Alto",
        "metrica": "",
        "tags_ind": "Consultoría,Research,Media",
        "tags_skill": "KPI,Análisis,Estrategia",
    },
    {
        "es": "Coordinación de equipos multifuncionales en proyectos de consultoría estratégica de investigación.",
        "en": "Led cross-functional integration between product, data science, and research teams.",
        "tipo": "Responsabilidad",
        "categoria": "Liderazgo",
        "impacto": "Medio",
        "metrica": "",
        "tags_ind": "Consultoría,Research",
        "tags_skill": "Liderazgo,Multidisciplinario",
    },
]


def main():
    print(f"Backup: data/master/backups/v2_pre_phase3.xlsx (ya creado)")
    cand_data = json.loads(CAND.read_text(encoding="utf-8"))

    wb = load_workbook(str(MASTER))
    ws_log = wb["03 Logros"]

    # Index existing logros: row_index, empresa_id, es, en
    rows_by_emp: dict[int, list[dict]] = defaultdict(list)
    max_id = 0
    for ridx, row in enumerate(ws_log.iter_rows(min_row=3, max_col=14, values_only=False), start=3):
        if row[0].value is None:
            continue
        try:
            emp_id = int(row[1].value) if row[1].value is not None else None
        except (TypeError, ValueError):
            continue
        rows_by_emp[emp_id].append({
            "row_idx": ridx,
            "id": int(row[0].value),
            "es": row[4].value or "",
            "en": row[5].value or "",
            "row": row,
        })
        max_id = max(max_id, int(row[0].value))

    print(f"Loaded {sum(len(v) for v in rows_by_emp.values())} existing logros, max id = {max_id}")

    en_fills = 0
    new_adds = 0
    next_id = max_id + 1

    for emp_key, c in cand_data.items():
        emp_id = int(emp_key)
        en_pool = sorted(c.get("EN_FOUND", []), key=lambda x: -x["quality"])
        if not en_pool:
            continue

        # Pass 1: try to fill empty EN in existing logros (one EN per existing row)
        existing = rows_by_emp.get(emp_id, [])
        for ex in existing:
            if ex["en"].strip():
                continue  # already has EN
            ex_tokens = tokens(ex["es"])
            best = (None, 0.0)
            for cand in en_pool:
                if cand.get("_used"):
                    continue
                score = jaccard(tokens(cand["text"]), ex_tokens)
                if score > best[1]:
                    best = (cand, score)
            if best[0] and best[0]["quality"] >= MIN_QUALITY_FOR_FILL_EN and best[1] >= MIN_MATCH_FOR_FILL_EN:
                ws_log.cell(row=ex["row_idx"], column=6, value=best[0]["text"])
                best[0]["_used"] = True
                en_fills += 1
                print(f"  ✓ filled EN for #{ex['id']} ({emp_id}): {best[0]['text'][:70]}…")

        # Pass 2: add top NEW candidates (unused, Q>=4)
        added = 0
        for cand in en_pool:
            if added >= MAX_NEW_PER_EMPRESA:
                break
            if cand.get("_used"):
                continue
            if cand["quality"] < MIN_QUALITY_FOR_NEW:
                continue
            # Skip obvious header/junk
            txt = cand["text"]
            if re.search(r"\d{2}/\d{4}", txt) or txt.count(" - ") >= 3:
                continue
            empresa_name = next((row["row"][2].value for row in existing if row["row"][2].value), "")
            new_row_idx = ws_log.max_row + 1
            new_values = [
                next_id, emp_id, empresa_name,
                "Logro",
                f"[Traducir desde EN] {txt}",  # ES placeholder
                txt,                            # EN canonical
                "Alto" if cand["quality"] >= 6 else "Medio",
                "Sí" if re.search(r"\d+\s*%|\$|\d+\s*x", txt, re.I) else "No",
                "",  # Métrica
                "",  # Categoría
                "",  # Tags Industria
                "",  # Tags Habilidad
                "Sí", "Sí",
            ]
            for col_idx, val in enumerate(new_values, start=1):
                ws_log.cell(row=new_row_idx, column=col_idx, value=val)
            next_id += 1
            new_adds += 1
            added += 1
            print(f"  + NEW #{next_id-1} ({emp_id}) Q{cand['quality']}: {txt[:70]}…")

    # Altazor curated (empresa #3, 0 logros in master)
    altazor_existing = rows_by_emp.get(3, [])
    empresa_altazor = "Altazor Intelligence"
    if not altazor_existing:
        print(f"\n🌟 Altazor curated additions:")
        for cur in ALTAZOR_CURATED:
            new_row_idx = ws_log.max_row + 1
            new_values = [
                next_id, 3, empresa_altazor,
                cur["tipo"],
                cur["es"],
                cur["en"],
                cur["impacto"],
                "Sí" if cur["metrica"] else "No",
                cur["metrica"],
                cur["categoria"],
                cur["tags_ind"],
                cur["tags_skill"],
                "Sí", "Sí",
            ]
            for col_idx, val in enumerate(new_values, start=1):
                ws_log.cell(row=new_row_idx, column=col_idx, value=val)
            print(f"  + ALTAZOR #{next_id}: {cur['es'][:70]}…")
            next_id += 1
            new_adds += 1

    print(f"\n💾 Saving master...")
    wb.save(str(MASTER))
    print(f"\n✅ Done — EN fills: {en_fills} · NEW logros added: {new_adds}")
    print(f"Master now has logros up to ID {next_id - 1}")


if __name__ == "__main__":
    main()
