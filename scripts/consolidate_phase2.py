"""
Consolidate Fase 2: Add the 10 missing puestos from 'Formato Datos CV - Raul Mercado.xlsx'
into 'cv_master_raul_mercado.xlsx' sheets '02 Experiencia' and '03 Logros'.

Process:
  1. Load both workbooks.
  2. Read existing puestos in master (sheet '02 Experiencia') → set of (empresa, fecha_inicio).
  3. Read all puestos in Formato Datos (sheet 'Experiencia').
  4. For each puesto not in master, append a new row with:
       - new sequential ID
       - copied ES fields
       - inferred/translated EN fields
       - inferred Nivel/Tipo defaults
  5. Also split 'LOGROS DESTACADOS' by ';' and append each as a row in '03 Logros'.
  6. Save master in-place (backup already taken).
"""

from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "data" / "master" / "cv_master_raul_mercado.xlsx"
SOURCE = ROOT / "data" / "raw" / "04_bases_datos_excel" / "Formato Datos CV - Raul Mercado.xlsx"


# ─── translation helpers ─────────────────────────────────────────────────────

PUESTO_EN = {
    "Consultor": "Consultant",
    "Consultor en Ventas, Preventa y Diseño de Soluciones": "Sales, Pre-Sales & Solutions Design Consultant",
    "Gerente de Marketing e Inteligencia de Negocios": "Marketing & Business Intelligence Manager",
    "Consultor Estratégico para Dirección General": "Strategic Consultant to General Management",
    "Moderador / Analista": "Moderator / Analyst",
    "Microempresario": "Founder / Small Business Owner",
    "Técnico en Streaming de Video": "Video Streaming Technician",
    "Asistente de Dirección y Dirección de Arte": "Assistant Director & Art Director",
    "Empresario Audiovisual": "Audiovisual Entrepreneur",
    "Técnico en Servicio": "Service Technician",
}

INDUSTRY_EN = {
    "Consultoría / Marketing / Insights": "Consulting / Marketing / Insights",
    "Tecnología / Servicios TI": "Technology / IT Services",
    "Investigación de Mercados": "Market Research",
    "Salud y Belleza": "Health & Beauty",
    "Tecnología / Audiovisual": "Technology / Audiovisual",
    "Producción Audiovisual": "Audiovisual Production",
    "Producción Audiovisual / Web": "Audiovisual Production / Web",
    "Tecnología / Reparación": "Technology / Repair",
}

COUNTRY_EN = {"México": "Mexico"}

EMPRESA_DESC_ES = {
    "Independiente": "Práctica independiente de consultoría en investigación y marketing.",
    "TIBA IT Services (KIO Networks)": "Unidad de servicios de TI dentro del grupo KIO Networks, foco en preventa y soluciones empresariales.",
    "TIBA IT Services by KIO Networks": "Unidad de servicios de TI dentro del grupo KIO Networks, foco en preventa y soluciones empresariales.",
    "NODO Investigación + Estrategia": "Firma pionera en investigación cualitativa y UX en México (raíz de Nodo Research).",
    "Xabón'art": "Microemprendimiento artesanal de productos de higiene y belleza sustentables.",
    "Focus Vision / Channel M3": "Proveedor global de streaming y soporte técnico para estudios de investigación.",
    "PTA / GBA y otras productoras": "Productoras audiovisuales para campañas publicitarias y contenido editorial.",
    "Onírica": "Casa propia de producción audiovisual y desarrollo web.",
    "Central de Registradoras S.A.": "Compañía dedicada al servicio y reparación de equipo de cómputo y cajas registradoras.",
}

EMPRESA_DESC_EN = {
    "Independiente": "Independent consulting practice in research and marketing.",
    "TIBA IT Services (KIO Networks)": "IT services unit within the KIO Networks group, focused on pre-sales and enterprise solutions.",
    "TIBA IT Services by KIO Networks": "IT services unit within the KIO Networks group, focused on pre-sales and enterprise solutions.",
    "NODO Investigación + Estrategia": "Pioneer qualitative research and UX firm in Mexico (root of Nodo Research).",
    "Xabón'art": "Artisanal micro-business of sustainable hygiene and beauty products.",
    "Focus Vision / Channel M3": "Global provider of streaming and technical support for market research studies.",
    "PTA / GBA y otras productoras": "Audiovisual production houses for advertising campaigns and editorial content.",
    "Onírica": "Own audiovisual production and web development studio.",
    "Central de Registradoras S.A.": "Company dedicated to servicing and repairing computer hardware and cash registers.",
}

# Defaults for fields the auxiliary xlsx doesn't have
NIVEL_DEFAULT = {
    "Consultor en Ventas, Preventa y Diseño de Soluciones": "Consultor Sr",
    "Gerente de Marketing e Inteligencia de Negocios": "Gerente",
    "Consultor Estratégico para Dirección General": "Consultor Sr",
    "Moderador / Analista": "Analista",
    "Consultor": "Consultor Sr",
    "Microempresario": "Fundador",
    "Técnico en Streaming de Video": "Técnico",
    "Asistente de Dirección y Dirección de Arte": "Asistente / Director de Arte",
    "Empresario Audiovisual": "Fundador",
    "Técnico en Servicio": "Técnico",
}

TIPO_DEFAULT = {
    "Consultor en Ventas, Preventa y Diseño de Soluciones": "Full-time",
    "Gerente de Marketing e Inteligencia de Negocios": "Full-time",
    "Consultor Estratégico para Dirección General": "Consultoría",
    "Moderador / Analista": "Full-time",
    "Consultor": "Freelance",
    "Microempresario": "Emprendimiento",
    "Técnico en Streaming de Video": "Freelance",
    "Asistente de Dirección y Dirección de Arte": "Freelance",
    "Empresario Audiovisual": "Emprendimiento",
    "Técnico en Servicio": "Full-time",
}


# ─── main ─────────────────────────────────────────────────────────────────────

def normalize(s: str) -> str:
    return (s or "").strip().lower().replace("\u00ed", "i")


def main():
    print(f"Loading master: {MASTER.name}")
    wb_master = load_workbook(str(MASTER))
    print(f"Loading source: {SOURCE.name}")
    wb_src = load_workbook(str(SOURCE), data_only=True)

    ws_exp = wb_master["02 Experiencia"]
    ws_log = wb_master["03 Logros"]

    # Existing puestos in master (rows after header — header is row 2 per dump)
    existing = set()
    max_existing_id = 0
    last_exp_row = ws_exp.max_row
    for row in ws_exp.iter_rows(min_row=3, max_col=10, values_only=True):
        if row[0] is None:
            continue
        id_, empresa, _industry, _country, _puesto_es, _puesto_en, _nivel, _tipo, fecha_inicio, _fecha_fin = row
        key = (normalize(str(empresa)), normalize(str(fecha_inicio)))
        existing.add(key)
        try:
            max_existing_id = max(max_existing_id, int(id_))
        except (TypeError, ValueError):
            pass
    print(f"Master currently has {len(existing)} puestos, max ID = {max_existing_id}")

    # Existing logros IDs
    max_logro_id = 0
    for row in ws_log.iter_rows(min_row=3, max_col=1, values_only=True):
        try:
            max_logro_id = max(max_logro_id, int(row[0]))
        except (TypeError, ValueError):
            pass
    print(f"Master currently has logro max ID = {max_logro_id}")

    # Read source puestos
    ws_src = wb_src["Experiencia"]
    src_rows = []
    for row in ws_src.iter_rows(min_row=1, max_col=8, values_only=True):
        # First col = NOMBRE POSICION, 2 header rows, then data
        puesto, empresa, industria, ubicacion, fecha_ini, fecha_fin, resp, logros = row[:8]
        if not empresa or empresa in ("COMPAÑÍA", "Razón social"):
            continue
        if not puesto:
            continue
        src_rows.append({
            "puesto_es": str(puesto).strip(),
            "empresa": str(empresa).strip(),
            "industria": str(industria or "").strip(),
            "pais": str(ubicacion or "México").strip(),
            "fecha_inicio": str(fecha_ini or "").strip(),
            "fecha_fin": str(fecha_fin or "").strip(),
            "responsabilidades": str(resp or "").strip(),
            "logros": str(logros or "").strip(),
        })
    print(f"Source has {len(src_rows)} puestos")

    # Find which are missing
    new_rows = []
    for r in src_rows:
        key = (normalize(r["empresa"]), normalize(r["fecha_inicio"]))
        if key not in existing:
            new_rows.append(r)
    print(f"--> {len(new_rows)} puestos NEW to add\n")

    # Append to master
    next_id = max_existing_id + 1
    next_logro_id = max_logro_id + 1
    append_row_exp = last_exp_row + 1

    for r in new_rows:
        puesto_en = PUESTO_EN.get(r["puesto_es"], r["puesto_es"])
        industria_en = INDUSTRY_EN.get(r["industria"], r["industria"])
        pais_en = COUNTRY_EN.get(r["pais"], r["pais"])
        nivel = NIVEL_DEFAULT.get(r["puesto_es"], "")
        tipo = TIPO_DEFAULT.get(r["puesto_es"], "")
        desc_es = EMPRESA_DESC_ES.get(r["empresa"], "")
        desc_en = EMPRESA_DESC_EN.get(r["empresa"], "")

        new_row = [
            next_id,              # c0 ID
            r["empresa"],         # c1 Empresa
            r["industria"],       # c2 Industria ES (master uses ES only here)
            pais_en if pais_en != r["pais"] else r["pais"],  # c3 País
            r["puesto_es"],       # c4 Puesto ES
            puesto_en,            # c5 Puesto EN
            nivel,                # c6 Nivel
            tipo,                 # c7 Tipo
            r["fecha_inicio"],    # c8 Fecha Inicio
            r["fecha_fin"],       # c9 Fecha Fin
            "No",                 # c10 ¿Actual?
            "",                   # c11 Dept / Área
            "",                   # c12 Reporta a
            "",                   # c13 Personas a cargo
            desc_es,              # c14 Descripción Empresa ES
            desc_en,              # c15 Descripción Empresa EN
        ]
        for col_idx, val in enumerate(new_row, start=1):
            ws_exp.cell(row=append_row_exp, column=col_idx, value=val)
        print(f"  + #{next_id} {r['empresa']} ({r['puesto_es']}) [{r['fecha_inicio']} - {r['fecha_fin']}]")

        # Logros — split by ';'
        if r["logros"]:
            chunks = [c.strip() for c in r["logros"].replace("\n", ";").split(";") if c.strip()]
            for chunk in chunks:
                logro_row = [
                    next_logro_id,    # ID Logro
                    next_id,          # ID Empresa
                    r["empresa"],     # Empresa
                    "Logro",          # Tipo
                    chunk,            # Logro ES
                    "",               # Achievement EN (pending)
                    "Medio",          # Impacto (default)
                    "No",             # Tiene Métrica
                    "",               # Métrica
                    "",               # Categoría
                    "",               # Tags Industria
                    "",               # Tags Habilidad
                    "Sí",             # Incluir CV ES
                    "",               # Incluir CV EN (blank since EN missing)
                ]
                logro_target_row = ws_log.max_row + 1
                for col_idx, val in enumerate(logro_row, start=1):
                    ws_log.cell(row=logro_target_row, column=col_idx, value=val)
                print(f"      ↳ logro #{next_logro_id}: {chunk[:60]}…")
                next_logro_id += 1

        # Responsabilidades as a single row
        if r["responsabilidades"]:
            resp_row = [
                next_logro_id,
                next_id,
                r["empresa"],
                "Responsabilidad",
                r["responsabilidades"],
                "",
                "Medio",
                "No",
                "",
                "Responsabilidad",
                "",
                "",
                "Sí",
                "",
            ]
            resp_target_row = ws_log.max_row + 1
            for col_idx, val in enumerate(resp_row, start=1):
                ws_log.cell(row=resp_target_row, column=col_idx, value=val)
            print(f"      ↳ resp #{next_logro_id}: {r['responsabilidades'][:60]}…")
            next_logro_id += 1

        next_id += 1
        append_row_exp += 1

    print(f"\nSaving master back to {MASTER.name}...")
    wb_master.save(str(MASTER))
    print("✅ Done.")


if __name__ == "__main__":
    main()
