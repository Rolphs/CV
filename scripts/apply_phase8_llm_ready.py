"""
Fase 8 · Hacer el master LLM-ready.

Cambios:
  1. Reescribe 00 Instrucciones como MANUAL OPERATIVO PARA LLM
  2. Agrega columnas a 03 Logros:
       - Seniority Fit (entry/mid/senior/director/exec/all)
       - Story / Evidence (backstory 1-2 líneas)
       - Pair With (IDs de logros que cuentan mejor historia juntos)
  3. Normaliza tags a vocabulario controlado (industria, habilidad, categoría)
  4. Agrega sheet 12 Recipes (recetas de CV pre-pensadas)
  5. Agrega sheet 13 Anti-Patterns (qué NO decir, palabras a evitar)
"""
from __future__ import annotations
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "data" / "master" / "cv_master_raul_mercado.xlsx"
BACKUP = ROOT / "data" / "master" / "backups" / "v8_pre_llm_ready.xlsx"

# Walmart palette (mantener consistencia con fase 7)
BLUE_100 = "0053E2"
GREEN_100 = "2A8703"
SPARK_100 = "FFC220"
GRAY_5 = "F5F5F5"


# ═══ VOCABULARIO CONTROLADO ════════════════════════════════════════════════
INDUSTRIA_MAP = {
    # Spanish → English canonical + consolidations
    "tecnología": "Tech", "tecnologia": "Tech",
    "consultoría": "Consulting", "consultoria": "Consulting", "consulting": "Consulting",
    "startups": "Startup", "startup": "Startup",
    "fmcg": "CPG",  # FMCG and CPG are synonyms, standardize on CPG
    "entretenimiento": "Entertainment",
    "marketing": "Marketing",
    "internacional": "Global",
    "tv": "TV",
    "theater": "Arts",
    "data": "Data",
    "broadcasting": "Broadcasting",
    "media": "Media",
    "streaming": "Streaming",
    "research": "Research-Industry",  # rare as industry tag
    "tech": "Tech",
    "adtech": "AdTech",
    "cpg": "CPG",
    "retail": "Retail",
    "b2b": "B2B",
    "telecom": "Telecom",
    "digital": "Digital",
    "arts": "Arts",
}

SKILL_MAP = {
    # Standardize to English (LLM-native) and consolidate
    "ia": "AI/ML", "ia generativa": "Generative AI", "ai": "AI/ML",
    "automatización": "Automation", "automatizacion": "Automation",
    "estrategia": "Strategy",
    "liderazgo": "Leadership", "leadership": "Leadership",
    "audiencias": "Audience Measurement",
    "research": "Research", "investigación": "Research", "investigacion": "Research",
    "insights": "Insights",
    "ux": "UX Research",
    "nps": "NPS / Customer Experience",
    "cx": "NPS / Customer Experience",
    "contenido": "Content Strategy",
    "client management": "Client Management",
    "account management": "Account Management",
    "procesos": "Process Design",
    "process design": "Process Design",
    "metodología": "Methodology", "metodologia": "Methodology",
    "project management": "Project Management",
    "project mgmt": "Project Management",
    "storytelling": "Storytelling",
    "python": "Python",
    "data pipelines": "Data Pipelines",
    "brand lift": "Brand Lift",
    "atribución": "Attribution", "atribucion": "Attribution",
    "hispanics": "Hispanic Market",
    "lanzamiento": "Product Launch",
    "inteligencia competitiva": "Competitive Intelligence",
    "cuentas": "Account Management",
    "multipaís": "Multi-Country", "multipais": "Multi-Country",
    "transformación digital": "Digital Transformation",
    "transformacion digital": "Digital Transformation",
    "consultoría": "Consulting", "consultoria": "Consulting",
    "estrategia de contenido": "Content Strategy",
    "mixed-methods": "Mixed Methods",
    "programación tv": "TV Programming", "programacion tv": "TV Programming",
    "retención": "Retention", "retencion": "Retention",
    "diseño de estudios": "Study Design", "diseno de estudios": "Study Design",
    "account growth": "Account Growth",
    "ventas consultivas": "Consultative Sales",
    "productividad": "Productivity",
    "customer insights": "Customer Insights",
    "kpi": "KPIs", "kpis": "KPIs",
    "análisis": "Analysis", "analisis": "Analysis",
    "multidisciplinario": "Cross-functional",
    "retail analytics": "Retail Analytics",
    "behavioral data": "Behavioral Data",
    "strategy": "Strategy",
    "budget": "Budget Management",
    "vendor mgmt": "Vendor Management",
    "vendor management": "Vendor Management",
    "photography": "Photography",
    "visual design": "Visual Design",
    "creative production": "Creative Production",
    "streaming": "Streaming",
}

CATEGORIA_MAP = {
    # Standardize categories to 8 canonical buckets
    "responsabilidad": "Responsibility",
    "innovación / ia": "Innovation",
    "innovacion / ia": "Innovation",
    "innovación / plataformas": "Innovation",
    "innovacion / plataformas": "Innovation",
    "innovación / metodología": "Innovation",
    "innovacion / metodologia": "Innovation",
    "innovación": "Innovation",
    "innovacion": "Innovation",
    "automatización / ia": "Innovation",
    "automatizacion / ia": "Innovation",
    "estrategia": "Strategy",
    "estrategia / arquitectura": "Strategy",
    "estrategia de insights": "Strategy",
    "estrategia de carrera": "Career Strategy",
    "impacto de negocio": "Business Impact",
    "crecimiento / revenue": "Business Impact",
    "crecimiento / retención": "Business Impact",
    "crecimiento / retencion": "Business Impact",
    "ventas / growth": "Business Impact",
    "eficiencia operacional": "Operational Excellence",
    "excelencia operativa": "Operational Excellence",
    "liderazgo": "Leadership",
    "liderazgo de investigación": "Leadership",
    "liderazgo de investigacion": "Leadership",
    "gestión de área": "Leadership",
    "gestion de area": "Leadership",
    "project leadership": "Leadership",
    "gestión de clientes": "Client Management",
    "gestion de clientes": "Client Management",
    "account management": "Client Management",
    "research": "Research",
    "ux / cx": "Innovation",
    "fotografía comercial": "Early Career",
    "fotografia comercial": "Early Career",
    "producción creativa": "Early Career",
    "produccion creativa": "Early Career",
    "soft skills origen": "Early Career",
}


def normalize_tag(tag, mapping, default_pass=True):
    if not tag:
        return tag
    key = tag.strip().lower()
    if key in mapping:
        return mapping[key]
    return tag.strip() if default_pass else tag


def normalize_tag_cell(cell_val, mapping):
    if not cell_val:
        return cell_val
    raw = [t.strip() for t in str(cell_val).replace(";", ",").split(",") if t.strip()]
    out = []
    seen = set()
    for t in raw:
        nt = normalize_tag(t, mapping)
        if nt.lower() not in seen:
            seen.add(nt.lower())
            out.append(nt)
    return ", ".join(out)


# ═══ ENRIQUECIMIENTO DE LOGROS (Seniority Fit · Story · Pair With) ═════════
# logro_id → (seniority_fit, story_evidence, pair_with_ids)
LOGRO_ENRICH = {
    # ── Canela Media (1) ────────────────────────────────────────────────
    1: ("director,exec", "Equipo de 12 analistas; antes tomaba 5 días por reporte, ahora 4 horas.",
        "51, 60"),
    2: ("director,exec", "Cliente Fortune 500 pedía insights de campañas; bajamos de 3 semanas a 5 días.",
        "1, 51"),
    3: ("director,exec", "Arquitectura usada por 8+ clientes Fortune 500 (incluye Procter, Coca-Cola, etc).",
        "51, 65"),
    4: ("senior,director", "Equipo bilingüe USA/MX, perfiles cuali + data scientist + media analyst.",
        "9, 54"),
    5: ("director,exec", "Canela pasó de player desconocido a referencia citada en pitch decks de competencia.",
        "3"),
    50: ("director,exec", "NPS de clientes pasó de 60 a 78 en 14 meses tras rediseño AI-driven de entregables.",
        "18, 63"),
    51: ("director,exec", "Combinación de PowerAutomate + Claude + DALL-E para reportes visuales auto-generados.",
        "1, 60"),
    60: ("senior,director", "Stack interno: ChatGPT Plus + Claude Pro + GPT-4 API, con prompts versionados en GitHub.",
        "1, 51"),

    # ── Yoor Company (2) ────────────────────────────────────────────────
    18: ("director,exec", "Fundé Yoor Company como CEO; cliente principal fue Coppel para investigación de UX bancario.",
        "63"),
    63: ("director,exec", "Estudio NPS pre/post rediseño UX; medición trimestral durante 6 meses.",
        "18"),

    # ── Altazor Intelligence (3) ────────────────────────────────────────
    64: ("director", "Pipeline para clientes de medios: ingestión → análisis IA → dashboard. POC a producción en 8 semanas.",
        "65, 66"),
    65: ("director", "KPIs predictivos para Televisa y otros: predicción de share basada en estudios de contenido.",
        "64"),
    66: ("senior,director", "Coordinación de equipos de 3 empresas en proyectos compartidos (product, data, research).",
        "64"),

    # ── Televisa Director (4) ──────────────────────────────────────────
    6: ("director,exec", "Investigación que llevó a cambio de horarios prime-time; medición vía Nielsen y panel propio.",
        "8, 52"),
    7: ("director,exec", "Estudios cuali con segmento 18-25; recomendaciones aplicadas en Las Estrellas y BLIM.",
        "8, 52"),
    8: ("director,exec", "Lideré research de pre-lanzamiento BLIM (2016-17) y de fortalecimiento Las Estrellas.",
        "6, 9, 54"),
    9: ("director,exec", "Marco metodológico que unificó research social + audiencias + digital para todo Grupo Televisa.",
        "8, 54"),
    52: ("director,exec", "Estudios juveniles aplicados a contenido Hoy y telenovelas digitales.",
        "7, 54"),
    54: ("director,exec", "Mezcla 60% cuali, 25% encuestas, 15% behavioral data.", "8, 9"),

    # ── Kantar / Millward Brown (5) ─────────────────────────────────────
    10: ("senior,director", "Rediseño de Brand Lift reports con storytelling más visual; clientes citan menos 'no entiendo'.",
        "55, 56"),
    11: ("senior,director", "Protocolo de re-engagement a 30/60/90 días con cuentas en riesgo.",
        "55, 56"),
    12: ("senior,director", "Cuentas: Coca-Cola, Procter, Telcel, Walmart, Bimbo.", "10, 55"),
    55: ("senior,director", "FMCG portfolio: Unilever, P&G, Coca-Cola, Bimbo.", "10, 11, 56"),
    56: ("senior,director", "Crecimiento orgánico de cuentas existentes via cross-sell de servicios Kantar.",
        "10, 55"),

    # ── TIBA Mkt/BI (11) ────────────────────────────────────────────────
    25: ("senior", "TIBA dependía de TI fragmentado; armé portafolio de gestión de infraestructura como producto.",
        "28, 31"),
    26: ("senior", "Análisis sectorial de gestión TIC en MX 2012-2014; identifiqué nichos no cubiertos.",
        "25"),
    27: ("senior,director", "Reportes ejecutivos para CEO KIO + board directors trimestralmente.",
        "14"),
    28: ("senior", "TIBA pasó de ser percibido como 'el spin-off' a marca con identidad propia frente a KIO.",
        "29, 30"),
    29: ("senior,director", "Definí mensajes, segmentación, sales materials y website para servicios B2B de gestión TIC.",
        "25, 28"),
    30: ("senior", "Newsletter interno, town halls, swag con valores corporativos.", "28"),
    31: ("senior,director", "Decks presentados al Consejo Directivo de KIO Networks (incluyendo Sergio Rosengaus).",
        "29"),

    # ── TIBA Sales (10) ─────────────────────────────────────────────────
    13: ("senior", "Combinación de prospección outbound + inteligencia competitiva sobre precios y contratos.",
        "22, 24, 58"),
    14: ("senior,director", "Mensual: reportes para Sergio Rosengaus (CEO) y board; trimestral: presentaciones.",
        "27, 31"),
    22: ("senior", "Ganamos contratos vs HP, IBM Global Services, Indra en licitaciones de gestión TIC.",
        "24, 58"),
    23: ("senior", "Soluciones tailored vs paquetes genéricos de competencia.", "22, 24"),
    24: ("senior", "Tradujimos pain points específicos en propuestas técnicas + comerciales.",
        "22, 23"),
    58: ("senior,director", "Pipeline duplicado año tras año por 3 años seguidos (2012-2014).",
        "13, 22"),

    # ── Nodo Research (8) ───────────────────────────────────────────────
    15: ("senior,director",
         "Aplicamos métodos UX cuando casi nadie en MX investigaba sitios web; clientes como Telmex, Bancomer.",
         "32, 34, 35"),
    16: ("senior,director", "Estudios en MX, Brasil, Chile, Argentina, Colombia, USA Hispanic, España.",
         "34"),
    32: ("senior,director", "Primer laboratorio UX con eye-tracker en MX; clientes pidieron capacitación interna.",
         "15, 35"),
    33: ("senior,director", "Presentaciones para Bancomer, Telmex, Walmart, P&G en sesiones de C-level.",
         "35"),
    34: ("senior,director", "Coordiné equipos locales en cada mercado; metodología propia documentada.",
         "16"),
    35: ("senior", "Cientos de focus groups y sesiones cuali en 5 años; especialidad UX/digital/social.",
         "15, 33"),

    # ── Independiente / BRICKO / EIDOS (6) ──────────────────────────────
    17: ("senior,director", "BRICKO Tequila y EIDOS Marketing fueron clientes principales 2020-2021 durante pandemia.",
         "19, 21, 61"),
    61: ("senior,director", "Cumplimiento perfecto entre 8 proyectos consecutivos con 5 clientes distintos.",
         "17, 19"),

    # ── Independiente 2014 (9) ──────────────────────────────────────────
    19: ("senior", "Período de transición post-TIBA; consultoría freelance breve.", "20, 21"),
    20: ("senior", "Cada cliente recibió enfoque distinto: B2C vs B2B, regional vs nacional.", "19, 21"),
    21: ("senior", "Equipos ad-hoc según necesidad; trabajo con agencias como external strategist.", "19, 20"),

    # ── Xabón'art (14) ──────────────────────────────────────────────────
    36: ("entry,mid", "Microempresa propia de artesanía; jabones y cosmética natural en mercados orgánicos.",
         "37, 38"),
    37: ("entry,mid", "Diseño visual + producción artesanal + venta en mercados como Bazar del Sábado.", "36, 38"),
    38: ("entry,mid", "Ciclo completo: formulación → producción → empaque → venta presencial.", "36, 37"),

    # ── Focus Vision (15) ───────────────────────────────────────────────
    39: ("entry,mid", "FocusVision era líder mundial en streaming de focus groups; soporte 24/7 a clientes globales.",
         "40"),
    40: ("entry,mid", "Conexiones internacionales en horarios variables; resolución técnica en vivo.", "39"),

    # ── PTA/GBA productoras (16) ────────────────────────────────────────
    41: ("entry", "Asistente en producciones para Coca-Cola, Telmex, Bimbo; aprendizaje de set y dirección.",
         "42"),
    42: ("entry", "Trabajo bajo directores reconocidos; aprendí storyboard, casting, dirección de arte.", "41"),

    # ── Onírica (17) ────────────────────────────────────────────────────
    43: ("entry,mid", "Productora propia; clientes corporativos + producción audiovisual + websites.",
         "44, 45"),
    44: ("entry,mid",
         "APEC México 2002: coordiné logística y producción audiovisual de presentaciones para delegaciones.",
         "43, 45"),
    45: ("entry,mid", "Emprendimiento creativo; aprendí gestión de proyectos creativos end-to-end.", "43, 44"),

    # ── Central Reg (18) ────────────────────────────────────────────────
    46: ("entry", "A los 16-17 años, primer empleo formal; reparé hardware empresarial.", "47, 48"),
    47: ("entry", "Aprendí diagnóstico de hardware, electrónica básica, atención a clientes empresariales.",
         "46, 48"),
    48: ("entry", "Cliente típico: supermercados, restaurantes, papelerías con sistemas POS.", "46, 47"),

    # ── Walmart Scintilla (19) ──────────────────────────────────────────
    67: ("senior", "Cuentas CPG en Scintilla / Walmart Data Ventures; insights de purchase behavior real.",
         "68"),
    68: ("senior,director",
         "Decisión deliberada de bajar de Director a Account Manager para acceder a behavioral data a escala retail.",
         "67"),

    # ── Nuevos (69-74) ──────────────────────────────────────────────────
    69: ("director,exec", "Diseñé proceso replicable para 3 áreas: telenovelas, deportes, noticias.", "70"),
    70: ("director,exec", "Presupuesto anual ~$2M MXN; 5 proveedores principales; KPIs trimestrales al board.",
         "69"),
    71: ("entry,mid", "Estudio Segarra: fotografía publicitaria; trabajé bajo Felipe Segarra.", ""),
    72: ("senior", "Question Mark fue startup boutique; clientes en Mkt research, branding, comunicación.", ""),
    73: ("entry,mid", "Producciones propias amateur 1996-2000; teatro infantil en escuelas.", "74"),
    74: ("entry,mid", "La experiencia escénica temprana hoy se traduce en presentaciones ejecutivas memorables.",
         "73"),
}


# ═══ NUEVO 00 INSTRUCCIONES PARA LLM ═══════════════════════════════════════
INSTRUCCIONES_LLM = [
    # (col_A_label_or_section, col_B_value_or_text_full)
    ("title", "CV MASTER · RAÚL MERCADO BUSTAMANTE — MANUAL OPERATIVO PARA LLM"),
    ("section", "PROPÓSITO"),
    ("text",
     "Este archivo es la BASE DE DATOS MAESTRA de la trayectoria profesional de Raúl Mercado. "
     "No es un CV terminado. Es la fuente desde la cual un LLM (Claude, GPT-4, etc.) construye "
     "CVs especializados y customizados para cada oportunidad. El usuario te lo dará como input "
     "junto con instrucciones específicas (industria target, idioma, nivel de puesto)."),
    ("section", "REGLAS CRÍTICAS — NO LAS VIOLES"),
    ("rule",
     "1. NO INVENTES MÉTRICAS. Si la celda 'Métrica / KPI' está vacía, no agregues números "
     "que no estén ya en el texto del logro. Si el logro no tiene número, redactalo cualitativo."),
    ("rule",
     "2. NO INVENTES FECHAS, EMPRESAS NI PUESTOS. Solo lo que aparece en sheet '02 Experiencia'."),
    ("rule",
     "3. RESPETA LOS TOGGLES 'Incluir CV ES/EN' en sheets 03, 05, 06, 07, 10. Si está 'No', "
     "ese item NO va al CV final salvo que el usuario lo pida explícitamente."),
    ("rule",
     "4. USA EL IDIOMA correcto: si el CV es para empresa USA, usa columna EN. Si LATAM, ES. "
     "NUNCA mezcles ambos en el output final."),
    ("rule",
     "5. ATIENDE 'Seniority Fit' en 03 Logros: si el puesto target es 'Manager', evita logros "
     "marcados solo como 'exec'. Si es 'Director/VP', evita los marcados solo como 'entry'."),
    ("rule",
     "6. USA SOLO LA SECCIÓN MÁS RECIENTE Y RELEVANTE. Para CV ejecutivo, omite trabajos "
     "anteriores a 2004 salvo que aporten narrativa (ej. emprendedor desde joven)."),
    ("rule",
     "7. LONGITUD: CVs profesionales 1-2 páginas. Selecciona 4-6 logros por puesto reciente, "
     "2-3 por puestos viejos. Total máximo ~15-20 logros en un CV."),
    ("section", "MAPA DE SHEETS"),
    ("sheet", "00 Instrucciones · ESTA SHEET. Manual operativo."),
    ("sheet", "01 Perfil · Datos contacto + bio ES/EN + headline + keywords ATS."),
    ("sheet", "02 Experiencia · Un registro por puesto. ID es la clave foránea para sheet 03."),
    ("sheet", "03 Logros · ⭐ CORAZÓN DEL ARCHIVO. Filtra aquí por industria/skill/seniority."),
    ("sheet", "04 Skills · Categorizadas, con nivel y años. Columna 'Destacar en' indica tipos de CV."),
    ("sheet", "05 Educación · Formación. Solo incluye en CV las marcadas 'Incluir CV'."),
    ("sheet", "06 Certificaciones · Con vigencia y URL."),
    ("sheet", "07 Voluntariado · Marcadas con 'Incluir CV'."),
    ("sheet", "08 Empresas Target · Tracker propio del usuario. NO usar como output."),
    ("sheet", "09 Perfiles Narrativos · Bio de 120 palabras por tipo de empresa. ⭐ Cópialo a CV."),
    ("sheet", "10 Conferencias y Speaker · Charlas, podcasts, cátedras."),
    ("sheet", "11 Reconocimientos y Membresías · Premios y afiliaciones (SOGEM, MENSA)."),
    ("sheet", "12 Recipes · 🍳 Recetas pre-armadas de CVs por tipo de puesto/industria. Empieza aquí si dudas."),
    ("sheet", "13 Anti-Patterns · ❌ Qué NO decir, palabras a evitar, claims que sonarían falsos."),
    ("section", "VOCABULARIO CONTROLADO DE TAGS"),
    ("vocab_industria",
     "Industria (canonical): Media, Streaming, TV, Broadcasting, AdTech, Entertainment, "
     "CPG, Retail, Tech, Digital, Data, Telecom, Consulting, B2B, Startup, Arts, Global."),
    ("vocab_skill",
     "Skill (canonical): Research, Insights, Audience Measurement, UX Research, Mixed Methods, "
     "Brand Lift, Attribution, AI/ML, Generative AI, Automation, Data Pipelines, Python, "
     "Strategy, Content Strategy, Leadership, Account Management, Client Management, "
     "Project Management, Vendor Management, Storytelling, Consultative Sales, Growth, "
     "Retail Analytics, Behavioral Data, Hispanic Market."),
    ("vocab_categoria",
     "Categoría (canonical): Business Impact, Operational Excellence, Innovation, Strategy, "
     "Leadership, Client Management, Research, Responsibility, Career Strategy, Early Career."),
    ("section", "FLUJO RECOMENDADO PARA CONSTRUIR UN CV"),
    ("step",
     "PASO 1 · Lee la solicitud del usuario: ¿qué empresa? ¿qué puesto? ¿qué nivel? ¿qué idioma?"),
    ("step",
     "PASO 2 · Busca en sheet 12 Recipes una receta que se ajuste. Si existe, úsala como base."),
    ("step",
     "PASO 3 · Selecciona perfil narrativo de sheet 09 según tipo de empresa."),
    ("step",
     "PASO 4 · Filtra logros en sheet 03: por Tags Industria/Skill, Seniority Fit, e 'Incluir CV'. "
     "Selecciona 4-6 por puesto reciente, 2-3 por puestos antiguos."),
    ("step",
     "PASO 5 · Para cada logro, si hay 'Pair With', considera incluir esos logros emparejados (cuentan mejor historia juntos)."),
    ("step",
     "PASO 6 · Selecciona 8-12 skills de sheet 04, priorizando las marcadas 'Destacar en' "
     "para el tipo de empresa target."),
    ("step",
     "PASO 7 · Agrega educación + certificaciones + reconocimientos relevantes (los marcados 'Incluir CV')."),
    ("step",
     "PASO 8 · Verifica contra sheet 13 Anti-Patterns que no estés usando lenguaje débil o claims problemáticos."),
    ("step",
     "PASO 9 · Output en formato pedido (Markdown, PDF-ready, plain text, JSON)."),
    ("section", "COMUNICACIÓN CON EL USUARIO"),
    ("text",
     "Cuando entregues un CV, incluye al final un breve resumen de las decisiones que tomaste: "
     "qué logros seleccionaste y por qué, qué dejaste fuera, y qué te gustaría confirmar con el usuario "
     "(ej: '¿Quieres mencionar la pausa 2020-2021 como sabático o como consultoría independiente?')."),
    ("section", "ÚLTIMA ACTUALIZACIÓN DEL MASTER"),
    ("text",
     "2026-05-17 · v8 LLM-ready · 12 sheets · 24 puestos · 69 logros · 19 skills · vocabulario controlado."),
]


# ═══ RECIPES (sheet 12) ════════════════════════════════════════════════════
RECIPES = [
    # (recipe_id, name, target, perfil_narrativo_idx, logro_ids, skill_categorias, notas)
    ("R01", "Director of Research & Insights · US Streaming/Media",
     "Empresas tipo Netflix, Hulu, Disney+, Paramount, NBCU para puestos de Director/Sr Director de Research.",
     "Streaming USA",
     "1, 3, 4, 5, 50, 51, 60, 8, 9, 54, 52, 6, 7, 55, 56, 64, 65",
     "Research, Analytics / IA, Estrategia, Liderazgo",
     "Idioma: EN. Énfasis: AI applications, scalable insight systems, audience measurement, Hispanic market. "
     "Incluir Yoor (founder mindset). Omitir trabajos pre-2007."),

    ("R02", "Research Director · LATAM Media Group",
     "Televisa-like, TelevisaUnivision, Grupo Salinas, Caracol, Globo, Cinemex, TV Azteca.",
     "Streaming USA",  # adapt
     "8, 9, 54, 52, 6, 7, 1, 3, 4, 50, 51, 64, 65, 15, 16, 33, 34",
     "Research, Estrategia, Liderazgo",
     "Idioma: ES. Énfasis: experiencia Televisa, conocimiento mercado mexicano, "
     "amplitud cuali+cuanti, integración con áreas creativas. Mencionar BLIM/Las Estrellas."),

    ("R03", "Sr Insights Manager · CPG Global Brand",
     "P&G, Unilever, Coca-Cola, Nestlé, Pepsico, Diageo, Bimbo.",
     "Streaming USA",  # we don't have a CPG-specific narrative yet
     "55, 56, 10, 11, 12, 1, 3, 50, 51, 64, 65, 67, 13, 22",
     "Research, Analytics / IA, Liderazgo",
     "Idioma: EN/ES según país. Énfasis: experiencia Kantar/Millward Brown con FMCG, "
     "Walmart Scintilla retail analytics (purchase behavior real), AI for client portfolios. "
     "Mencionar cuentas globales gestionadas."),

    ("R04", "Head of Research / VP Insights · Tech/AdTech Startup",
     "TripleLift, Captify, Samba TV, Comscore, Nielsen IQ.",
     "Streaming USA",
     "1, 3, 4, 5, 50, 51, 60, 64, 65, 66, 18, 63, 52, 8, 25, 26",
     "Research, Analytics / IA, Estrategia, Liderazgo",
     "Idioma: EN. Énfasis: AI-native research, scalable systems, founder mindset (Yoor, Xabón'art, Onírica), "
     "consultoría estratégica. Mencionar Pioneer in UX in Mexico (L32) para credibilidad técnica."),

    ("R05", "Senior Strategic Consultant · Research/Innovation Firm",
     "Kantar, IPSOS, Nielsen, GfK, Ascential, Bain, McKinsey Insights practice.",
     "Streaming USA",
     "3, 4, 5, 9, 54, 64, 65, 66, 17, 61, 55, 56, 27, 31, 14, 33, 15",
     "Research, Estrategia, Liderazgo",
     "Idioma: EN/ES. Énfasis: capacidad de board-level reporting (L14, L27, L31, L33), "
     "transformación de áreas (L9, L29), liderazgo internacional multipaís (L16, L34)."),

    ("R06", "Account Manager / Sr Account Manager · Retail Analytics Platform",
     "Para roles AM/Sr AM en plataformas tipo Walmart Scintilla, Amazon Marketing Cloud, Criteo, dunnhumby.",
     "Streaming USA",
     "67, 68, 13, 22, 23, 24, 25, 12, 55, 56, 10, 11, 1, 50, 63, 18",
     "Research, Estrategia, Liderazgo, Analytics / IA",
     "Idioma: EN. Énfasis: combinación rara de research depth + commercial acumen + AI fluency + retail data. "
     "Mencionar la decisión deliberada de L68 si el reclutador pregunta por qué bajó de Director."),
]


# ═══ ANTI-PATTERNS (sheet 13) ══════════════════════════════════════════════
ANTI_PATTERNS = [
    # (id, categoria, evita, en_lugar_usa, razon)
    ("A01", "Métricas falsas",
     "Inventar porcentajes o cifras como '20%', '50%', '$1M' cuando no están en el master.",
     "Si no hay número en la celda 'Métrica / KPI' del logro, usa lenguaje cualitativo: "
     "'significativamente', 'medible', 'sostenido'.",
     "Cualquier número inventado puede ser refutado en entrevista o background check."),

    ("A02", "Lenguaje débil",
     "'Ayudé a', 'Participé en', 'Apoyé en', 'Colaboré con', 'Estuve involucrado en', "
     "'Helped', 'Assisted', 'Contributed to', 'Was involved in'.",
     "Verbos de acción fuertes: 'Lideré', 'Diseñé', 'Implementé', 'Construí', "
     "'Led', 'Designed', 'Built', 'Drove', 'Owned'.",
     "Los verbos débiles diluyen el ownership y suenan a 'estuve en la reunión'."),

    ("A03", "Buzzwords vacíos",
     "'Sinergia', 'paradigma', 'disruptivo', 'unicornio', 'rockstar', 'ninja', "
     "'pasionado de', 'innovador', 'orientado a resultados'.",
     "Sé concreto: en lugar de 'orientado a resultados', describe el resultado. "
     "En lugar de 'innovador', cita el output innovador (ej. 'primer framework UX en MX').",
     "Los reclutadores escanean y descartan estos términos por sobreuso."),

    ("A04", "Sobre-claim de IA",
     "Decir 'experto en IA', 'machine learning engineer', 'data scientist' si no son el core del puesto.",
     "Posiciónate como practitioner: 'aplico LLMs (Claude, GPT-4) en flujos de research', "
     "'integro IA en procesos de insights'. Sé específico sobre las herramientas reales que usas.",
     "Los hiring managers técnicos detectan al instante claims inflados. Tu valor real es "
     "research + AI applied, no AI engineering."),

    ("A05", "Mezclar idiomas en CV",
     "Headers en español con bullets en inglés. Bio EN seguida de skills ES.",
     "Comprométete con UN solo idioma por CV. El idioma se decide por el país del empleador.",
     "Mezclar señala falta de cuidado y dificulta lectura ATS/recruiter."),

    ("A06", "Listar todas las herramientas",
     "Listar Excel, Word, PowerPoint, Outlook, Google Docs, Slack, Zoom, Teams, etc.",
     "Solo herramientas diferenciadoras: SQL, Python, Looker, Power BI, Tableau, Figma, "
     "Salesforce, R, plataformas específicas de research (Qualtrics, Quirks, Confirmit, Recollective).",
     "Herramientas básicas son commodity; las específicas son señal real de capability."),

    ("A07", "Justificar gaps con palabras vagas",
     "'Período de reflexión', 'sabático', 'búsqueda activa' para el gap 2020-2021.",
     "Si te preguntan por el gap 2020-21: 'Durante la pandemia, consultoría independiente "
     "con BRICKO y EIDOS apoyando transformación digital de PYMEs' (es la verdad, está en el master).",
     "Hay un logro real para ese período (L17, L61). Úsalo."),

    ("A08", "Tono pasivo / impersonal",
     "'Se diseñó', 'Fue implementado', 'Was led by team', 'Was developed'.",
     "Primera persona implícita con verbos en pasado: 'Diseñé', 'Implementé', 'Led', 'Built'. "
     "Es CV tuyo, no de un comité.",
     "El pasivo oculta agencia y hace dudar de ownership."),

    ("A09", "Pedir más de lo que dice el master para puesto target",
     "Inflar Walmart Scintilla a 'Director' o agregar People Manager si no lo es.",
     "Refleja el puesto real. La narrativa de L68 (career choice) explica el step laterallateral si surge en entrevista.",
     "Cualquier inflación se detecta en LinkedIn/referencias y mata credibilidad."),

    ("A10", "Olvidar el contexto del lector",
     "Usar acrónimos sin explicar: 'BLIM', 'SOGEM', 'APEC', 'CPG' en CVs para audiencias no familiares.",
     "Primera mención: explica brevemente. 'BLIM (Televisa's streaming platform)', "
     "'SOGEM (Mexican Writers Society)'. Después, abrevia.",
     "Acrónimos sin contexto frustran al lector y señalan que escribes para ti, no para él."),

    ("A11", "Listar exhaustivamente todas las cuentas",
     "'Trabajé con Coca-Cola, Pepsi, P&G, Unilever, Nestlé, Bimbo, Telcel, AT&T, Walmart, Soriana, Chedraui...'",
     "Selecciona 4-5 cuentas reconocibles y diversas: 'Brands like Coca-Cola, P&G, Walmart and BLIM '.",
     "Listas largas se ven como relleno; selección curada se ve como pensamiento."),

    ("A12", "Logros sin contexto",
     "'Reduje cancelaciones 5%' sin decir desde qué base.",
     "'Reduje cancelaciones de proyectos de 18% a <5% mediante…'. Cuando el master tiene la cifra base, úsala.",
     "Sin base, el número no transmite magnitud."),

    ("A13", "Mencionar edad o años exactos de carrera",
     "'30 años de experiencia', 'desde 1994', 'soy de los 70s'.",
     "'20+ years across…', 'Two decades of…', sin atar a edad. Bias etario es real.",
     "Salvo que la longevidad sea diferenciadora positiva (rara vez), atar a edad es ATS-friendly y bias-prone."),
]


# ═══ HELPER STYLES ═════════════════════════════════════════════════════════
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor=BLUE_100)
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor=GREEN_100)
RULE_FONT = Font(name="Calibri", size=11, bold=True, color="9B1B1B")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=GREEN_100)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_instrucciones(ws):
    """Wipe + rewrite 00 Instrucciones."""
    # Unmerge first (delete_rows doesn't clean merged ranges properly)
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))
    # Clear existing
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)

    row = 1
    for kind, text in INSTRUCCIONES_LLM:
        if kind == "title":
            ws.cell(row=row, column=1, value=text)
            ws.cell(row=row, column=1).font = TITLE_FONT
            ws.cell(row=row, column=1).fill = TITLE_FILL
            ws.cell(row=row, column=1).alignment = CENTER
            ws.row_dimensions[row].height = 30
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        elif kind == "section":
            row += 1  # blank row before section
            ws.cell(row=row, column=1, value=text)
            ws.cell(row=row, column=1).font = SECTION_FONT
            ws.cell(row=row, column=1).fill = SECTION_FILL
            ws.cell(row=row, column=1).alignment = CENTER
            ws.row_dimensions[row].height = 24
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        elif kind in ("text", "rule", "step", "vocab_industria", "vocab_skill", "vocab_categoria"):
            ws.cell(row=row, column=1, value=text)
            ws.cell(row=row, column=1).alignment = LEFT_WRAP
            if kind == "rule":
                ws.cell(row=row, column=1).font = RULE_FONT
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            # Set row height heuristic
            est_lines = max(2, len(text) // 80 + 1)
            ws.row_dimensions[row].height = 15 * est_lines
        elif kind == "sheet":
            # Format: "10 Conferencias · texto"
            parts = text.split(" · ", 1)
            ws.cell(row=row, column=1, value=parts[0] if parts else text)
            ws.cell(row=row, column=1).font = Font(bold=True, color=BLUE_100)
            if len(parts) > 1:
                ws.cell(row=row, column=2, value=parts[1])
                ws.cell(row=row, column=2).alignment = LEFT_WRAP
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

    # Column widths
    widths = [32, 28, 28, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def add_columns_to_logros(ws):
    """Add Seniority Fit, Story / Evidence, Pair With columns. Compute values."""
    headers = [c.value for c in ws[2]]
    # Find insertion point (before "Incluir CV ES" if exists)
    incluir_idx = None
    for i, h in enumerate(headers, start=1):
        if h == "Incluir CV ES":
            incluir_idx = i
            break
    if incluir_idx is None:
        incluir_idx = ws.max_column + 1

    # Insert 3 columns at position incluir_idx
    ws.insert_cols(incluir_idx, amount=3)
    ws.cell(row=2, column=incluir_idx, value="Seniority Fit")
    ws.cell(row=2, column=incluir_idx + 1, value="Story / Evidence")
    ws.cell(row=2, column=incluir_idx + 2, value="Pair With (IDs)")

    # Style headers
    for off in range(3):
        cell = ws.cell(row=2, column=incluir_idx + off)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    # Fill values per logro ID
    for r in range(3, ws.max_row + 1):
        lid = ws.cell(row=r, column=1).value
        if lid is None:
            continue
        try:
            lid_int = int(lid)
        except (ValueError, TypeError):
            continue
        enrich = LOGRO_ENRICH.get(lid_int)
        if enrich:
            sen, story, pair = enrich
        else:
            sen, story, pair = "all", "", ""
        ws.cell(row=r, column=incluir_idx, value=sen)
        ws.cell(row=r, column=incluir_idx + 1, value=story)
        ws.cell(row=r, column=incluir_idx + 2, value=pair)
        for off in range(3):
            ws.cell(row=r, column=incluir_idx + off).alignment = LEFT_WRAP

    # Column widths
    ws.column_dimensions[get_column_letter(incluir_idx)].width = 22
    ws.column_dimensions[get_column_letter(incluir_idx + 1)].width = 52
    ws.column_dimensions[get_column_letter(incluir_idx + 2)].width = 14


def normalize_logro_tags(ws):
    """Normalize Tags Industria, Tags Habilidad, Categoría in 03 Logros."""
    headers = [c.value for c in ws[2]]
    idx_ind = headers.index("Tags Industria") + 1 if "Tags Industria" in headers else None
    idx_skill = headers.index("Tags Habilidad") + 1 if "Tags Habilidad" in headers else None
    idx_cat = headers.index("Categoría") + 1 if "Categoría" in headers else None

    n_changes = 0
    for r in range(3, ws.max_row + 1):
        if idx_ind:
            v = ws.cell(row=r, column=idx_ind).value
            nv = normalize_tag_cell(v, INDUSTRIA_MAP)
            if nv != v:
                ws.cell(row=r, column=idx_ind, value=nv)
                n_changes += 1
        if idx_skill:
            v = ws.cell(row=r, column=idx_skill).value
            nv = normalize_tag_cell(v, SKILL_MAP)
            if nv != v:
                ws.cell(row=r, column=idx_skill, value=nv)
                n_changes += 1
        if idx_cat:
            v = ws.cell(row=r, column=idx_cat).value
            if v:
                key = str(v).strip().lower()
                if key in CATEGORIA_MAP:
                    ws.cell(row=r, column=idx_cat, value=CATEGORIA_MAP[key])
                    n_changes += 1
    return n_changes


def create_recipes_sheet(wb):
    if "12 Recipes" in wb.sheetnames:
        del wb["12 Recipes"]
    ws = wb.create_sheet("12 Recipes", index=wb.sheetnames.index("11 Reconocimientos y Membresías") + 1)
    ws.cell(row=1, column=1,
            value="12 RECIPES · Plantillas pre-armadas de CVs por tipo de puesto")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).fill = TITLE_FILL
    ws.cell(row=1, column=1).alignment = CENTER
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:G1")

    headers = ["Recipe ID", "Nombre / Target Role", "Cuándo usar",
               "Perfil Narrativo (idx)", "Logros a usar (IDs)",
               "Skills (categorías)", "Notas críticas para el LLM"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    ws.row_dimensions[2].height = 30

    for r_idx, rec in enumerate(RECIPES, start=3):
        for c_idx, val in enumerate(rec, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = LEFT_WRAP

    widths = [10, 38, 38, 18, 50, 28, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    ws.sheet_properties.tabColor = SPARK_100


def create_antipatterns_sheet(wb):
    if "13 Anti-Patterns" in wb.sheetnames:
        del wb["13 Anti-Patterns"]
    ws = wb.create_sheet("13 Anti-Patterns", index=wb.sheetnames.index("12 Recipes") + 1)
    ws.cell(row=1, column=1,
            value="13 ANTI-PATTERNS · Qué NO hacer al construir un CV desde este master")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).fill = TITLE_FILL
    ws.cell(row=1, column=1).alignment = CENTER
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:E1")

    headers = ["ID", "Categoría", "Evita esto", "En lugar usa", "Por qué"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    ws.row_dimensions[2].height = 30

    for r_idx, ap in enumerate(ANTI_PATTERNS, start=3):
        for c_idx, val in enumerate(ap, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = LEFT_WRAP

    widths = [6, 22, 55, 55, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    ws.sheet_properties.tabColor = "9B1B1B"  # dark red (anti-pattern signal)


def main():
    print(f"📂 Master: {MASTER}")
    if not BACKUP.exists():
        shutil.copy2(MASTER, BACKUP)
        print(f"🛟 Backup creado: {BACKUP.name}")

    wb = load_workbook(str(MASTER))

    # ── 1. Reescribir 00 Instrucciones ──
    print("\n[1] Reescribiendo 00 Instrucciones como manual LLM…")
    write_instrucciones(wb["00 Instrucciones"])
    print(f"    ✅ {len(INSTRUCCIONES_LLM)} bloques escritos")

    # ── 2. Agregar columnas a 03 Logros ──
    print("\n[2] Agregando columnas Seniority Fit · Story · Pair With a 03 Logros…")
    add_columns_to_logros(wb["03 Logros"])
    enriched = sum(1 for lid in LOGRO_ENRICH if LOGRO_ENRICH[lid][1])
    print(f"    ✅ {len(LOGRO_ENRICH)} logros con seniority asignado, {enriched} con story")

    # ── 3. Normalizar tags ──
    print("\n[3] Normalizando tags a vocabulario controlado…")
    n_changes = normalize_logro_tags(wb["03 Logros"])
    print(f"    ✅ {n_changes} celdas de tags normalizadas")

    # ── 4. Crear sheet 12 Recipes ──
    print("\n[4] Creando sheet 12 Recipes…")
    create_recipes_sheet(wb)
    print(f"    ✅ {len(RECIPES)} recipes agregadas")

    # ── 5. Crear sheet 13 Anti-Patterns ──
    print("\n[5] Creando sheet 13 Anti-Patterns…")
    create_antipatterns_sheet(wb)
    print(f"    ✅ {len(ANTI_PATTERNS)} anti-patterns agregadas")

    # Sheet activa al abrir: 00 (para que un humano vea el manual)
    wb.active = wb.sheetnames.index("00 Instrucciones")

    wb.save(str(MASTER))
    print(f"\n✅ Master LLM-ready guardado")
    print(f"\n📊 Sheets finales:")
    for s in wb.sheetnames:
        print(f"   • {s}")


if __name__ == "__main__":
    main()
