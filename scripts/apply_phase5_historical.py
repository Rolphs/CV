"""
Fase 5 · Aplicar hallazgos históricos (2010-2022) al master.

Cambios:
  - 02 Experiencia: +8 puestos (IDs 20-27)
      A1-A7: puestos pre-2007 / paralelos (Mago, Reforma, Cinemex,
             Trabulsi, Teatro, Segarra, Question Mark)
      B1:    subpuesto Televisa "Strategic Information Manager" 2017-2018
  - 05 Educación: +1 fila (Curso Literatura Latinoamericana UIA 1993, Dehesa)
  - 10 Conferencias y Speaker: nueva sheet con 8 entradas
  - 11 Reconocimientos y Membresías: nueva sheet con 2 entradas
"""
from __future__ import annotations
import shutil
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "data" / "master" / "cv_master_raul_mercado.xlsx"
BACKUP = ROOT / "data" / "master" / "backups" / "v5_pre_historical.xlsx"


# ─── Datos a insertar ───────────────────────────────────────────────────────
# 02 Experiencia · columnas:
# ID | Empresa | Industria | País | Puesto ES | Puesto EN | Nivel | Tipo |
# Fecha Inicio | Fecha Fin | ¿Actual? | Dept | Reporta a | Personas |
# Desc Empresa ES | Desc Empresa EN
NEW_POSITIONS = [
    (20, "Mago Ilusionista (Independiente)", "Entretenimiento / Eventos", "México",
     "Mago Ilusionista", "Magician / Illusionist", "Independiente", "Freelance",
     "1987", "1992", "No", "", "", "",
     "Presentaciones en fiestas infantiles y eventos privados. Primera experiencia profesional, desde la infancia.",
     "Performances at children's parties and private events. First professional experience, from childhood."),

    (21, "Periódico Reforma", "Medios / Prensa", "México",
     "Voceador (con Aura† y Dehesa†)", "Newspaper Crier (with Aura† and Dehesa†)",
     "Apoyo", "Tiempo parcial",
     "1994", "1994", "No", "Distribución", "", "",
     "Voceador del periódico Reforma, colaborando junto a figuras literarias de la talla de Alejandro Aura (†) y Germán Dehesa (†).",
     "Newspaper crier for Reforma, working alongside literary figures Alejandro Aura (†) and Germán Dehesa (†)."),

    (22, "Cinemex", "Entretenimiento / Cine", "México",
     "Staff", "Staff", "Operativo", "Tiempo parcial",
     "1997", "1997", "No", "Operaciones", "", "",
     "Personal operativo en cadena de salas de cine.",
     "Operational staff at movie theater chain."),

    (23, "Estudio Fotográfico Trabulsi", "Fotografía / Artes Visuales", "México",
     "Asistente de Fotógrafo", "Photographer Assistant", "Apoyo", "Tiempo parcial",
     "1998", "1999", "No", "", "Fotógrafo principal", "",
     "Asistencia técnica y creativa en sesiones fotográficas profesionales.",
     "Technical and creative assistance in professional photography sessions."),

    (24, "Dirección de Teatro (Independiente)", "Artes Escénicas", "México",
     "Director de Producción y Montaje", "Production & Staging Director",
     "Independiente", "Freelance",
     "1996", "2000", "No", "", "", "Variable según producción",
     "Producción y montaje de obras musicales y obras de teatro para público infantil.",
     "Production and staging of musical works and children's theater plays."),

    (25, "Estudio Segarra", "Fotografía Publicitaria", "México",
     "Fotógrafo", "Photographer", "Freelance", "Freelance",
     "02/2006", "03/2007", "No", "", "", "",
     "Fotografía de producto para clientes como Avón, Barcel y Bísquets Obregón.",
     "Product photography for clients including Avón, Barcel and Bísquets Obregón."),

    (26, "Question Mark (Marketing Studio)", "Marketing / Research", "México",
     "Project Leader", "Project Leader", "Senior", "Contrato",
     "06/2014", "10/2014", "No", "", "Lourdes Lamasney / Ian Reider", "",
     "Startup de Marketing Studio fundada por Lourdes Lamasney e Ian Reider. Liderazgo de proyectos de marketing e insights para clientes en diversos segmentos.",
     "Marketing Studio startup founded by Lourdes Lamasney and Ian Reider. Led marketing and insights projects for clients across multiple segments."),

    (27, "Televisa", "Medios / TV", "México",
     "Strategic Information Manager", "Strategic Information Manager",
     "Gerencia", "Tiempo completo",
     "08/2017", "04/2018", "No", "Investigación Social", "Dirección de Investigación", "Equipo cross-funcional",
     "Diseño e implementación del nuevo proceso de investigación social: desde la extracción de insights valiosos hasta la transferencia de conocimiento accionable. Gestión de proveedores, presupuesto, KPIs y proyectos. Precede al rol de Director de Investigación Social.",
     "Designed and implemented a new social research process: from extracting valuable insights to transferring actionable knowledge. Vendor management, budgeting, KPIs and project management. Preceded the Social Research Director role."),
]


# 05 Educación · columnas:
# Tipo | Institución | Título ES | Degree EN | Área | País | Año ini | Año fin |
# Estado | Relevancia
NEW_EDUCATION = [
    ("Curso / Formación Cultural", "Universidad Iberoamericana, CDMX",
     "Curso de Literatura Latinoamericana", "Latin American Literature Course",
     "Humanidades / Literatura", "México", 1993, 1993, "Completo",
     "Impartido por Germán Dehesa (†) — formación temprana en pensamiento crítico y narrativa cultural."),
]


# 10 Conferencias y Speaker
SPEAKER_HEADERS = [
    "Año", "Tipo", "Título ES", "Title EN", "Organización", "Audiencia",
    "Descripción ES", "Description EN", "Periodo", "Incluir CV ES", "Incluir CV EN"
]
SPEAKER_ROWS = [
    ("2013–actual", "Podcast (Host)", "La Máquina de las Ucronías", "La Máquina de las Ucronías",
     "Independiente", "Público general / Pensadores",
     "Podcast de exploración de historias alternativas, filosofía y ciencias sociales. Host desde 2013.",
     "Podcast exploring alternate histories, philosophy and social sciences. Host since 2013.",
     "Continuo", "Sí", "Sí"),

    (2016, "Seminario", "Story Telling para Investigadores", "Storytelling for Researchers",
     "AMAI (Asoc. Mexicana de Agencias de Investigación de Mercado)",
     "Investigadores de mercado",
     "Parte del programa de educación continua para profesionales de investigación de mercado.",
     "Part of the continuous education program for market research professionals.",
     "Único", "Sí", "Sí"),

    (2014, "Conferencia", "Realidades Virtuales", "Virtual Realities",
     "Textbook Writers Association", "Editores y autores",
     "Tendencias tecnológicas, hábitos de jóvenes y nuevas formas de vinculación adultos-niños.",
     "Tech trends, habits amongst youth and new ways of adult-children linking.",
     "Único", "Sí", "Sí"),

    (2013, "Cátedra Invitada", "Introducción al Análisis Cualitativo", "Introduction to Qualitative Analysis",
     "Universidad Anáhuac", "Estudiantes de maestría",
     "Parte del currículum de maestría en investigación.",
     "Part of master's degree curriculum in research.",
     "Único", "Sí", "Sí"),

    (2013, "Seminario", "Management 3.0", "Management 3.0",
     "Instituto Mexicano de Teleservicios (IMT)", "Ejecutivos",
     "Tendencias modernas de gestión y liderazgo.",
     "Modern management and leadership trends.",
     "Único", "Sí", "No"),

    (2012, "Conferencia", "Macrotendencias Tecnológicas", "Technology Macro Trends",
     "SixSigma de México", "Líderes ITC",
     "Qué es la innovación y qué viene a continuación; dirigida a líderes ITC.",
     "What is innovation and what comes next; attended by ITC leaders.",
     "Único", "Sí", "Sí"),

    (2011, "Seminario", "Nuevas Mujeres / Innovación", "New Women / Innovation",
     "Grupo Industrial Saltillo (GIS)", "Senior management de GIS",
     "Seminario de innovación para alta gerencia de GIS.",
     "Innovation seminar for GIS senior management.",
     "Único", "Sí", "No"),

    (2011, "Seminario", "Teoría del Conocimiento / Filosofía y Estructura Social",
     "Knowledge Theory / Philosophy and Society Structure",
     "Bachillerato (público especializado)", "Estudiantes y profesores de preparatoria",
     "Seminarios para estudiantes y profesores de preparatoria sobre teoría del conocimiento y estructura social.",
     "Seminars for high school students and teachers on knowledge theory and society structure.",
     "Único", "Sí", "No"),
]


# 11 Reconocimientos y Membresías
RECOG_HEADERS = [
    "Tipo", "Nombre ES", "Name EN", "Organización", "Año",
    "Descripción ES", "Description EN", "Incluir CV ES", "Incluir CV EN"
]
RECOG_ROWS = [
    ("Reconocimiento", "Mención Honorífica · Concurso de Cuento y Poesía",
     "Honorable Mention · Short Story and Poetry Contest",
     "Sociedad General de Escritores de México (SOGEM)", 2004,
     "Mención honorífica en concurso literario; autor de cuento y poesía.",
     "Honorable mention in literary contest; author of short story and poetry.",
     "Sí", "Sí"),

    ("Membresía", "Miembro de MENSA México", "Member of MENSA Mexico",
     "MENSA · High-IQ Society", "Vigente",
     "Miembro de la sociedad internacional de alto coeficiente intelectual, capítulo México.",
     "Member of the international high-IQ society, Mexico chapter.",
     "Sí", "Sí"),
]


# ─── Estilos para nuevas sheets ─────────────────────────────────────────────
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="0053E2")  # walmart blue.100
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2A8703")  # walmart green
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP = Alignment(vertical="top", wrap_text=True)


def make_sheet(wb, name: str, title: str, headers: list[str], rows: list[tuple]):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    # row 1: title
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).fill = TITLE_FILL
    ws.cell(row=1, column=1).alignment = CENTER
    ws.row_dimensions[1].height = 28
    # row 2: headers
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.row_dimensions[2].height = 22
    # data rows
    for r_idx, row in enumerate(rows, start=3):
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.alignment = WRAP
    # column widths
    widths = {
        "Año": 12, "Tipo": 18, "Organización": 38, "Audiencia": 26,
        "Periodo": 14, "Incluir CV ES": 14, "Incluir CV EN": 14,
    }
    for col, h in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=2, column=col).column_letter].width = widths.get(h, 30)
    ws.freeze_panes = "A3"
    return ws


def main():
    print(f"📂 Master: {MASTER}")
    print(f"🛟 Backup ya existe: {BACKUP.name}")
    if not BACKUP.exists():
        shutil.copy2(MASTER, BACKUP)
        print(f"   creado ahora")

    wb = load_workbook(str(MASTER))

    # ── 02 Experiencia: append puestos ────────────────────────────────────
    ws_exp = wb["02 Experiencia"]
    # detectar última fila con ID
    last_id_row = ws_exp.max_row
    existing_ids = {
        ws_exp.cell(row=r, column=1).value
        for r in range(3, last_id_row + 1)
        if ws_exp.cell(row=r, column=1).value is not None
    }
    print(f"\n[02 Experiencia] IDs existentes: {sorted(int(x) for x in existing_ids if str(x).isdigit())}")
    added_pos = 0
    for row in NEW_POSITIONS:
        if row[0] in existing_ids:
            print(f"  ⏭  ID {row[0]} ya existe, skip")
            continue
        ws_exp.append(row)
        print(f"  ➕ ID {row[0]:>2} · {row[1]} · {row[4]}")
        added_pos += 1

    # ── 05 Educación: append fila ─────────────────────────────────────────
    ws_edu = wb["05 Educación"]
    existing_edu = set()
    for r in range(3, ws_edu.max_row + 1):
        title = ws_edu.cell(row=r, column=3).value
        if title:
            existing_edu.add(str(title).lower())
    added_edu = 0
    for row in NEW_EDUCATION:
        if row[2].lower() in existing_edu:
            print(f"  ⏭  Educación '{row[2]}' ya existe, skip")
            continue
        ws_edu.append(row)
        print(f"\n[05 Educación] ➕ {row[2]} ({row[6]})")
        added_edu += 1

    # ── 10 Conferencias y Speaker (nueva) ─────────────────────────────────
    make_sheet(wb, "10 Conferencias y Speaker",
               "CONFERENCIAS Y SPEAKER ENGAGEMENTS — Charlas, seminarios, cátedras y podcasts",
               SPEAKER_HEADERS, SPEAKER_ROWS)
    print(f"\n[10 Conferencias y Speaker] ✨ sheet creada con {len(SPEAKER_ROWS)} entradas")

    # ── 11 Reconocimientos y Membresías (nueva) ───────────────────────────
    make_sheet(wb, "11 Reconocimientos y Membresías",
               "RECONOCIMIENTOS Y MEMBRESÍAS — Premios, distinciones y afiliaciones",
               RECOG_HEADERS, RECOG_ROWS)
    print(f"[11 Reconocimientos y Membresías] ✨ sheet creada con {len(RECOG_ROWS)} entradas")

    wb.save(str(MASTER))
    print(f"\n✅ Master guardado")
    print(f"   • {added_pos} puestos añadidos a 02 Experiencia")
    print(f"   • {added_edu} entradas añadidas a 05 Educación")
    print(f"   • 2 sheets nuevas (10 y 11)")
    print(f"\n📊 Resumen final del master:")
    wb2 = load_workbook(str(MASTER), data_only=True)
    for s in wb2.sheetnames:
        count = sum(1 for r in wb2[s].iter_rows(min_row=3, values_only=True)
                    if any(c is not None for c in r))
        print(f"   {s}: {count} filas con datos")


if __name__ == "__main__":
    main()
