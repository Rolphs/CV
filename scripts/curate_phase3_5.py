"""
Fase 3.5 · Curación de redundancias + traducción de placeholders.

Steps:
  1. Backup master (v3_pre_curation).
  2. Delete redundant rows (sub-set of another logro with better phrasing).
  3. Re-number remaining IDs to keep sequence clean.
  4. Translate [Traducir desde EN] placeholders to Spanish.
  5. Clean PDF artifacts (ligatures, bullet noise) in EN column.
  6. Save master.
"""

from __future__ import annotations
import re
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "data" / "master" / "cv_master_raul_mercado.xlsx"
BACKUP = ROOT / "data" / "master" / "backups" / "v3_pre_curation.xlsx"

# IDs to delete (redundant cluster losers)
DELETE_IDS = {49, 53, 57, 59, 62}

# Translations + EN cleanup for kept rows
# Format: id -> (es_translation, en_clean_or_None)
EDITS = {
    50: (
        "Aumento del 30% en NPS de clientes mediante diseño de experiencia potenciado por IA.",
        None,
    ),
    51: (
        "Liderazgo en la integración de flujos de trabajo con IA y sistemas escalables de insights, alcanzando 90% de automatización en reportería e incrementando velocidad-de-insight en 30%.",
        "Led the integration of AI-driven workflows and scalable insight systems, achieving 90% automation in reporting and increasing speed-to-insight by 30%.",
    ),
    52: (
        "Generación de +$2M en impacto de ingresos mediante estrategias de contenido basadas en insights y campañas de targeting juvenil.",
        "Delivered $2M+ revenue impact through insight-driven content strategies and youth targeting initiatives.",
    ),
    54: (
        "Liderazgo de 150+ proyectos combinando investigación cualitativa, encuestas y analytics conductuales para informar programación de TV, estrategia digital y desarrollo de audiencias.",
        "Led 150+ projects combining qualitative research, surveys, and behavioral analytics to shape TV programming, digital strategy, and audience development.",
    ),
    55: (
        "Gestión de portafolios globales de clientes FMCG, reduciendo churn por debajo del 5% e incrementando satisfacción en 15% YoY.",
        "Managed global FMCG client portfolios, reducing churn below 5% and increasing satisfaction by 15% YoY.",
    ),
    56: (
        "Logro de <5% de churn de clientes y crecimiento de 60% en cuentas estratégicas mediante diseño innovador de estudios y entrega consistente.",
        "Achieved <5% client churn and grew strategic accounts by 60% through innovative study design and consistent delivery.",
    ),
    58: (
        "Generación de +$5M en ventas y duplicación anual de la base de clientes mediante estrategias de adquisición basadas en insights.",
        "Delivered $5M+ in sales and consistently doubled the client base year-over-year through insight-driven acquisition strategies.",
    ),
    60: (
        "Aplicación estratégica de ChatGPT, Claude y GPT-4 para aceleración y síntesis en flujos de investigación.",
        "Strategic use of ChatGPT, Claude, and GPT-4 for research workflow acceleration and synthesis.",
    ),
    61: (
        "Gestión de pipeline robusto de clientes con 100% de entrega a tiempo y alta satisfacción sostenida basada en rigor de diseño, storytelling y profundidad de insight.",
        "Managed a robust client pipeline with 100% on-time delivery and consistently high satisfaction based on design rigor, storytelling, and insight depth.",
    ),
    63: (
        "Incremento del 30% en Net Promoter Score mediante aplicación de insights del cliente para optimizar UX y engagement.",
        "Drove a 30% increase in Net Promoter Score by applying customer insights to optimize UX and engagement.",
    ),
}

# Tag/Categoría enrichment for the 18 new logros (IDs 49-66, kept after dedupe)
ENRICHMENT = {
    50: {"categoria": "Innovación / IA", "tags_ind": "Media,Streaming", "tags_skill": "IA,CX,NPS"},
    51: {"categoria": "Automatización / IA", "tags_ind": "Media,Streaming", "tags_skill": "IA,Automatización,Insights"},
    52: {"categoria": "Crecimiento / Revenue", "tags_ind": "Media,Broadcasting", "tags_skill": "Insights,Estrategia de Contenido"},
    54: {"categoria": "Liderazgo de Investigación", "tags_ind": "Media,Broadcasting", "tags_skill": "Mixed-methods,Programación TV"},
    55: {"categoria": "Gestión de Clientes", "tags_ind": "FMCG,Consultoría", "tags_skill": "Account Management,Retención"},
    56: {"categoria": "Crecimiento / Retención", "tags_ind": "FMCG,Consultoría", "tags_skill": "Diseño de Estudios,Account Growth"},
    58: {"categoria": "Ventas / Growth", "tags_ind": "Tecnología,B2B", "tags_skill": "Ventas Consultivas,Insights"},
    60: {"categoria": "Innovación / IA", "tags_ind": "Tecnología,B2B", "tags_skill": "IA Generativa,Productividad"},
    61: {"categoria": "Excelencia Operativa", "tags_ind": "Consultoría", "tags_skill": "Project Management,Storytelling"},
    63: {"categoria": "UX / CX", "tags_ind": "Startup,Tech", "tags_skill": "UX,NPS,Customer Insights"},
    64: {"categoria": "Innovación / IA", "tags_ind": "Consultoría,Research", "tags_skill": "IA,Automatización,Research"},
    65: {"categoria": "Estrategia / Arquitectura", "tags_ind": "Consultoría,Research,Media", "tags_skill": "KPI,Análisis,Estrategia"},
    66: {"categoria": "Liderazgo", "tags_ind": "Consultoría,Research", "tags_skill": "Liderazgo,Multidisciplinario"},
}

LIGATURES = {"ﬂ": "fl", "ﬃ": "ffi", "ﬁ": "fi", "ﬀ": "ff", "ﬄ": "ffl"}


def clean_pdf_text(t: str) -> str:
    if not t:
        return t
    for lig, rep in LIGATURES.items():
        t = t.replace(lig, rep)
    t = re.sub(r"^\s*[•●▪■◾▶►‣]+\s*", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def main():
    import shutil
    print(f"📦 Backup → {BACKUP.name}")
    shutil.copy2(MASTER, BACKUP)

    wb = load_workbook(str(MASTER))
    ws = wb["03 Logros"]

    # Collect all data rows
    all_rows = []  # list of (row_idx, [values])
    for ridx, row in enumerate(ws.iter_rows(min_row=3, max_col=14, values_only=False), start=3):
        if row[0].value is None:
            continue
        values = [cell.value for cell in row]
        all_rows.append((ridx, values))

    # Filter out deleted IDs, keep ID-to-newID mapping (sequential renumber)
    kept = []
    deleted = []
    for ridx, values in all_rows:
        rid = int(values[0])
        if rid in DELETE_IDS:
            deleted.append(rid)
            continue
        kept.append((rid, values))
    print(f"🗑️  Deleted IDs: {sorted(deleted)}")
    print(f"✅ Kept: {len(kept)} rows")

    # Apply translations + cleanup
    for rid, values in kept:
        if rid in EDITS:
            es_new, en_new = EDITS[rid]
            values[4] = es_new
            if en_new is not None:
                values[5] = en_new
        # always run cleanup on EN column
        values[5] = clean_pdf_text(values[5] or "")
        # apply enrichment
        if rid in ENRICHMENT:
            e = ENRICHMENT[rid]
            values[9] = e["categoria"]   # Categoría (col 10)
            values[10] = e["tags_ind"]    # Tags Industria (col 11)
            values[11] = e["tags_skill"]  # Tags Habilidad (col 12)

    # Wipe sheet data area (rows 3+) and rewrite from kept
    last_row = ws.max_row
    # delete from bottom up so indices don't shift
    for ridx in range(last_row, 2, -1):
        ws.delete_rows(ridx, 1)

    # Write back (preserving original IDs to keep traceability)
    target_row = 3
    for rid, values in kept:
        for cidx, val in enumerate(values, start=1):
            ws.cell(row=target_row, column=cidx, value=val)
        target_row += 1

    # Save
    wb.save(str(MASTER))
    print(f"💾 Saved master")
    print(f"\n📊 Final logros count: {len(kept)} (was 66 before, -{len(deleted)} dedup)")


if __name__ == "__main__":
    main()
