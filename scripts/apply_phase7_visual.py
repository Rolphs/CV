"""
Fase 7 · Polish VISUAL del master para que se vea IMPRESIONANTE.

Cambios:
  1. Actualiza 00 Instrucciones para reflejar las 12 sheets (agregar 10 y 11)
  2. Aplica diseño consistente a todas las sheets de datos (02-11):
       - Row 1: Título con bg blue.100 (#0053E2), texto blanco bold 14pt
       - Row 2: Headers con bg green (#2A8703), texto blanco bold 11pt
       - Rows 3+: alternancia sutil (gray.5 / blanco), wrap text
       - Bordes sutiles
       - Freeze panes en A3 (encabezados pinned)
       - Column widths optimizados por tipo de columna
       - Tab colors por categoría:
           Meta (00, 01): gris
           Core (02 Experiencia, 03 Logros): walmart blue
           Inventories (04-07, 10, 11): walmart green
           Strategy (08, 09): walmart spark (amarillo)
"""
from __future__ import annotations
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "data" / "master" / "cv_master_raul_mercado.xlsx"
BACKUP = ROOT / "data" / "master" / "backups" / "v7_pre_visual.xlsx"


# ─── Walmart palette ───────────────────────────────────────────────────────
BLUE_100 = "0053E2"      # primary
SPARK_100 = "FFC220"     # accent
GREEN_100 = "2A8703"
GREEN_10 = "D9EBC9"
GRAY_5 = "F5F5F5"
GRAY_10 = "E8E8E8"
GRAY_160 = "1A1A1A"


# ─── Styles ────────────────────────────────────────────────────────────────
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor=BLUE_100)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=GREEN_100)
BAND_FILL = PatternFill("solid", fgColor=GRAY_5)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
SUBTLE_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)


# ─── Column width hints by header name ─────────────────────────────────────
COL_WIDTH_HINTS = {
    # ids and short codes
    "ID": 6, "ID Logro": 9, "ID Empresa": 11,
    "Año": 10, "Tipo": 16, "País": 12, "Nivel": 14, "Estado": 14,
    "Fecha Inicio": 13, "Fecha Fin": 13, "¿Actual?": 10,
    "Vigencia": 13, "Año inicio": 12, "Año fin": 11,
    "Año inicio.1": 12, "Personas a cargo": 18,
    "Incluir CV ES": 14, "Incluir CV EN": 14, "Periodo": 14,
    "Tiene Métrica": 14, "Impacto": 11,
    # medium-length
    "Empresa": 28, "Empresa / Company": 28, "Industria": 22, "Industria / Industry": 22,
    "Dept / Área": 18, "Reporta a": 22, "Status": 14, "Año aplicación": 14,
    "Categoría": 22, "Institución": 32, "Organización": 32,
    "Audiencia": 26, "Skills a enfatizar": 32,
    "Habilidad ES": 26, "Skill EN": 26, "Años exp.": 11,
    "Área": 20,
    # bilingual title pairs
    "Puesto ES": 32, "Puesto EN": 32,
    "Título / Grado ES": 32, "Degree EN": 32,
    "Certificado ES": 28, "Certificate EN": 28,
    "Actividad ES": 32, "Activity EN": 32,
    "Título ES": 32, "Title EN": 32,
    "Nombre ES": 32, "Name EN": 32,
    # long-text columns
    "Descripción Empresa ES": 42, "Descripción Empresa EN": 42,
    "Logro / Resp. ES": 50, "Achievement EN": 50,
    "Descripción ES": 42, "Description EN": 42,
    "Notas / Customización": 38,
    "Tags Industria": 22, "Tags Habilidad": 24, "Métrica / KPI": 22,
    "Relevancia para CV": 38, "Relevancia": 38,
    "URL / Credencial": 30, "Público beneficiado": 26,
    "Puesto al que aplico": 26, "Logros a destacar (IDs)": 22,
    "Fecha aplicación": 14, "Resultado": 18,
    "Tipo de empresa": 22, "Perfil ES (≈120 palabras)": 60,
    "Profile EN (≈120 words)": 60,
}


# ─── Tab colors by sheet ───────────────────────────────────────────────────
TAB_COLORS = {
    "00 Instrucciones": "808080",   # gris
    "01 Perfil": "A0A0A0",          # gris claro
    "02 Experiencia": BLUE_100,     # azul ⭐ core
    "03 Logros": BLUE_100,          # azul ⭐ core
    "04 Skills": GREEN_100,
    "05 Educación": GREEN_100,
    "06 Certificaciones": GREEN_100,
    "07 Voluntariado": GREEN_100,
    "08 Empresas Target": SPARK_100,
    "09 Perfiles Narrativos": SPARK_100,
    "10 Conferencias y Speaker": GREEN_100,
    "11 Reconocimientos y Membresías": GREEN_100,
}


def polish_sheet(ws, has_title_row=True):
    """Apply visual polish to a data sheet (rows 1-2 = title/headers, 3+ = data)."""
    max_col = ws.max_column
    max_row = ws.max_row

    # Title row (row 1)
    if has_title_row and max_col > 0:
        ws.cell(row=1, column=1).font = TITLE_FONT
        ws.cell(row=1, column=1).fill = TITLE_FILL
        ws.cell(row=1, column=1).alignment = CENTER
        ws.row_dimensions[1].height = 30
        # merge across all columns (if not already merged)
        try:
            existing_merges = [str(r) for r in ws.merged_cells.ranges]
            target_range = f"A1:{get_column_letter(max_col)}1"
            if target_range not in existing_merges:
                ws.merge_cells(target_range)
        except Exception:
            pass

    # Header row (row 2)
    for c in range(1, max_col + 1):
        cell = ws.cell(row=2, column=c)
        if cell.value is not None:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = SUBTLE_BORDER
    ws.row_dimensions[2].height = 30

    # Banded rows + wrap on data
    for r in range(3, max_row + 1):
        if r % 2 == 0:
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = BAND_FILL
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = LEFT_WRAP
            cell.border = SUBTLE_BORDER

    # Column widths from hints (header name → width)
    for c in range(1, max_col + 1):
        header_val = ws.cell(row=2, column=c).value
        header_name = str(header_val).strip() if header_val is not None else ""
        width = COL_WIDTH_HINTS.get(header_name, 22)
        ws.column_dimensions[get_column_letter(c)].width = width

    # Freeze panes
    ws.freeze_panes = "A3"


def polish_meta_sheet(ws):
    """Light polish for 00/01 sheets that have different layout (key→value rows)."""
    max_col = ws.max_column
    max_row = ws.max_row
    # Title row
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).fill = TITLE_FILL
    ws.cell(row=1, column=1).alignment = CENTER
    ws.row_dimensions[1].height = 30
    try:
        ws.merge_cells(f"A1:{get_column_letter(max_col)}1")
    except Exception:
        pass

    # Section header rows (col A has bold all-caps text and col B is empty)
    # We detect: row where col A is uppercase string and col B is None
    section_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor=GREEN_100)
    for r in range(2, max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if a and isinstance(a, str) and a.strip() and a == a.upper() and not b:
            ws.cell(row=r, column=1).font = section_font
            ws.cell(row=r, column=1).fill = section_fill
            ws.cell(row=r, column=1).alignment = CENTER
            try:
                ws.merge_cells(f"A{r}:{get_column_letter(max_col)}{r}")
            except Exception:
                pass

    # Make col A bold-ish for keys (when col B has value)
    key_font = Font(name="Calibri", size=10, bold=True, color=GRAY_160)
    for r in range(2, max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if b is not None and a and isinstance(a, str) and a != a.upper():
            ws.cell(row=r, column=1).font = key_font

    # Column widths
    widths = [4, 36, 80, 30, 30, 30, 30]
    for i, w in enumerate(widths[:max_col], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Wrap on long cells
    wrap = Alignment(vertical="top", wrap_text=True)
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell.alignment = wrap

    ws.freeze_panes = "A2"


# ─── 00 Instrucciones: actualizar para incluir sheets 10 y 11 ──────────────
def update_instrucciones(ws):
    """Add rows for sheets 10 and 11 + refresh counts."""
    # Find row that mentions "09 Perfiles Narrativos"
    insert_after = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=2).value
        if a and "09 Perfiles Narrativos" in str(a):
            insert_after = r
            break
    if insert_after is None:
        return False
    # Insert 2 rows after
    ws.insert_rows(insert_after + 1, amount=2)
    ws.cell(row=insert_after + 1, column=2, value="10 Conferencias y Speaker")
    ws.cell(row=insert_after + 1, column=3,
            value="Charlas, seminarios, cátedras y podcasts. Útil para sección 'Speaker / Thought Leadership'.")
    ws.cell(row=insert_after + 2, column=2, value="11 Reconocimientos y Membresías")
    ws.cell(row=insert_after + 2, column=3,
            value="Premios, distinciones y afiliaciones profesionales (SOGEM, MENSA, etc.).")
    return True


def main():
    print(f"📂 Master: {MASTER}")
    if not BACKUP.exists():
        shutil.copy2(MASTER, BACKUP)
        print(f"🛟 Backup creado: {BACKUP.name}")

    wb = load_workbook(str(MASTER))
    print(f"\nPolishing {len(wb.sheetnames)} sheets...\n")

    # Update 00 Instrucciones content
    ws_00 = wb["00 Instrucciones"]
    if update_instrucciones(ws_00):
        print("[00] ✏️  Agregadas referencias a sheets 10 y 11")
    polish_meta_sheet(ws_00)
    print("[00] 🎨 Polish aplicado")

    # 01 Perfil: meta layout
    polish_meta_sheet(wb["01 Perfil"])
    print("[01] 🎨 Polish aplicado")

    # 02-11: data sheets
    for sname in wb.sheetnames:
        if sname in ("00 Instrucciones", "01 Perfil"):
            continue
        polish_sheet(wb[sname])
        print(f"[{sname[:2]}] 🎨 Polish aplicado")

    # Tab colors
    for sname, color in TAB_COLORS.items():
        if sname in wb.sheetnames:
            wb[sname].sheet_properties.tabColor = color
    print("\n🌈 Tab colors aplicados")

    # Set active sheet to 01 Perfil (opens there)
    wb.active = wb.sheetnames.index("01 Perfil")

    wb.save(str(MASTER))
    print(f"\n✅ Master pulido y guardado")
    print(f"\n📊 Sheets finales:")
    for s in wb.sheetnames:
        print(f"   • {s}")


if __name__ == "__main__":
    main()
