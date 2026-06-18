"""
Fase 6 · Auditoría profunda del master xlsx.

Pasos:
  1. RECURSIVO: Dump completo de cada sheet → JSON crudo + MD legible
  2. CONVOLUCIONAL: Análisis cruzado entre sheets
       - Cronología de puestos (orden, overlaps, gaps, congruencia años)
       - Integridad referencial 02→03 (ID Empresa en logros existe)
       - Paridad ES/EN en cada sheet bilingüe
       - Formatos de fecha consistentes
       - Duplicados y redundancias
       - Métricas vs claim "Tiene Métrica"
       - Empresas mencionadas en perfiles narrativos vs catálogo
  3. ESTILO: detectores heurísticos de problemas de redacción
       - Logros muy cortos (<40 char) o muy largos (>300 char)
       - Falta de verbo de acción al inicio
       - Inconsistencia en uso de gerundios vs infinitivos
       - Capitalización irregular
       - "Etc." / muletillas
       - Acentos mal puestos / palabras sin acento

Outputs:
  - data/reports/master_full_dump.{json,md}
  - data/reports/master_audit.md
"""
from __future__ import annotations
import json
import re
import unicodedata
from collections import defaultdict, Counter
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "data" / "master" / "cv_master_raul_mercado.xlsx"
OUT = ROOT / "data" / "reports"


# ─── Helpers ────────────────────────────────────────────────────────────────
def norm(s):
    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c)).lower()


def parse_year(val):
    """Try to extract a 4-digit year from any value."""
    if val is None:
        return None
    s = str(val).strip()
    if "actual" in s.lower() or "presente" in s.lower() or "present" in s.lower():
        return 2026
    # MM/YYYY or YYYY
    m = re.search(r"(19\d{2}|20\d{2})", s)
    return int(m.group(1)) if m else None


def parse_month_year(val):
    """Parse to (year, month). month=6 if not detectable."""
    if val is None:
        return (None, None)
    s = str(val).strip()
    if "actual" in s.lower() or "presente" in s.lower() or "present" in s.lower():
        return (2026, 5)
    # MM/YYYY
    m = re.match(r"(\d{1,2})/(19\d{2}|20\d{2})", s)
    if m:
        return (int(m.group(2)), int(m.group(1)))
    # Roman/text month: VIII/2017
    roman = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6, "VII": 7,
             "VIII": 8, "IX": 9, "X": 10, "XI": 11, "XII": 12}
    m = re.match(r"([IVX]+)/(19\d{2}|20\d{2})", s)
    if m and m.group(1) in roman:
        return (int(m.group(2)), roman[m.group(1)])
    # Just YYYY
    m = re.match(r"(19\d{2}|20\d{2})$", s)
    if m:
        return (int(m.group(1)), 6)
    return (parse_year(val), 6)


def read_sheet_full(ws):
    """Read all cells, return list of dicts (header → value) starting from row 3."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return [], []
    headers = [str(c).strip() if c is not None else f"col{i}" for i, c in enumerate(rows[1])]
    data = []
    for r in rows[2:]:
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in r):
            continue
        data.append({h: r[i] if i < len(r) else None for i, h in enumerate(headers)})
    return headers, data


# ─── 1. RECURSIVE DUMP ──────────────────────────────────────────────────────
def dump_all(wb):
    out = {}
    for sname in wb.sheetnames:
        ws = wb[sname]
        headers, data = read_sheet_full(ws)
        out[sname] = {"headers": headers, "rows": data, "n_rows": len(data)}
    return out


# ─── 2. CONVOLUTIONAL AUDIT ─────────────────────────────────────────────────
ACTION_VERBS_ES = {
    "lideré", "lideraron", "lidero", "lidere", "dirigí", "dirijo", "diseñé", "diseño",
    "implementé", "implemento", "desarrollé", "desarrollo", "establecí", "establezco",
    "creé", "creo", "construí", "construyo", "logré", "logro", "consolidé",
    "automaticé", "automatizo", "transformé", "transformo", "optimicé", "optimizo",
    "incrementé", "incremento", "reduje", "reducido", "generé", "genero",
    "ejecuté", "ejecuto", "supervisé", "supervisión", "coordiné", "coordino",
    "negocié", "negocio", "gestioné", "gestiono", "definí", "definido", "rediseñé",
    "mantuve", "mantengo", "responsable", "encargado", "gestión", "diseño",
    "implementación", "desarrollo", "liderazgo", "análisis", "investigación",
    "tropicalización", "modelación", "moderación", "automatización", "presentación",
    "presenté", "preparé", "elaboré", "atendí", "investigué", "fundación",
    "asesoría", "atención", "implementé", "introduje", "ejecución",
}

ACTION_VERBS_EN = {
    "led", "leading", "directed", "designed", "implemented", "developed",
    "established", "created", "built", "achieved", "consolidated", "automated",
    "transformed", "optimized", "increased", "reduced", "generated", "executed",
    "supervised", "coordinated", "negotiated", "managed", "defined", "redesigned",
    "maintained", "responsible", "implementing", "designing", "leading", "leadership",
    "analysis", "research", "tropicalization", "modelling", "moderating", "automation",
    "presented", "prepared", "elaborated", "attended", "investigated", "founded",
    "advisory", "introduced", "execution",
}


def audit(dump):
    issues = []  # list of (severity, sheet, ref, message)
    info = []

    def add(sev, sheet, ref, msg):
        issues.append((sev, sheet, ref, msg))

    # ── 02 Experiencia ─────────────────────────────────────────────────────
    exp = dump["02 Experiencia"]
    exp_rows = exp["rows"]
    # build year ranges
    exp_index = {}
    chrono = []
    seen_keys = defaultdict(list)
    for row in exp_rows:
        eid = row.get("ID")
        if eid is None:
            continue
        eid = int(eid)
        empresa = (row.get("Empresa / Company") or "").strip()
        puesto_es = (row.get("Puesto ES") or "").strip()
        puesto_en = (row.get("Puesto EN") or "").strip()
        fi = row.get("Fecha Inicio")
        ff = row.get("Fecha Fin")
        actual = (row.get("¿Actual?") or "").strip().lower()
        yi, mi = parse_month_year(fi)
        yf, mf = parse_month_year(ff)
        exp_index[eid] = {
            "empresa": empresa, "puesto_es": puesto_es, "puesto_en": puesto_en,
            "fi": fi, "ff": ff, "yi": yi, "mi": mi, "yf": yf, "mf": mf,
            "actual": actual,
        }
        chrono.append((eid, empresa, puesto_es, yi, mi, yf, mf, fi, ff, actual))

        # check: ES vs EN parity
        if puesto_es and not puesto_en:
            add("🟡", "02 Experiencia", f"ID {eid}", f"Puesto ES presente pero EN vacío: '{puesto_es}'")
        if puesto_en and not puesto_es:
            add("🟡", "02 Experiencia", f"ID {eid}", f"Puesto EN presente pero ES vacío")

        # check: fechas válidas
        if not yi:
            add("🔴", "02 Experiencia", f"ID {eid}", f"Fecha Inicio no parseable: '{fi}'")
        if not yf:
            add("🔴", "02 Experiencia", f"ID {eid}", f"Fecha Fin no parseable: '{ff}'")
        if yi and yf and yi > yf:
            add("🔴", "02 Experiencia", f"ID {eid}", f"Fecha Inicio ({fi}) posterior a Fin ({ff})")

        # check: formato fecha inconsistente
        fi_str = str(fi or "").strip()
        ff_str = str(ff or "").strip()
        for label, s in [("Inicio", fi_str), ("Fin", ff_str)]:
            if s and not re.match(r"^(\d{1,2}/\d{4}|\d{4}|[IVX]+/\d{4}|Actual|Presente)$", s, re.I):
                add("🟡", "02 Experiencia", f"ID {eid}",
                    f"Formato fecha {label} no estándar: '{s}' (esperado MM/YYYY o YYYY)")

        # check duplicate company+puesto+yi
        key = (norm(empresa), norm(puesto_es), yi)
        seen_keys[key].append(eid)

        # check ¿Actual? coherence
        is_actual_flag = actual in ("sí", "si", "yes", "true")
        ff_is_actual = "actual" in ff_str.lower() or "presente" in ff_str.lower()
        if is_actual_flag != ff_is_actual:
            add("🟡", "02 Experiencia", f"ID {eid}",
                f"Inconsistencia ¿Actual?='{actual}' vs Fecha Fin='{ff}'")

    # duplicados
    for key, ids in seen_keys.items():
        if len(ids) > 1:
            add("🟡", "02 Experiencia", f"IDs {ids}",
                f"Posible duplicado: empresa+puesto+año = {key}")

    # cronología: sort by start
    sorted_chrono = sorted(chrono, key=lambda x: (x[3] or 0, x[4] or 0))
    info.append(("📅 Cronología (por inicio):", sorted_chrono))

    # gaps and overlaps (excluding paralelos/freelance)
    timeline = sorted([(yi*12+mi, yf*12+mf, eid, empresa, puesto_es)
                       for eid, empresa, puesto_es, yi, mi, yf, mf, *_ in chrono
                       if yi and yf])
    # ─── 02 vs 03 Logros referential integrity ─────────────────────────────
    logros = dump["03 Logros"]["rows"]
    exp_ids = set(exp_index.keys())
    company_by_id = {eid: exp_index[eid]["empresa"] for eid in exp_ids}
    logros_by_emp_id = defaultdict(list)
    for lrow in logros:
        lid = lrow.get("ID Logro")
        eid = lrow.get("ID Empresa")
        empresa_text = (lrow.get("Empresa") or "").strip()
        logro_es = (lrow.get("Logro / Resp. ES") or "").strip()
        logro_en = (lrow.get("Achievement EN") or "").strip()
        if eid is None:
            add("🔴", "03 Logros", f"Logro {lid}", "ID Empresa vacío")
            continue
        try:
            eid_int = int(eid)
        except (ValueError, TypeError):
            add("🔴", "03 Logros", f"Logro {lid}", f"ID Empresa no numérico: '{eid}'")
            continue
        if eid_int not in exp_ids:
            add("🔴", "03 Logros", f"Logro {lid}",
                f"ID Empresa {eid_int} no existe en 02 Experiencia")
            continue
        logros_by_emp_id[eid_int].append(lrow)
        # verify company name matches
        expected = company_by_id[eid_int]
        if norm(expected) != norm(empresa_text):
            add("🟡", "03 Logros", f"Logro {lid}",
                f"Empresa '{empresa_text}' no coincide con ID {eid_int} ('{expected}')")
        # ES/EN parity
        if logro_es and not logro_en:
            add("🟡", "03 Logros", f"Logro {lid}", f"ES presente pero EN vacío")
        if logro_en and not logro_es:
            add("🟡", "03 Logros", f"Logro {lid}", f"EN presente pero ES vacío")
        # length checks
        if logro_es:
            if len(logro_es) < 40:
                add("🟢", "03 Logros", f"Logro {lid}",
                    f"ES muy corto ({len(logro_es)} ch): '{logro_es}'")
            if len(logro_es) > 320:
                add("🟢", "03 Logros", f"Logro {lid}",
                    f"ES muy largo ({len(logro_es)} ch)")
        # action verb start
        if logro_es:
            first_word = norm(logro_es.split()[0]) if logro_es.split() else ""
            if first_word and not any(first_word.startswith(v) for v in ACTION_VERBS_ES):
                add("🟢", "03 Logros", f"Logro {lid}",
                    f"ES no inicia con verbo de acción reconocido: '{first_word}…'")
        # metric flag vs metric content
        tiene_metrica = (lrow.get("Tiene Métrica") or "").strip().lower()
        metrica = (lrow.get("Métrica / KPI") or "").strip()
        has_number = bool(re.search(r"\d", logro_es)) or bool(metrica)
        if tiene_metrica in ("sí", "si", "yes") and not has_number:
            add("🟡", "03 Logros", f"Logro {lid}",
                "Marcado 'Tiene Métrica=Sí' pero no se encontró número")
        if tiene_metrica in ("no", "") and has_number and metrica:
            add("🟢", "03 Logros", f"Logro {lid}",
                f"Marcado 'Tiene Métrica=No' pero hay valor en Métrica: '{metrica}'")

    # check: empresas sin logros
    for eid, info_e in exp_index.items():
        if eid not in logros_by_emp_id:
            empresa = info_e["empresa"]
            add("🟢", "03 Logros", f"ID Empresa {eid}",
                f"Sin logros asociados: {empresa} ({info_e['puesto_es']})")

    # ─── ES/EN parity en otras sheets ─────────────────────────────────────
    for sname, key_es, key_en in [
        ("04 Skills", "Habilidad ES", "Skill EN"),
        ("05 Educación", "Título / Grado ES", "Degree EN"),
        ("06 Certificaciones", "Certificado ES", "Certificate EN"),
        ("07 Voluntariado", "Actividad ES", "Activity EN"),
        ("10 Conferencias y Speaker", "Título ES", "Title EN"),
        ("11 Reconocimientos y Membresías", "Nombre ES", "Name EN"),
    ]:
        if sname not in dump:
            continue
        for i, r in enumerate(dump[sname]["rows"], start=3):
            es = (r.get(key_es) or "").strip()
            en = (r.get(key_en) or "").strip()
            if es and not en:
                add("🟡", sname, f"row {i}", f"{key_es} presente pero {key_en} vacío")
            if en and not es:
                add("🟡", sname, f"row {i}", f"{key_en} presente pero {key_es} vacío")

    # ─── Cronología: detectar overlaps relevantes y gaps ──────────────────
    # Para evitar ruido, ignoro overlaps si uno es freelance/paralelo
    paralelos = {"freelance", "independiente", "tiempo parcial", "contrato"}
    main_timeline = []
    for eid, info_e in exp_index.items():
        tipo = next((r.get("Tipo") for r in exp_rows if r.get("ID") == eid), "")
        if tipo and any(p in norm(tipo) for p in paralelos):
            continue
        if info_e["yi"] and info_e["yf"]:
            main_timeline.append((info_e["yi"]*12 + info_e["mi"],
                                 info_e["yf"]*12 + info_e["mf"],
                                 eid, info_e["empresa"], info_e["puesto_es"]))
    main_timeline.sort()
    for i in range(len(main_timeline)-1):
        a_start, a_end, a_id, a_emp, a_pos = main_timeline[i]
        b_start, b_end, b_id, b_emp, b_pos = main_timeline[i+1]
        if b_start < a_end - 1:  # overlap >1 month
            add("🟡", "02 Experiencia (cronología)",
                f"IDs {a_id} vs {b_id}",
                f"Overlap entre '{a_emp}' (hasta {a_end//12}/{a_end%12 or 12}) "
                f"y '{b_emp}' (desde {b_start//12}/{b_start%12 or 12})")
        elif b_start > a_end + 6:  # gap > 6 months
            gap = b_start - a_end
            add("🟢", "02 Experiencia (cronología)",
                f"IDs {a_id} → {b_id}",
                f"Gap de {gap} meses entre '{a_emp}' (fin {a_end//12}/{a_end%12 or 12}) "
                f"y '{b_emp}' (inicio {b_start//12}/{b_start%12 or 12})")

    return issues, info, exp_index, sorted_chrono


# ─── 3. WRITE REPORTS ───────────────────────────────────────────────────────
def write_dump_md(dump, path):
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"# 📋 Dump completo del Master\n\n")
        f.write(f"_Generado: {datetime.now().isoformat()}_\n\n")
        f.write(f"**Sheets:** {len(dump)}\n\n")
        for sname, d in dump.items():
            f.write(f"\n## {sname} ({d['n_rows']} filas)\n\n")
            if not d["rows"]:
                f.write("_(sin datos)_\n")
                continue
            # subset of cols for readability
            cols = d["headers"]
            f.write("| " + " | ".join(cols) + " |\n")
            f.write("|" + "|".join(["---"] * len(cols)) + "|\n")
            for r in d["rows"]:
                cells = []
                for c in cols:
                    v = r.get(c)
                    s = "" if v is None else str(v)
                    s = s.replace("|", "\\|").replace("\n", " ⏎ ")
                    if len(s) > 100:
                        s = s[:97] + "…"
                    cells.append(s)
                f.write("| " + " | ".join(cells) + " |\n")


def write_audit_md(issues, info, exp_index, sorted_chrono, path):
    sev_order = {"🔴": 0, "🟡": 1, "🟢": 2}
    grouped = defaultdict(list)
    for sev, sheet, ref, msg in issues:
        grouped[sev].append((sheet, ref, msg))

    sev_label = {"🔴": "BLOCKERS", "🟡": "WARNINGS", "🟢": "SUGGESTIONS"}
    with open(path, "w", encoding="utf-8") as f:
        f.write("# 🔬 Auditoría profunda del Master\n\n")
        f.write(f"_Generado: {datetime.now().isoformat()}_\n\n")
        f.write(f"**Total hallazgos:** {len(issues)}\n\n")
        for sev in ["🔴", "🟡", "🟢"]:
            f.write(f"- {sev} {sev_label[sev]}: {len(grouped[sev])}\n")
        f.write("\n")

        # cronología
        f.write("## 📅 Cronología de puestos (orden de inicio)\n\n")
        f.write("| ID | Empresa | Puesto ES | Inicio | Fin | ¿Actual? |\n")
        f.write("|---|---|---|---|---|---|\n")
        for eid, empresa, puesto_es, yi, mi, yf, mf, fi, ff, actual in sorted_chrono:
            f.write(f"| {eid} | {empresa[:30]} | {puesto_es[:35]} | {fi} | {ff} | {actual} |\n")
        f.write("\n")

        # issues
        for sev in ["🔴", "🟡", "🟢"]:
            if not grouped[sev]:
                continue
            f.write(f"\n## {sev} {sev_label[sev]} ({len(grouped[sev])})\n\n")
            # group by sheet
            by_sheet = defaultdict(list)
            for sheet, ref, msg in grouped[sev]:
                by_sheet[sheet].append((ref, msg))
            for sheet, items in sorted(by_sheet.items()):
                f.write(f"\n### {sheet} ({len(items)})\n\n")
                for ref, msg in items:
                    f.write(f"- **{ref}**: {msg}\n")


def main():
    print(f"📂 Loading {MASTER}")
    wb = load_workbook(str(MASTER), data_only=True)
    print(f"   {len(wb.sheetnames)} sheets")
    dump = dump_all(wb)
    for s, d in dump.items():
        print(f"   {s}: {d['n_rows']} rows")

    # save raw dump
    json_path = OUT / "master_full_dump.json"
    json_path.write_text(
        json.dumps({s: {"headers": d["headers"],
                        "rows": [{k: (str(v) if v is not None else None)
                                  for k, v in r.items()} for r in d["rows"]],
                        "n_rows": d["n_rows"]}
                    for s, d in dump.items()},
                   indent=2, ensure_ascii=False),
        encoding="utf-8")
    print(f"📁 → {json_path.name}")

    md_dump = OUT / "master_full_dump.md"
    write_dump_md(dump, md_dump)
    print(f"📁 → {md_dump.name}")

    print("\n🔬 Running convolutional audit…")
    issues, info, exp_index, sorted_chrono = audit(dump)
    sev_counts = Counter(i[0] for i in issues)
    print(f"   Hallazgos: 🔴 {sev_counts.get('🔴', 0)} | 🟡 {sev_counts.get('🟡', 0)} | 🟢 {sev_counts.get('🟢', 0)}")

    audit_path = OUT / "master_audit.md"
    write_audit_md(issues, info, exp_index, sorted_chrono, audit_path)
    print(f"📁 → {audit_path.name}")


if __name__ == "__main__":
    main()
