"""Dump full content of key XLSX files so we can design the consolidation schema."""
from pathlib import Path
from openpyxl import load_workbook
import json

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "data" / "raw"
MASTER_DIR = ROOT / "data" / "master"
OUT = ROOT / "data" / "reports"

TARGETS = [
    MASTER_DIR / "cv_master_raul_mercado.xlsx",
    RAW / "04_bases_datos_excel" / "Formato Datos CV - Raul Mercado.xlsx",
]

def dump_sheet(ws, max_rows=200):
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # stop if 10 consecutive completely-empty rows
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            rows.append(None)  # marker, we'll trim later
        else:
            rows.append([("" if c is None else str(c)) for c in row])
        if i >= max_rows:
            break
    # trim trailing None markers
    while rows and rows[-1] is None:
        rows.pop()
    return rows

def main():
    dumps = {}
    for path in TARGETS:
        if not path.exists():
            print(f"⚠️  skip (no existe): {path}")
            continue
        wb = load_workbook(str(path), data_only=True, read_only=True)
        sheets = {}
        for name in wb.sheetnames:
            ws = wb[name]
            sheets[name] = dump_sheet(ws)
        wb.close()
        dumps[path.name] = sheets

    out_json = OUT / "xlsx_full_dump.json"
    out_json.write_text(json.dumps(dumps, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"wrote {out_json}")

    # markdown view, only non-empty rows
    md = OUT / "xlsx_full_dump.md"
    with open(md, "w", encoding="utf-8") as f:
        for fname, sheets in dumps.items():
            f.write(f"\n# {fname}\n")
            for sname, rows in sheets.items():
                f.write(f"\n## {sname}\n\n")
                non_empty = [r for r in rows if r is not None]
                if not non_empty:
                    f.write("_(vacía)_\n")
                    continue
                widths = max(len(r) for r in non_empty)
                f.write("| " + " | ".join(f"c{i}" for i in range(widths)) + " |\n")
                f.write("|" + "|".join(["---"] * widths) + "|\n")
                for r in non_empty:
                    padded = r + [""] * (widths - len(r))
                    cells = [(c[:60] + "…") if len(c) > 60 else c for c in padded]
                    cells = [c.replace("|", "\\|").replace("\n", " ⏎ ") for c in cells]
                    f.write("| " + " | ".join(cells) + " |\n")
    print(f"wrote {md}")

if __name__ == "__main__":
    main()
