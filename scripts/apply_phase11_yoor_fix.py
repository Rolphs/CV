"""Fix Yoor Company end date in master xlsx.

Apr 2022 - Oct 2022 (incorrect) → Apr 2022 - Feb 2023 (correct per Rolph 2026-05-17).
Also rename empresa to 'Yoorco' (commercial name) per user preference.
"""
import shutil
import sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")
ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "data" / "master" / "cv_master_raul_mercado.xlsx"
BACKUP = ROOT / "data" / "master" / "backups" / "v11_pre_yoor_fix.xlsx"

BACKUP.parent.mkdir(parents=True, exist_ok=True)
shutil.copy2(MASTER, BACKUP)
print(f"💾 Backup: {BACKUP.relative_to(ROOT)}")

wb = load_workbook(MASTER)
ws = wb["02 Experiencia"]

# Find header row (with column names) and Yoor row
header_row = None
for r in range(1, 6):
    vals = [ws.cell(row=r, column=c).value for c in range(1, 20)]
    if any(v and "empresa" in str(v).lower() for v in vals):
        header_row = r
        break

if not header_row:
    raise SystemExit("❌ Could not find header row in '02 Experiencia'")

headers = {ws.cell(row=header_row, column=c).value: c
           for c in range(1, 25)
           if ws.cell(row=header_row, column=c).value}
print(f"📋 Header row: {header_row} · columns: {list(headers.keys())[:8]}...")

# Find target columns (be defensive about naming)
def find_col(*needles):
    for h, c in headers.items():
        hl = str(h).lower()
        if all(n in hl for n in needles):
            return c, h
    return None, None

empresa_col, _ = find_col("empresa")
fin_col, fin_name = find_col("fecha", "fin")
print(f"📍 empresa col={empresa_col} · fecha_fin col={fin_col} ('{fin_name}')")

# Walk rows looking for Yoor
yoor_row = None
for r in range(header_row + 1, ws.max_row + 1):
    val = ws.cell(row=r, column=empresa_col).value
    if val and "yoor" in str(val).lower():
        yoor_row = r
        old_empresa = val
        old_fin = ws.cell(row=r, column=fin_col).value
        print(f"\n🎯 Found Yoor at row {r}")
        print(f"   empresa BEFORE : {old_empresa!r}")
        print(f"   fecha_fin BEFORE: {old_fin!r}")

        # Update
        ws.cell(row=r, column=empresa_col).value = "Yoorco (Yoor Company)"
        ws.cell(row=r, column=fin_col).value = "02/2023"

        print(f"   empresa AFTER  : {ws.cell(row=r, column=empresa_col).value!r}")
        print(f"   fecha_fin AFTER: {ws.cell(row=r, column=fin_col).value!r}")
        break

if not yoor_row:
    raise SystemExit("❌ Yoor row not found in '02 Experiencia'")

wb.save(MASTER)
print(f"\n✅ Saved: {MASTER.relative_to(ROOT)}")
print("\nNext: python scripts/export_master_json.py")
