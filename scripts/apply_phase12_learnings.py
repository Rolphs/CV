"""
Fase 12 · Apply learnings from PPG/Comex application session.

Adds to master:
  M1 · Logro L75 — Yoorco team & founder mindset (was deliberately omitted for
       PPG, but valuable for future founder-friendly recipes like R04).
  M2 · Recipe R07 — Director Data Intelligence · Corporate MX/LATAM
       (no existing recipe matched PPG-like corporate director roles).
  M3 · Anti-pattern A14 — never invent target industry experience.
  M4 · Anti-pattern A15 — when no single recipe fits, build a hybrid and
       document it in application.yaml.
  M5 · Anti-pattern A16 — mirror JD verbs verbatim, never invent content.

Idempotent: skips rows whose ID already exists.
"""
import shutil
import sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "data" / "master" / "cv_master_raul_mercado.xlsx"
BACKUP = ROOT / "data" / "master" / "backups" / "v12_pre_learnings.xlsx"


# ─────────────────────────── Payloads ───────────────────────────

LOGRO_L75 = {
    "ID Logro": 75,
    "ID Empresa": 2,  # Yoorco / Yoor Company
    "Empresa": "Yoorco (Yoor Company)",
    "Tipo": "Logro",
    "Logro / Resp. ES": (
        "Construí y lideré un pequeño equipo de producto, validando mentalidad "
        "emprendedora, accountability comercial y ejecución end-to-end desde "
        "la idea hasta el MVP."
    ),
    "Achievement EN": (
        "Built and led a small product team, validating founder mindset, "
        "commercial accountability and end-to-end execution from idea to MVP."
    ),
    "Impacto": (
        "Demuestra capacidad de bootstrapping y liderazgo desde cero, útil "
        "para recipes con sesgo founder/startup (R04 Tech-AdTech)."
    ),
    "Tiene Métrica": "No",
    "Métrica / KPI": "",
    "Categoría": "Liderazgo",
    "Tags Industria": "['Tech', 'Startup']",
    "Tags Habilidad": "['Liderazgo', 'Producto', 'Emprendimiento']",
    "Seniority Fit": "['director', 'exec']",
    "Story / Evidence": (
        "Equipo de 2 personas reportando directo al fundador. Pivots semanales, "
        "ciclo idea→prototipo→feedback en <2 semanas."
    ),
    "Pair With (IDs)": "18, 63",  # Yoorco's NPS logros
    "Incluir CV ES": None,
    "Incluir CV EN": None,
}


RECIPE_R07 = {
    "Recipe ID": "R07",
    "Nombre / Target Role": (
        "Director Data Intelligence · Corporate FMCG / Coatings / Retail · MX & LATAM"
    ),
    "Cuándo usar": (
        "PPG, Comex, Sherwin-Williams, Bimbo, FEMSA, Coca-Cola FEMSA, "
        "Nestlé MX, Walmart MX, Liverpool. Roles Director/Sr Director de "
        "BI / Data Intelligence / Insights donde el énfasis está en pricing, "
        "promotion, profitability y dashboards para decisiones comerciales."
    ),
    "Perfil Narrativo (idx)": "CPG / Retail",
    "Logros a usar (IDs)": (
        # Walmart Scintilla (current, CPG retail analytics)
        "67, "
        # Canela Media (AI scalable systems, F500 clients, NPS)
        "51, 3, 50, 4, "
        # Altazor (AI pipelines, predictive KPIs)
        "64, 65, "
        # Televisa (large-scale research portfolio + board)
        "54, 9, "
        # Kantar Millward Brown (CPG brand health, pricing)
        "10, 11, 12, 55, 56, "
        # TIBA / KIO (B2B competitive intelligence, board reports, pricing)
        "13, 14, 27, "
        # Nodo (UX pioneer, multi-industry C-level)
        "33, 15"
    ),
    "Skills (categorías)": "Research, Analytics / IA, Estrategia, Liderazgo",
    "Notas críticas para el LLM": (
        "Idioma: EN o ES según JD. Énfasis: pricing & promotion strategies, "
        "profitability protection, KPI ecosystems, board-level reporting, "
        "cross-functional collaboration (marketing+finance+commercial+ops). "
        "Header debe llevar 'Director' explícito (NUNCA 'Manager'). "
        "Walmart Scintilla rebautizado como 'Strategic Analytics Lead' (no "
        "'Account Manager') por dignidad y consistencia con seniority claim. "
        "Incluir Yoorco (founder mindset) y entrada Independent Consultant si "
        "hay gap cronológico real. NO mencionar industria specific (paint/"
        "coatings/etc) si no la tienes en master — pivota a 'consumer-facing "
        "categories' o 'transversal data intelligence'. Mirror verbos del JD "
        "verbatim (oversee, guide, foster, ensure, identify). Pages target: 3."
    ),
}


ANTI_PATTERNS_NEW = [
    {
        "ID": "A14",
        "Categoría": "Industria inventada",
        "Evita esto": (
            "Mencionar experiencia en industrias específicas que NO están en el "
            "master (ej. 'architectural coatings', 'paint industry', 'aerospace')."
        ),
        "En lugar usa": (
            "Generaliza a categorías que SÍ tienes: 'consumer-facing categories', "
            "'transversal data intelligence', 'CPG analytics' (si tienes Kantar/"
            "Scintilla). El reclutador valora pattern-matching, no claims falsas."
        ),
        "Por qué": (
            "Cualquier industria afirmada que no puedas defender en entrevista "
            "destruye credibilidad. Aprendido con PPG Comex (architectural "
            "coatings) — nunca mencionado, match rate igual quedó 90%."
        ),
    },
    {
        "ID": "A15",
        "Categoría": "Recipe forzado",
        "Evita esto": (
            "Forzar un recipe que no calza 100% sin documentar las desviaciones "
            "en application.yaml (ej. usar R03 'Manager' para puesto Director)."
        ),
        "En lugar usa": (
            "Si ningún recipe calza, construye un híbrido (ej. R03+R05) y "
            "documenta en application.yaml el racional: qué logros se "
            "inyectaron de cada uno, qué seniority se ajustó, qué énfasis "
            "industria se pivoteó."
        ),
        "Por qué": (
            "PPG no encajaba en R01-R06 (director-level corporate FMCG con "
            "pricing). Inventamos híbrido R03+R05 → 90% match rate y CV "
            "defensible. La trazabilidad permite repetir el patrón."
        ),
    },
    {
        "ID": "A16",
        "Categoría": "Mirror JD verbatim",
        "Evita esto": (
            "Usar tu lenguaje genérico cuando el JD usa verbos específicos "
            "(ej. 'led' cuando el JD pide 'oversee'/'guide'/'foster')."
        ),
        "En lugar usa": (
            "Mirror exacto de los verbos del JD en tus bullets — sin inventar "
            "contenido nuevo, solo reformulando logros reales del master con "
            "el vocabulario del puesto target."
        ),
        "Por qué": (
            "ATS y reclutadores escanean por keyword overlap. PPG: subimos de "
            "60% → 92% match rate solo reescribiendo bullets existentes con "
            "verbos del JD (oversee, guide, foster, identify, ensure)."
        ),
    },
]


# ─────────────────────────── Engine ───────────────────────────


def _header_map(ws, header_row: int = 2) -> dict[str, int]:
    """Build {column_name: col_index} from a header row."""
    return {
        ws.cell(row=header_row, column=c).value: c
        for c in range(1, 30)
        if ws.cell(row=header_row, column=c).value
    }


def _existing_ids(ws, id_col: int, header_row: int) -> set:
    """Set of all non-empty values in id_col below header_row."""
    out = set()
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=id_col).value
        if v is not None and v != "":
            out.add(v)
    return out


def _first_empty_row(ws, header_row: int, id_col: int) -> int:
    """First row index where id_col is empty (= where to append)."""
    for r in range(header_row + 1, ws.max_row + 2):
        if ws.cell(row=r, column=id_col).value in (None, ""):
            return r
    return ws.max_row + 1


def _append_row(ws, payload: dict, headers: dict[str, int], target_row: int):
    """Write each payload field into its mapped column on target_row."""
    for field_name, value in payload.items():
        col = headers.get(field_name)
        if col is None:
            print(f"   ⚠ Column '{field_name}' not found in sheet — skipped")
            continue
        ws.cell(row=target_row, column=col).value = value


def apply_logro(wb, payload: dict, id_field: str = "ID Logro"):
    ws = wb["03 Logros"]
    headers = _header_map(ws)
    id_col = headers[id_field]
    existing = _existing_ids(ws, id_col, header_row=2)
    new_id = payload[id_field]
    if new_id in existing:
        print(f"   ⏭  L{new_id} already in master — skipped")
        return False
    row = _first_empty_row(ws, header_row=2, id_col=id_col)
    _append_row(ws, payload, headers, row)
    print(f"   ✅ Added L{new_id} at row {row}")
    return True


def apply_recipe(wb, payload: dict, id_field: str = "Recipe ID"):
    ws = wb["12 Recipes"]
    headers = _header_map(ws)
    id_col = headers[id_field]
    existing = _existing_ids(ws, id_col, header_row=2)
    new_id = payload[id_field]
    if new_id in existing:
        print(f"   ⏭  {new_id} already in master — skipped")
        return False
    row = _first_empty_row(ws, header_row=2, id_col=id_col)
    _append_row(ws, payload, headers, row)
    print(f"   ✅ Added {new_id} at row {row}")
    return True


def apply_anti_pattern(wb, payload: dict, id_field: str = "ID"):
    ws = wb["13 Anti-Patterns"]
    headers = _header_map(ws)
    id_col = headers[id_field]
    existing = _existing_ids(ws, id_col, header_row=2)
    new_id = payload[id_field]
    if new_id in existing:
        print(f"   ⏭  {new_id} already in master — skipped")
        return False
    row = _first_empty_row(ws, header_row=2, id_col=id_col)
    _append_row(ws, payload, headers, row)
    print(f"   ✅ Added {new_id} at row {row}")
    return True


# ─────────────────────────── Main ───────────────────────────


def main():
    print(f"📂 Loading {MASTER.name}")
    BACKUP.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(MASTER, BACKUP)
    print(f"💾 Backup: {BACKUP.relative_to(ROOT)}")

    wb = load_workbook(MASTER)

    changed = 0
    print("\n🍳 M1 · Logro L75 (Yoorco team & founder mindset)")
    changed += int(apply_logro(wb, LOGRO_L75))

    print("\n🧪 M2 · Recipe R07 (Director Data Intelligence Corporate)")
    changed += int(apply_recipe(wb, RECIPE_R07))

    print("\n🚫 M3-M5 · Anti-patterns A14, A15, A16")
    for ap in ANTI_PATTERNS_NEW:
        changed += int(apply_anti_pattern(wb, ap))

    if changed:
        wb.save(MASTER)
        print(f"\n✅ Saved {changed} change(s) to {MASTER.name}")
        print("\nNext:  python scripts/export_master_json.py")
    else:
        print("\nℹ Nothing to add — all IDs already present.")


if __name__ == "__main__":
    main()
