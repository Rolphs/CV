"""
Fase 13 · Refactor Bricko in master per Rolph 2026-05-17 clarifications.

Changes:
  1. Empresa #6 "Independiente (BRICKO / EIDOS)" →  "Bricko"
     - tipo: Freelance → Empresa
     - puesto: Strategic Consultant → Managing Partner
     - personas_a_cargo: 0 → 5-10
     - reporta_a: Directo al cliente → Co-founders / Partners
     - descripcion_*: focused on Magento-based e-commerce for SMEs/startups
     - industria: Consultoría / Digital → E-commerce / Magento
  2. Empresa NUEVA #25 "Eidos" (separated from Bricko, placeholder for future
     enrichment).
  3. Refactor logro L17 to mention Magento and remove "EIDOS" reference.
  4. Add L76 (qualitative: managing partner / 5-10 team / Magento e-commerce).
  5. Add L77 (qualitative: commercial + UX research + A/B testing leadership).

NO METRICS INVENTED — Rolph chose qualitative-only per A14 anti-pattern.
"""
import shutil
import sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "data" / "master" / "cv_master_raul_mercado.xlsx"
BACKUP = ROOT / "data" / "master" / "backups" / "v13_pre_bricko_refactor.xlsx"


# ─────────────────────────── Payloads ───────────────────────────


BRICKO_UPDATES = {
    "Empresa / Company": "Bricko",
    "Industria / Industry": "E-commerce / Magento",
    "Tipo": "Empresa",
    "Puesto ES": "Managing Partner",
    "Puesto EN": "Managing Partner",
    "Descripción Empresa ES": "Empresa de desarrollo e-commerce especializada en implementaciones Magento para PYMEs y startups durante la aceleración digital del COVID-19.",
    "Descripción Empresa EN": "E-commerce development company specializing in Magento-based implementations for SMEs and startups during the COVID-19 digital acceleration.",
    "Personas a cargo": "5-10",
    "Reporta a": "Co-founders / Partners",
    "Dept / Área": "Dirección / Operaciones",
}


EIDOS_NEW = {
    "ID": 28,
    "Empresa / Company": "Eidos",
    "Industria / Industry": "TBD",
    "País": "México",
    "Puesto ES": "TBD",
    "Puesto EN": "TBD",
    "Nivel": "TBD",
    "Tipo": "TBD",
    "Fecha Inicio": "TBD",
    "Fecha Fin": "TBD",
    "Actual": False,
    "Descripción Empresa ES": "TBD — placeholder creado 2026-05-17 cuando se separó de Bricko. Pendiente input de Rolph: fechas, rol, alcance y logros.",
    "Descripción Empresa EN": "TBD — placeholder created 2026-05-17 when separated from Bricko. Awaiting Rolph's input on dates, role, scope and achievements.",
}


L17_REFACTOR = (
    "Designed and executed Magento-based e-commerce implementations and "
    "digital-transformation projects for SMEs and startups during the "
    "COVID-19 pandemic at Bricko."
)


LOGRO_L76 = {
    "ID Logro": 76,
    "ID Empresa": 6,
    "Empresa": "Bricko",
    "Tipo": "Logro",
    "Logro / Resp. ES": (
        "Co-lideré Bricko como Managing Partner, construyendo y operando un "
        "equipo de 5-10 personas que entregó implementaciones de e-commerce "
        "Magento para PYMEs y startups durante la aceleración digital del "
        "COVID-19."
    ),
    "Achievement EN": (
        "Co-led Bricko as Managing Partner, building and operating a 5-10 "
        "person team that delivered Magento-based e-commerce implementations "
        "for SMEs and startups during the COVID-19 digital acceleration."
    ),
    "Impacto": (
        "Demuestra liderazgo de empresa propia con equipo, ejecución comercial "
        "y delivery técnico en stack e-commerce. Cubre cronología 2020-2021 "
        "con narrativa de Founder/Partner sustentada."
    ),
    "Tiene Métrica": "No",
    "Métrica / KPI": "",
    "Categoría": "Liderazgo",
    "Tags Industria": "['E-commerce', 'Tech', 'SMB']",
    "Tags Habilidad": "['Liderazgo', 'E-commerce', 'Magento', 'Gestión de equipo']",
    "Seniority Fit": "['director', 'exec']",
    "Story / Evidence": (
        "Bricko = empresa formal (no paraguas freelance). 5-10 personas. "
        "Reporting horizontal a co-founders. Eidos era separada (placeholder #25)."
    ),
    "Pair With (IDs)": "17, 61, 77",
    "Incluir CV ES": None,
    "Incluir CV EN": None,
}


LOGRO_L77 = {
    "ID Logro": 77,
    "ID Empresa": 6,
    "Empresa": "Bricko",
    "Tipo": "Logro",
    "Logro / Resp. ES": (
        "Lideré la estrategia comercial, relaciones con clientes y calidad de "
        "entrega del portafolio Bricko, aplicando frameworks de research-led "
        "UX design, A/B testing y optimización de conversión a cada lanzamiento "
        "de e-commerce."
    ),
    "Achievement EN": (
        "Owned commercial strategy, client relationships and delivery quality "
        "across the Bricko portfolio, applying research-led UX design, A/B "
        "testing and conversion-optimization frameworks to every e-commerce "
        "launch."
    ),
    "Impacto": (
        "Conecta el origen Research/UX con la ejecución comercial e-commerce. "
        "Útil para recipes que valoran T-shaped leadership (research + business + tech)."
    ),
    "Tiene Métrica": "No",
    "Métrica / KPI": "",
    "Categoría": "Estrategia comercial",
    "Tags Industria": "['E-commerce', 'SMB', 'Startup']",
    "Tags Habilidad": "['UX', 'A/B Testing', 'Conversión', 'Estrategia Comercial']",
    "Seniority Fit": "['director', 'exec']",
    "Story / Evidence": (
        "Cada cliente Bricko recibía un proceso research-first: descubrimiento, "
        "UX design, prototipo, A/B testing y launch. Diferenciador vs agencias "
        "Magento puras (solo dev)."
    ),
    "Pair With (IDs)": "17, 61, 76",
    "Incluir CV ES": None,
    "Incluir CV EN": None,
}


# ─────────────────────────── Engine ───────────────────────────


def _header_map(ws, header_row: int = 2) -> dict[str, int]:
    return {
        ws.cell(row=header_row, column=c).value: c
        for c in range(1, 30)
        if ws.cell(row=header_row, column=c).value
    }


def _existing_ids(ws, id_col: int, header_row: int) -> set:
    out = set()
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=id_col).value
        if v not in (None, ""):
            out.add(v)
    return out


def _first_empty_row(ws, header_row: int, id_col: int) -> int:
    for r in range(header_row + 1, ws.max_row + 2):
        if ws.cell(row=r, column=id_col).value in (None, ""):
            return r
    return ws.max_row + 1


def _append_row(ws, payload: dict, headers: dict[str, int], target_row: int):
    for field_name, value in payload.items():
        col = headers.get(field_name)
        if col is None:
            print(f"   ⚠ Column '{field_name}' not in sheet — skipped")
            continue
        ws.cell(row=target_row, column=col).value = value


def _update_row(ws, row: int, payload: dict, headers: dict[str, int]):
    for field_name, value in payload.items():
        col = headers.get(field_name)
        if col is None:
            print(f"   ⚠ Column '{field_name}' not in sheet — skipped")
            continue
        ws.cell(row=row, column=col).value = value


# ─────────────────────────── Operations ───────────────────────────


def update_bricko_empresa(wb):
    ws = wb["02 Experiencia"]
    headers = _header_map(ws)
    empresa_col = headers.get("Empresa / Company")

    target_row = None
    for r in range(3, ws.max_row + 1):
        v = ws.cell(row=r, column=empresa_col).value
        if v and "brick" in str(v).lower():
            target_row = r
            break

    if target_row is None:
        print("   ❌ Bricko empresa row not found")
        return False

    old_name = ws.cell(row=target_row, column=empresa_col).value
    print(f"   📍 Bricko row={target_row} · current name: {old_name!r}")
    _update_row(ws, target_row, BRICKO_UPDATES, headers)
    print(f"   ✅ Updated to: 'Bricko' · Managing Partner · 5-10 people · Magento e-commerce")
    return True


def insert_eidos_placeholder(wb):
    ws = wb["02 Experiencia"]
    headers = _header_map(ws)
    id_col = headers["ID"]
    existing = _existing_ids(ws, id_col, header_row=2)
    new_id = EIDOS_NEW["ID"]
    if new_id in existing:
        print(f"   ⏭  Eidos #{new_id} already exists — skipped")
        return False
    row = _first_empty_row(ws, header_row=2, id_col=id_col)
    _append_row(ws, EIDOS_NEW, headers, row)
    print(f"   ✅ Added Eidos (#{new_id}) at row {row} as TBD placeholder")
    return True


def refactor_l17(wb):
    ws = wb["03 Logros"]
    headers = _header_map(ws)
    id_col = headers["ID Logro"]
    en_col = headers["Achievement EN"]
    es_col = headers["Logro / Resp. ES"]
    empresa_col = headers["Empresa"]

    target_row = None
    for r in range(3, ws.max_row + 1):
        if ws.cell(row=r, column=id_col).value == 17:
            target_row = r
            break

    if target_row is None:
        print("   ❌ L17 not found")
        return False

    old_en = ws.cell(row=target_row, column=en_col).value
    print(f"   📍 L17 BEFORE: {old_en[:100]}...")
    ws.cell(row=target_row, column=en_col).value = L17_REFACTOR
    ws.cell(row=target_row, column=empresa_col).value = "Bricko"
    # ES: simple translation alignment
    ws.cell(row=target_row, column=es_col).value = (
        "Diseñé y ejecuté implementaciones de e-commerce Magento y proyectos "
        "de transformación digital para PYMEs y startups durante la pandemia "
        "de COVID-19 en Bricko."
    )
    print(f"   ✅ L17 AFTER: {L17_REFACTOR[:100]}...")
    return True


def add_logro(wb, payload):
    ws = wb["03 Logros"]
    headers = _header_map(ws)
    id_col = headers["ID Logro"]
    existing = _existing_ids(ws, id_col, header_row=2)
    new_id = payload["ID Logro"]
    if new_id in existing:
        print(f"   ⏭  L{new_id} already exists — skipped")
        return False
    row = _first_empty_row(ws, header_row=2, id_col=id_col)
    _append_row(ws, payload, headers, row)
    print(f"   ✅ Added L{new_id} at row {row}")
    return True


# ─────────────────────────── Main ───────────────────────────


def main():
    print(f"📂 Loading {MASTER.name}")
    BACKUP.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(MASTER, BACKUP)
    print(f"💾 Backup: {BACKUP.relative_to(ROOT)}")

    wb = load_workbook(MASTER)
    changed = 0

    print("\n🏢 1 · Refactor empresa #6 (Bricko)")
    changed += int(update_bricko_empresa(wb))

    print("\n🏢 2 · Insert empresa Eidos (#25) placeholder")
    changed += int(insert_eidos_placeholder(wb))

    print("\n🍳 3 · Refactor L17 (remove EIDOS mention, add Magento)")
    changed += int(refactor_l17(wb))

    print("\n🍳 4 · Add L76 (Managing Partner + 5-10 team)")
    changed += int(add_logro(wb, LOGRO_L76))

    print("\n🍳 5 · Add L77 (UX research + A/B + commercial strategy)")
    changed += int(add_logro(wb, LOGRO_L77))

    if changed:
        wb.save(MASTER)
        print(f"\n✅ Saved {changed} change(s) to {MASTER.name}")
        print("\nNext:  python scripts/export_master_json.py")
    else:
        print("\nℹ Nothing changed — already applied.")


if __name__ == "__main__":
    main()
