"""
Fase 6 · Polish profundo del master.

Aplica TODOS los fixes detectados en la auditoría:
  A. Consolidación KIO/TIBA: borra IDs 7 y 12 (duplicados), reasigna logros
  B. Consolidación NODO: borra ID 13 (duplicado), reasigna logros a ID 8
  C. Televisa fecha inicio: ID 4 cambia 08/2017 → 04/2018 (era ascenso del 27)
  D. Normalización nombres empresa en 9 logros (Kantar, Independiente)
  E. Reescritura de 63 logros con verbo de acción en ES y EN
  F. Expansión de 6 logros muy cortos (L25, L29, L32, L36, L44, L46)
  G. 6 logros nuevos para puestos sin logros (L69-L74)
  H. Bonus: arregla anacronismo (L60 ChatGPT/GPT-4 estaba en KIO 2012-2014,
     se mueve a Canela Media donde corresponde por época)
"""
from __future__ import annotations
import shutil
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "data" / "master" / "cv_master_raul_mercado.xlsx"
BACKUP = ROOT / "data" / "master" / "backups" / "v6_pre_polish.xlsx"


# ─── E + F: Reescritura de logros (ID → (new_es, new_en)) ──────────────────
# Cada uno reescrito para iniciar con verbo de acción punchy
LOGRO_REWRITES = {
    1: ("Automaticé el 90% de los reportes de investigación mediante IA, eliminando tareas manuales repetitivas.",
        "Automated 90% of research reports through AI, eliminating repetitive manual tasks."),
    2: ("Reduje en 30% el time-to-market de insights para clientes mediante pipelines de datos automatizados.",
        "Reduced client insights time-to-market by 30% through automated data pipelines."),
    3: ("Diseñé una arquitectura escalable de insights con modelos de atribución basados en IA para clientes Fortune 500.",
        "Designed a scalable insights architecture with AI-based attribution models for Fortune 500 clients."),
    4: ("Lideré un equipo de investigación multidisciplinario (datos, cuali, cuanti) para clientes de medios en español en EE.UU.",
        "Led a multidisciplinary research team (data, qual, quant) serving U.S. Spanish-language media clients."),
    5: ("Posicioné a Canela Media como referente técnico en investigación de audiencias hispanas frente a anunciantes.",
        "Positioned Canela Media as a technical reference in Hispanic audience research for advertisers."),
    6: ("Incrementé en 5% la audiencia prime-time mediante investigación estratégica de contenidos y rediseño de parrilla.",
        "Increased prime-time audience by 5% through strategic content research and programming redesign."),
    7: ("Impulsé un crecimiento del 15% en el segmento juvenil mediante estudios de insights y recomendaciones editoriales basadas en datos.",
        "Drove 15% growth in the youth segment through insights studies and data-driven editorial recommendations."),
    8: ("Lideré los estudios de audiencias para el lanzamiento y fortalecimiento de las plataformas BLIM y Las Estrellas.",
        "Led audience research for the launch and growth of BLIM and Las Estrellas streaming platforms."),
    9: ("Diseñé la estrategia integral de investigación social y de audiencias para el principal grupo de medios de Latinoamérica.",
        "Designed the comprehensive social and audience research strategy for Latin America's leading media group."),
    10: ("Incrementé en 15% la satisfacción de clientes mediante el rediseño de procesos de entrega y comunicación de resultados.",
         "Increased client satisfaction by 15% through redesigned delivery processes and results communication."),
    11: ("Reduje las cancelaciones de proyectos por debajo del 5% mediante protocolos de gestión proactiva de cuentas.",
         "Reduced project cancellations below 5% through proactive account management protocols."),
    12: ("Gestioné un portafolio de estudios cuali-cuantitativos para marcas globales de CPG, telecom y retail.",
         "Managed a portfolio of qual-quant studies for global CPG, telecom and retail brands."),
    13: ("Dupliqué la base anual de clientes mediante inteligencia competitiva y estudios de mercado B2B para servicios TIC.",
         "Doubled the annual client base through competitive intelligence and B2B market studies for ICT services."),
    14: ("Entregué insights estratégicos de mercado a dirección ejecutiva y al consejo de administración.",
         "Delivered strategic market insights to executive leadership and the board of directors."),
    15: ("Encabecé la introducción de metodologías UX en México, desarrollando los primeros frameworks de investigación de usabilidad digital del país.",
         "Pioneered UX methodologies in Mexico, developing the country's first digital usability research frameworks."),
    16: ("Gestioné cuentas nacionales e internacionales con estudios multicategoría y multipaís en más de 10 mercados.",
         "Managed national and international accounts with multi-category and multi-country studies across 10+ markets."),
    17: ("Diseñé y ejecuté estrategias de transformación digital para PYMEs y startups durante la pandemia COVID-19 a través de BRICKO y EIDOS.",
         "Designed and executed digital transformation strategies for SMEs and startups during the COVID-19 pandemic via BRICKO and EIDOS."),
    18: ("Incrementé en 30% el NPS de usuarios mediante rediseño de experiencia centrado en el usuario y potenciado por IA.",
         "Increased user NPS by 30% through AI-powered user-centered experience redesign."),
    19: ("Conduje proyectos como consultor independiente y partner estratégico para múltiples clientes.",
         "Led projects as independent consultant and strategic partner for multiple clients."),
    20: ("Diseñé soluciones adaptadas por segmento, integrando equipos ad-hoc según las necesidades de cada cliente.",
         "Designed segment-tailored solutions, integrating ad-hoc teams based on each client's needs."),
    21: ("Lideré proyectos de investigación y marketing, formando equipos ad-hoc según cada cliente y colaborando con agencias como socio estratégico.",
         "Led research and marketing projects, building ad-hoc teams per client and partnering with agencies as a strategic ally."),
    22: ("Gané contratos clave frente a empresas internacionales en licitaciones de soporte de infraestructura TIC.",
         "Won key contracts against international enterprises in ICT infrastructure support bids."),
    23: ("Personalicé soluciones tecnológicas alineadas a las necesidades específicas de cada cliente corporativo.",
         "Customized technology solutions aligned with each corporate client's specific needs."),
    24: ("Traduje necesidades del cliente en propuestas de alto valor para contratos de soporte de infraestructura TIC.",
         "Translated client needs into high-value proposals for ICT infrastructure support contracts."),
    25: ("Dupliqué la base anual de clientes mediante el diseño de un portafolio competitivo basado en análisis sectorial y la conceptualización de nuevos servicios TIC.",
         "Doubled the annual client base through a competitive portfolio design grounded in sector analysis and new ICT service conceptualization."),
    26: ("Consolidé un portafolio de servicios competitivo basado en análisis sectorial profundo y benchmarking de la industria.",
         "Consolidated a competitive service portfolio based on deep sector analysis and industry benchmarking."),
    27: ("Lideré la estrategia de endomarketing y análisis competitivo, generando reportes trimestrales para el consejo directivo.",
         "Led the endo-marketing strategy and competitive analysis, generating quarterly reports for the board of directors."),
    28: ("Consolidé el posicionamiento institucional de la marca mediante una estrategia de comunicación corporativa coherente.",
         "Consolidated institutional brand positioning through a coherent corporate communication strategy."),
    29: ("Diseñé y ejecuté la estrategia integral de marketing B2B para servicios de gestión y soporte de infraestructura TIC, ampliando el reconocimiento en el mercado empresarial.",
         "Designed and executed the comprehensive B2B marketing strategy for ICT management and infrastructure support services, expanding enterprise market recognition."),
    30: ("Generé mayor visibilidad interna y sentido de pertenencia mediante una estrategia de endomarketing alineada a los valores corporativos.",
         "Built greater internal visibility and belonging through an endo-marketing strategy aligned with corporate values."),
    31: ("Diseñé portafolios de servicio, desarrollé presentaciones para el Consejo y conceptualicé nuevos servicios TIC con enfoque en endomarketing.",
         "Designed service portfolios, developed board-level presentations and conceptualized new ICT services with an endo-marketing focus."),
    32: ("Encabecé la introducción de la investigación de usabilidad y UX en México, desarrollando los primeros frameworks aplicados a software y tecnología digital del país.",
         "Pioneered usability and UX research in Mexico, developing the country's first frameworks applied to software and digital technology."),
    33: ("Presenté resultados de investigación ante ejecutivos C-level y comités directivos de empresas multinacionales.",
         "Presented research results to C-level executives and steering committees of multinational companies."),
    34: ("Lideré estudios internacionales multipaís coordinando equipos locales en más de 10 mercados latinoamericanos.",
         "Led international multi-country studies coordinating local teams across 10+ Latin American markets."),
    35: ("Moderé estudios cualitativos, conduje análisis y presenté resultados aplicando metodologías de usabilidad, UX y redes sociales.",
         "Moderated qualitative studies, conducted analysis and presented results applying usability, UX and social media methodologies."),
    36: ("Establecí un modelo de negocio sustentable elaborando productos artesanales de salud y belleza con ingredientes naturales y procesos eco-conscientes.",
         "Established a sustainable business model crafting artisanal health and beauty products with natural ingredients and eco-conscious processes."),
    37: ("Integré las funciones de diseño, producción y comercialización en una microempresa propia con foco artesanal.",
         "Integrated design, production and commercialization functions in a self-owned artisanal micro-business."),
    38: ("Desarrollé, produje y comercialicé productos artesanales de salud y belleza, gestionando el ciclo completo del negocio.",
         "Developed, produced and commercialized artisanal health and beauty products, managing the full business cycle."),
    39: ("Brindé soporte técnico clave para la transmisión estable de estudios remotos en plataformas de streaming pioneras.",
         "Provided key technical support for stable streaming of remote research studies on pioneering platforms."),
    40: ("Operé la transmisión y soporte técnico de sesiones de investigación en línea para clientes internacionales.",
         "Operated streaming and technical support for online research sessions for international clients."),
    41: ("Participé en producciones publicitarias relevantes para clientes nacionales reconocidos como asistente de dirección y arte.",
         "Contributed to notable advertising productions for renowned national clients as assistant director and art assistant."),
    42: ("Realicé visualizaciones, asistencias de dirección y dirección de arte para campañas publicitarias de alto perfil.",
         "Produced visualizations, assistant directing and art direction for high-profile advertising campaigns."),
    43: ("Produje contenido audiovisual end-to-end para clientes corporativos y eventos internacionales.",
         "Produced end-to-end audiovisual content for corporate clients and international events."),
    44: ("Desarrollé productos multimedia y experiencias audiovisuales integrales, incluyendo coordinación logística para la cumbre APEC México 2002.",
         "Developed multimedia products and end-to-end audiovisual experiences, including logistics coordination for the APEC Mexico 2002 summit."),
    45: ("Lideré producción audiovisual y desarrollo de sitios web como emprendedor en una startup creativa.",
         "Led audiovisual production and web development as an entrepreneur in a creative startup."),
    46: ("Construí fundamentos técnicos sólidos en hardware, soporte de TI y sistemas de punto-de-venta atendiendo a clientes empresariales desde edad temprana.",
         "Built solid technical foundations in hardware, IT support and point-of-sale systems serving enterprise clients from an early age."),
    47: ("Adquirí fundamentos en hardware, diagnóstico y soporte técnico atendiendo a clientes empresariales.",
         "Gained foundations in hardware, diagnostics and technical support serving enterprise clients."),
    48: ("Realicé servicio y reparación de computadoras, cajas registradoras y periféricos para clientes empresariales.",
         "Performed service and repair of computers, cash registers and peripherals for enterprise clients."),
    50: ("Aumenté en 30% el NPS de clientes mediante el diseño de experiencia potenciado por IA.",
         "Increased client NPS by 30% through AI-enhanced customer experience design."),
    51: ("Lideré la integración de flujos de trabajo con IA y sistemas escalables de insights, alcanzando 90% de automatización en reportería e incrementando velocidad-de-insight en 30%.",
         "Led the integration of AI-driven workflows and scalable insight systems, achieving 90% automation in reporting and increasing speed-to-insight by 30%."),
    52: ("Generé más de $2M en impacto de ingresos mediante estrategias de contenido basadas en insights y campañas de targeting juvenil.",
         "Delivered $2M+ in revenue impact through insight-driven content strategies and youth targeting initiatives."),
    54: ("Lideré más de 150 proyectos combinando investigación cualitativa, encuestas y analytics conductuales para informar programación de TV, estrategia digital y desarrollo de audiencias.",
         "Led 150+ projects combining qualitative research, surveys and behavioral analytics to shape TV programming, digital strategy and audience development."),
    55: ("Gestioné portafolios globales de clientes FMCG, reduciendo churn por debajo del 5% e incrementando satisfacción en 15% YoY.",
         "Managed global FMCG client portfolios, reducing churn below 5% and increasing satisfaction by 15% YoY."),
    56: ("Logré menos del 5% de churn de clientes y un crecimiento del 60% en cuentas estratégicas mediante diseño innovador de estudios y entrega consistente.",
         "Achieved under 5% client churn and 60% strategic-account growth through innovative study design and consistent delivery."),
    58: ("Generé más de $5M en ventas y dupliqué la base anual de clientes mediante estrategias de adquisición basadas en insights.",
         "Delivered $5M+ in sales and consistently doubled the annual client base through insight-driven acquisition strategies."),
    60: ("Apliqué estratégicamente ChatGPT, Claude y GPT-4 para acelerar y sintetizar flujos de investigación de audiencias.",
         "Strategically applied ChatGPT, Claude and GPT-4 to accelerate and synthesize audience research workflows."),
    61: ("Gestioné un pipeline robusto de clientes con 100% de entrega a tiempo y alta satisfacción sostenida basada en rigor de diseño, storytelling y profundidad de insight.",
         "Managed a robust client pipeline with 100% on-time delivery and consistently high satisfaction based on design rigor, storytelling and insight depth."),
    63: ("Incrementé en 30% el Net Promoter Score aplicando insights del cliente para optimizar UX y engagement.",
         "Drove a 30% increase in Net Promoter Score by applying customer insights to optimize UX and engagement."),
    64: ("Construí pipelines de insights potenciados por IA, reduciendo los tiempos de entrega hasta en 90%.",
         "Built AI-powered insight pipelines reducing reporting turnaround by up to 90%."),
    65: ("Diseñé ecosistemas predictivos de KPIs que permitieron optimización en tiempo real del desempeño de medios.",
         "Designed predictive KPI ecosystems enabling real-time media performance optimization."),
    66: ("Lideré la integración cross-funcional entre equipos de producto, ciencia de datos e investigación en proyectos de consultoría estratégica.",
         "Led cross-functional integration between product, data science and research teams in strategic consulting projects."),
    67: ("Gestioné cuentas CPG estratégicas en la plataforma de retail analytics de Walmart, traduciendo datos transaccionales en decisiones de surtido, precio y promoción.",
         "Managed strategic CPG accounts on Walmart's retail analytics platform, translating transactional data into assortment, pricing and promotion decisions."),
    68: ("Elegí deliberadamente cercanía a datos conductuales a escala retail por encima de seniority de título, alineando trayectoria con el siguiente salto a métodos conductuales aplicados.",
         "Deliberately chose proximity to retail-scale behavioral data over title seniority, aligning career path with applied behavioral methods."),
}


# ─── A + B: reasignación de ID Empresa (logro_id → new_emp_id) ─────────────
# Tras consolidar KIO/TIBA y NODO
LOGRO_EMP_REASSIGN = {
    # Antes ID 7 (KIO Networks - delete):
    13: (10, "TIBA IT Services (KIO Networks)"),   # duplicación clientes → sales role
    14: (11, "TIBA IT Services (KIO Networks)"),   # insights to board → BI/Mkt role
    58: (10, "TIBA IT Services (KIO Networks)"),   # $5M sales → sales role
    60: (1,  "Canela Media"),                      # ChatGPT/GPT-4 era 2023+, mover a Canela
    # Antes ID 12 (TIBA Strategic - delete):
    28: (11, "TIBA IT Services (KIO Networks)"),
    29: (11, "TIBA IT Services (KIO Networks)"),
    30: (11, "TIBA IT Services (KIO Networks)"),
    31: (11, "TIBA IT Services (KIO Networks)"),
    # Antes ID 13 (NODO duplicado - delete):
    32: (8, "Nodo Research"),
    33: (8, "Nodo Research"),
    34: (8, "Nodo Research"),
    35: (8, "Nodo Research"),
}

# ─── D: normalización de nombres de empresa en logros existentes ───────────
LOGRO_NAME_FIX = {
    10: "Kantar (Millward Brown)",
    11: "Kantar (Millward Brown)",
    12: "Kantar (Millward Brown)",
    17: "Independiente (BRICKO / EIDOS)",
    55: "Kantar (Millward Brown)",
    56: "Kantar (Millward Brown)",
    61: "Independiente (BRICKO / EIDOS)",
}


# ─── A + B: IDs de experiencia a borrar ────────────────────────────────────
EXP_TO_DELETE = [7, 12, 13]

# ─── C: Televisa fecha inicio fix ──────────────────────────────────────────
TELEVISA_DATE_FIX = {
    "id": 4,
    "new_fecha_inicio": "04/2018",
}

# ─── B (cont): consolidación de puesto NODO ID 8 ───────────────────────────
NODO_PUESTO_UPDATE = {
    "id": 8,
    "puesto_es": "Investigador Senior · Moderador / Analista · Account Manager",
    "puesto_en": "Senior Researcher · Moderator / Analyst · Account Manager",
}


# ─── G: logros NUEVOS para puestos sin logros ──────────────────────────────
NEW_LOGROS = [
    # (id_emp, empresa, tipo, logro_es, logro_en, impacto, tiene_metrica, metrica, categoria, tags_ind, tags_skill, incl_es, incl_en)
    (27, "Televisa", "Logro",
     "Diseñé e implementé un nuevo proceso de investigación social end-to-end, desde extracción de insights hasta transferencia de conocimiento accionable.",
     "Designed and implemented a new end-to-end social research process, from insight extraction to actionable knowledge transfer.",
     "Alto", "No", "", "Estrategia de Insights", "Media,TV", "Research,Strategy,Process Design", "Sí", "Sí"),
    (27, "Televisa", "Logro",
     "Establecí marco de gestión de proveedores, presupuesto anual y KPIs cualitativos para el área de Investigación Social, elevando su relevancia estratégica en la corporación.",
     "Established vendor management framework, annual budget and qualitative KPIs for the Social Research area, elevating its strategic relevance in the corporation.",
     "Alto", "No", "", "Gestión de Área", "Media,TV", "Leadership,Budget,KPIs,Vendor Mgmt", "Sí", "Sí"),
    (25, "Estudio Segarra", "Logro",
     "Realicé fotografía de producto para marcas reconocidas como Avón, Barcel y Bísquets Obregón, asegurando consistencia visual y entrega oportuna en cada sesión.",
     "Delivered product photography for renowned brands including Avón, Barcel and Bísquets Obregón, ensuring visual consistency and on-time delivery on every shoot.",
     "Medio", "No", "", "Fotografía Comercial", "FMCG,Retail", "Photography,Visual Design", "No", "No"),
    (26, "Question Mark (Marketing Studio)", "Logro",
     "Lideré proyectos de marketing e insights para clientes en diversos segmentos como Project Leader en la startup Question Mark, fundada por Lourdes Lamasney e Ian Reider.",
     "Led marketing and insights projects across diverse client segments as Project Leader at the Question Mark startup, founded by Lourdes Lamasney and Ian Reider.",
     "Medio", "No", "", "Project Leadership", "Marketing,Consulting", "Project Management,Insights,Startup", "Sí", "Sí"),
    (24, "Dirección de Teatro (Independiente)", "Logro",
     "Produje y monté obras de teatro musicales y para público infantil de manera independiente, gestionando todo el ciclo creativo y operativo desde concepto hasta puesta en escena.",
     "Produced and staged independent musical and children's theater plays, managing the full creative and operational cycle from concept to performance.",
     "Bajo", "No", "", "Producción Creativa", "Arts,Theater", "Creative Production,Project Mgmt", "No", "No"),
    (24, "Dirección de Teatro (Independiente)", "Logro",
     "Desarrollé habilidades tempranas de liderazgo creativo y storytelling escénico aplicables hoy a la presentación de insights y narrativas de marca.",
     "Built early skills in creative leadership and stage storytelling, now applied to insights presentation and brand narratives.",
     "Bajo", "No", "", "Soft Skills Origen", "Arts", "Storytelling,Leadership", "No", "No"),
]


# ─── EJECUCIÓN ─────────────────────────────────────────────────────────────
def main():
    print(f"📂 Master: {MASTER}")
    if not BACKUP.exists():
        shutil.copy2(MASTER, BACKUP)
        print(f"🛟 Backup creado: {BACKUP.name}")
    else:
        print(f"🛟 Backup ya existe: {BACKUP.name}")

    wb = load_workbook(str(MASTER))

    # ── 02 Experiencia: localizar índices por ID ──
    ws_exp = wb["02 Experiencia"]
    exp_row_by_id = {}
    for r_idx in range(3, ws_exp.max_row + 1):
        eid = ws_exp.cell(row=r_idx, column=1).value
        if eid is not None:
            exp_row_by_id[int(eid)] = r_idx

    # ── 03 Logros: localizar índices por ID Logro ──
    ws_log = wb["03 Logros"]
    log_row_by_id = {}
    for r_idx in range(3, ws_log.max_row + 1):
        lid = ws_log.cell(row=r_idx, column=1).value
        if lid is not None:
            log_row_by_id[int(lid)] = r_idx

    print(f"\nFound {len(exp_row_by_id)} experiencia rows, {len(log_row_by_id)} logros rows")

    # ─── A+B: reasignar logros (cambiar ID Empresa + Empresa) ────────────
    print(f"\n[A+B] Reasignando {len(LOGRO_EMP_REASSIGN)} logros antes de borrar empresas duplicadas:")
    for lid, (new_eid, new_name) in LOGRO_EMP_REASSIGN.items():
        if lid not in log_row_by_id:
            print(f"  ⚠️  Logro {lid} no encontrado")
            continue
        r = log_row_by_id[lid]
        old_eid = ws_log.cell(row=r, column=2).value
        old_name = ws_log.cell(row=r, column=3).value
        ws_log.cell(row=r, column=2, value=new_eid)
        ws_log.cell(row=r, column=3, value=new_name)
        print(f"  🔀 L{lid}: ID Emp {old_eid} → {new_eid} ({old_name!r} → {new_name!r})")

    # ─── D: normalizar nombres de empresa en logros ──────────────────────
    print(f"\n[D] Normalizando nombres de empresa en {len(LOGRO_NAME_FIX)} logros:")
    for lid, new_name in LOGRO_NAME_FIX.items():
        if lid not in log_row_by_id:
            continue
        r = log_row_by_id[lid]
        old_name = ws_log.cell(row=r, column=3).value
        ws_log.cell(row=r, column=3, value=new_name)
        print(f"  ✏  L{lid}: {old_name!r} → {new_name!r}")

    # ─── E+F: reescritura de logros ──────────────────────────────────────
    print(f"\n[E+F] Reescribiendo {len(LOGRO_REWRITES)} logros con verbos de acción:")
    for lid, (new_es, new_en) in LOGRO_REWRITES.items():
        if lid not in log_row_by_id:
            print(f"  ⚠️  Logro {lid} no encontrado")
            continue
        r = log_row_by_id[lid]
        ws_log.cell(row=r, column=5, value=new_es)  # col E = Logro ES
        ws_log.cell(row=r, column=6, value=new_en)  # col F = Achievement EN
    print(f"  ✅ {len(LOGRO_REWRITES)} reescrituras aplicadas")

    # ─── G: agregar 6 logros nuevos ───────────────────────────────────────
    print(f"\n[G] Agregando {len(NEW_LOGROS)} logros nuevos:")
    # Encontrar el siguiente ID disponible
    max_lid = max(int(ws_log.cell(row=r, column=1).value)
                  for r in range(3, ws_log.max_row + 1)
                  if ws_log.cell(row=r, column=1).value is not None)
    next_lid = max_lid + 1
    for logro_data in NEW_LOGROS:
        row = (next_lid,) + logro_data
        ws_log.append(row)
        print(f"  ➕ L{next_lid} (Emp {logro_data[0]} · {logro_data[1]}): {logro_data[2]}")
        next_lid += 1

    # ─── B (cont): actualizar puesto NODO ID 8 ────────────────────────────
    if NODO_PUESTO_UPDATE["id"] in exp_row_by_id:
        r = exp_row_by_id[NODO_PUESTO_UPDATE["id"]]
        ws_exp.cell(row=r, column=5, value=NODO_PUESTO_UPDATE["puesto_es"])
        ws_exp.cell(row=r, column=6, value=NODO_PUESTO_UPDATE["puesto_en"])
        print(f"\n[B] Actualizado puesto NODO ID 8: '{NODO_PUESTO_UPDATE['puesto_es']}'")

    # ─── C: Televisa fecha inicio ─────────────────────────────────────────
    r = exp_row_by_id[TELEVISA_DATE_FIX["id"]]
    old_fi = ws_exp.cell(row=r, column=9).value
    ws_exp.cell(row=r, column=9, value=TELEVISA_DATE_FIX["new_fecha_inicio"])
    print(f"\n[C] Televisa ID 4 fecha inicio: {old_fi!r} → {TELEVISA_DATE_FIX['new_fecha_inicio']!r}")

    # ─── A+B: borrar filas de experiencia consolidadas ────────────────────
    print(f"\n[A+B] Borrando {len(EXP_TO_DELETE)} entradas de experiencia duplicadas:")
    # Borrar de mayor a menor para no invalidar indices
    for eid in sorted(EXP_TO_DELETE, reverse=True):
        if eid in exp_row_by_id:
            r = exp_row_by_id[eid]
            empresa = ws_exp.cell(row=r, column=2).value
            puesto = ws_exp.cell(row=r, column=5).value
            ws_exp.delete_rows(r)
            print(f"  🗑  ID {eid} ({empresa} · {puesto}) eliminado (fila {r})")

    wb.save(str(MASTER))
    print(f"\n✅ Master guardado")

    # Verificación rápida
    print(f"\n📊 Estado final:")
    wb2 = load_workbook(str(MASTER), data_only=True)
    for s in ["02 Experiencia", "03 Logros"]:
        count = sum(1 for r in wb2[s].iter_rows(min_row=3, values_only=True)
                    if r[0] is not None)
        print(f"   {s}: {count} filas con ID")


if __name__ == "__main__":
    main()
